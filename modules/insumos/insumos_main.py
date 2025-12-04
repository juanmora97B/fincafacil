"""
Módulo de Gestión de Insumos y Semillas
Permite registrar, actualizar y dar seguimiento a las insumos de la finca
Incluye: Edición, Eliminación, Importación Excel, Imágenes, Asignación a trabajadores
"""
import customtkinter as ctk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
import sqlite3
import sys
import os
from PIL import Image
import shutil
import openpyxl
import unicodedata

sys.path.append(os.path.join(os.path.dirname(__file__), '../..'))
from database import db


class InsumosModule(ctk.CTkFrame):
    def __init__(self, master):
        super().__init__(master)
        self.pack(fill="both", expand=True)
        self.foto_path_actual = None  # Ruta de la foto seleccionada
        self.insumo_editando_id = None  # ID de insumo en edición
        self.crear_widgets()
        self.cargar_fincas()
        self.cargar_trabajadores()
        self.cargar_insumos()

    # ----------------------- Helpers de Normalización y Parsing -----------------------
    def _normalize_text(self, texto: str) -> str:
        """Normaliza texto para comparaciones (lowercase, sin acentos, espacios simples)."""
        if texto is None:
            return ""
        t = str(texto).strip().lower()
        # Remover acentos
        t = unicodedata.normalize('NFD', t)
        t = ''.join(c for c in t if unicodedata.category(c) != 'Mn')
        # Colapsar espacios múltiples
        while "  " in t:
            t = t.replace("  ", " ")
        return t

    def _parse_valor(self, valor_raw):
        """Convierte distintos formatos de valor monetario a float seguro.
        Acepta: 170000, 170.000, 170,000, $170000.50, 170000,50
        Retorna None si vacío. Lanza ValueError si imposible convertir.
        """
        if valor_raw is None:
            return None
        s = str(valor_raw).strip()
        if s == "":
            return None
        # Quitar símbolo de moneda y espacios
        s = s.replace('$', '').replace('₡', '').replace('€', '').replace('USD', '').strip()
        # Reemplazar separadores decimales comunes
        # Caso: si hay tanto ',' como '.' decidir último como decimal y quitar el otro
        if ',' in s and '.' in s:
            # Tomar el que aparece más a la derecha como decimal
            last_comma = s.rfind(',')
            last_dot = s.rfind('.')
            if last_comma > last_dot:  # coma es decimal
                s = s.replace('.', '')  # quitar puntos miles
                s = s.replace(',', '.')  # coma decimal a punto
            else:  # punto es decimal
                s = s.replace(',', '')  # quitar comas miles
        else:
            # Solo comas: asumir coma decimal si hay una y parte decimal corta, si más de una -> quitar todas como miles
            if ',' in s and s.count(',') == 1 and len(s.split(',')[-1]) <= 2:
                s = s.replace(',', '.')
            else:
                s = s.replace(',', '')
            # Solo puntos: si múltiples puntos -> quitar como miles excepto último
            if s.count('.') > 1:
                parts = s.split('.')
                s = ''.join(parts[:-1]) + '.' + parts[-1]
        # Quitar espacios finales
        s = s.strip()
        try:
            return float(s)
        except ValueError:
            raise ValueError(f"Formato de valor inválido: '{valor_raw}'")

    def crear_widgets(self):
        # Contenedor principal sin scroll para ocupar todo el espacio
        container = ctk.CTkFrame(self, fg_color="transparent")
        container.pack(fill="both", expand=True, padx=2, pady=5)
        
        # Título
        titulo = ctk.CTkLabel(
            container,
            text="🔧 Gestión de Insumos y Semillas",
            font=("Segoe UI", 22, "bold")
        )
        titulo.pack(pady=(5, 3))

        # Frame principal con tabs
        self.notebook = ttk.Notebook(container)
        self.notebook.pack(fill="both", expand=True, padx=2, pady=(3, 10))

        # Tab: Registrar Insumo
        self.frame_registro = ctk.CTkFrame(self.notebook)
        self.notebook.add(self.frame_registro, text="➕ Nueva Insumo")

        # Tab: Catálogo
        self.frame_catalogo = ctk.CTkFrame(self.notebook)
        self.notebook.add(self.frame_catalogo, text="📋 Catálogo")

        # Tab: Movimientos de Insumos
        self.frame_mantenimientos = ctk.CTkFrame(self.notebook)
        self.notebook.add(self.frame_mantenimientos, text="� Movimientos")

        # Crear contenido
        self.crear_formulario_registro()
        self.crear_catalogo()
        self.crear_mantenimientos()

    def crear_formulario_registro(self):
        """Formulario para registrar nueva insumo"""
        main_frame = ctk.CTkScrollableFrame(self.frame_registro)
        main_frame.pack(fill="both", expand=True, padx=10, pady=5)

        ctk.CTkLabel(
            main_frame,
            text="📝 Nuevo Insumo",
            font=("Segoe UI", 18, "bold")
        ).pack(pady=(0, 5))

        # Información básica
        info_frame = ctk.CTkFrame(main_frame, corner_radius=10)
        info_frame.pack(fill="both", expand=True, pady=5)
        
        ctk.CTkLabel(info_frame, text="📋 INFORMACIÓN BÁSICA", 
                    font=("Segoe UI", 14, "bold")).pack(anchor="w", padx=10, pady=10)

        # Código y Nombre
        row1 = ctk.CTkFrame(info_frame, fg_color="transparent")
        row1.pack(fill="x", padx=10, pady=5)
        ctk.CTkLabel(row1, text="Código *:", width=150).pack(side="left", padx=5)
        self.entry_codigo = ctk.CTkEntry(row1, width=200)
        self.entry_codigo.pack(side="left", padx=5)
        
        ctk.CTkLabel(row1, text="Nombre *:", width=100).pack(side="left", padx=5)
        self.entry_nombre = ctk.CTkEntry(row1, width=300)
        self.entry_nombre.pack(side="left", padx=5)

        # Categoría y Finca
        row2 = ctk.CTkFrame(info_frame, fg_color="transparent")
        row2.pack(fill="x", padx=10, pady=5)
        ctk.CTkLabel(row2, text="Categoría *:", width=150).pack(side="left", padx=5)
        self.combo_categoria = ctk.CTkComboBox(
            row2,
            values=["Fertilizante", "Alimento", "Medicamento", "Semilla", "Vacuna", "Otro"],
            width=200
        )
        self.combo_categoria.pack(side="left", padx=5)
        
        ctk.CTkLabel(row2, text="Finca *:", width=100).pack(side="left", padx=5)
        self.combo_finca = ctk.CTkComboBox(row2, width=300)
        self.combo_finca.pack(side="left", padx=5)

        # Unidad de Medida y Proveedor
        row3 = ctk.CTkFrame(info_frame, fg_color="transparent")
        row3.pack(fill="x", padx=10, pady=5)
        ctk.CTkLabel(row3, text="Unidad de Medida *:", width=150).pack(side="left", padx=5)
        self.entry_unidad_medida = ctk.CTkEntry(row3, width=200, placeholder_text="kg, litros, dosis, etc.")
        self.entry_unidad_medida.pack(side="left", padx=5)
        
        ctk.CTkLabel(row3, text="Proveedor *:", width=100).pack(side="left", padx=5)
        self.entry_proveedor_principal = ctk.CTkEntry(row3, width=300)
        self.entry_proveedor_principal.pack(side="left", padx=5)

        # Cantidad Inicial y Fecha de Adquisición
        row4 = ctk.CTkFrame(info_frame, fg_color="transparent")
        row4.pack(fill="x", padx=10, pady=5)
        ctk.CTkLabel(row4, text="Cantidad Inicial *:", width=150).pack(side="left", padx=5)
        self.entry_stock_actual = ctk.CTkEntry(row4, width=200, placeholder_text="Cantidad disponible")
        self.entry_stock_actual.insert(0, "0")
        self.entry_stock_actual.pack(side="left", padx=5)
        
        ctk.CTkLabel(row4, text="Fecha Adquisición:", width=120).pack(side="left", padx=5)
        self.entry_fecha_adq = ctk.CTkEntry(row4, width=180, placeholder_text="YYYY-MM-DD")
        self.entry_fecha_adq.insert(0, datetime.now().strftime("%Y-%m-%d"))
        self.entry_fecha_adq.pack(side="left", padx=5)

        # Lote y Estado
        row5 = ctk.CTkFrame(info_frame, fg_color="transparent")
        row5.pack(fill="x", padx=10, pady=5)
        ctk.CTkLabel(row5, text="Lote/Serie:", width=150).pack(side="left", padx=5)
        self.entry_serie = ctk.CTkEntry(row5, width=200, placeholder_text="Número de lote")
        self.entry_serie.pack(side="left", padx=5)
        
        ctk.CTkLabel(row5, text="Estado:", width=100).pack(side="left", padx=5)
        self.combo_estado = ctk.CTkComboBox(
            row5,
            values=["Disponible", "Agotado", "En Reposición"],
            width=300
        )
        self.combo_estado.set("Disponible")
        self.combo_estado.pack(side="left", padx=5)

        # Ubicación y Responsable
        row6 = ctk.CTkFrame(info_frame, fg_color="transparent")
        row6.pack(fill="x", padx=10, pady=5)
        ctk.CTkLabel(row6, text="Ubicación:", width=150).pack(side="left", padx=5)
        self.entry_ubicacion = ctk.CTkEntry(row6, width=200, placeholder_text="Bodega, almacén, etc.")
        self.entry_ubicacion.pack(side="left", padx=5)
        
        ctk.CTkLabel(row6, text="Responsable *:", width=100).pack(side="left", padx=5)
        self.combo_responsable = ctk.CTkComboBox(row6, width=300)
        self.combo_responsable.pack(side="left", padx=5)
        
        # Stock en Bodega (campo oculto, se calcula automático)
        self.entry_stock_bodega = ctk.CTkEntry(info_frame, width=0)
        self.entry_stock_bodega.pack_forget()
        
        # Imagen de la insumo
        img_frame = ctk.CTkFrame(info_frame, corner_radius=10)
        img_frame.pack(fill="x", padx=10, pady=10)
        
        ctk.CTkLabel(img_frame, text="📷 Imagen:", width=150).pack(side="left", padx=5)
        ctk.CTkButton(
            img_frame,
            text="📁 Seleccionar Imagen",
            command=self.seleccionar_imagen,
            width=180
        ).pack(side="left", padx=5)
        
        self.label_imagen_nombre = ctk.CTkLabel(img_frame, text="Sin imagen", text_color="gray")
        self.label_imagen_nombre.pack(side="left", padx=10)
        
        ctk.CTkButton(
            img_frame,
            text="👁️ Ver",
            command=self.ver_imagen,
            width=80
        ).pack(side="left", padx=5)
        
        ctk.CTkButton(
            img_frame,
            text="🗑️ Quitar",
            command=self.quitar_imagen,
            width=80,
            fg_color="#D32F2F"
        ).pack(side="left", padx=5)

        # Información financiera
        fin_frame = ctk.CTkFrame(main_frame, corner_radius=10)
        fin_frame.pack(fill="x", pady=10)
        
        ctk.CTkLabel(fin_frame, text="💰 INFORMACIÓN DE COSTOS", 
                    font=("Segoe UI", 14, "bold")).pack(anchor="w", padx=10, pady=10)

        row7 = ctk.CTkFrame(fin_frame, fg_color="transparent")
        row7.pack(fill="x", padx=10, pady=5)
        ctk.CTkLabel(row7, text="Valor Unitario *:", width=150).pack(side="left", padx=5)
        self.entry_valor = ctk.CTkEntry(row7, width=150, placeholder_text="Costo por unidad")
        self.entry_valor.pack(side="left", padx=5)
        
        ctk.CTkLabel(row7, text="Stock Mínimo:", width=120).pack(side="left", padx=5)
        self.entry_vida_util = ctk.CTkEntry(row7, width=100, placeholder_text="Nivel mínimo")
        self.entry_vida_util.insert(0, "10")
        self.entry_vida_util.pack(side="left", padx=5)

        # Descripción y Observaciones
        row8 = ctk.CTkFrame(main_frame, fg_color="transparent")
        row8.pack(fill="both", expand=True, pady=5)
        ctk.CTkLabel(row8, text="Descripción:", width=150).pack(side="left", padx=5, anchor="n")
        self.text_descripcion = ctk.CTkTextbox(row8, width=600, height=100)
        self.text_descripcion.pack(side="left", padx=5, fill="both", expand=True)

        row9 = ctk.CTkFrame(main_frame, fg_color="transparent")
        row9.pack(fill="both", expand=True, pady=5)
        ctk.CTkLabel(row9, text="Observaciones *:", width=150).pack(side="left", padx=5, anchor="n")
        self.text_observaciones = ctk.CTkTextbox(row8, width=600, height=100)
        self.text_observaciones.pack(side="left", padx=5, fill="both", expand=True)

        # Botones
        btn_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        btn_frame.pack(pady=5)
        ctk.CTkButton(
            btn_frame, 
            text="💾 Guardar Insumo", 
            command=self.guardar_insumo,
            fg_color="green",
            hover_color="#006400"
        ).pack(side="left", padx=5)
        ctk.CTkButton(
            btn_frame,
            text="🔄 Limpiar Formulario",
            command=self.limpiar_formulario
        ).pack(side="left", padx=5)

    def crear_catalogo(self):
        """Crea el catálogo de insumos"""
        main_frame = ctk.CTkFrame(self.frame_catalogo)
        main_frame.pack(fill="both", expand=True, padx=10, pady=10)

        ctk.CTkLabel(
            main_frame,
            text="📋 Catálogo de Insumos y Semillas",
            font=("Segoe UI", 18, "bold")
        ).pack(pady=(0, 10))

        # Filtros
        filtro_frame = ctk.CTkFrame(main_frame)
        filtro_frame.pack(fill="x", pady=10)
        
        ctk.CTkLabel(filtro_frame, text="Filtrar por Finca:").pack(side="left", padx=5)
        self.combo_filtro_finca = ctk.CTkComboBox(filtro_frame, width=200)
        self.combo_filtro_finca.pack(side="left", padx=5)
        
        ctk.CTkLabel(filtro_frame, text="Estado:").pack(side="left", padx=5)
        self.combo_filtro_estado = ctk.CTkComboBox(
            filtro_frame,
            values=["Todos", "Disponible", "Agotado", "En Reposición"],
            width=180
        )
        self.combo_filtro_estado.set("Todos")
        self.combo_filtro_estado.pack(side="left", padx=5)
        
        ctk.CTkButton(
            filtro_frame,
            text="🔍 Filtrar",
            command=self.cargar_insumos
        ).pack(side="left", padx=5)

        # Frame para tabla (centrar y permitir barra a la derecha)
        tabla_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        tabla_frame.pack(fill="both", expand=True, pady=(5, 5))

        self.tabla = ttk.Treeview(
            tabla_frame,
            columns=("codigo", "nombre", "categoria", "finca", "stock_actual", "stock_bodega", "responsable", "estado"),
            show="headings",
            height=15
        )
        columnas = [
            ("codigo", "Código", 90),
            ("nombre", "Nombre", 180),
            ("categoria", "Categoría", 120),
            ("finca", "Finca", 130),
            ("stock_actual", "Cantidad Total", 100),
            ("stock_bodega", "En Bodega", 90),
            ("responsable", "Responsable", 140),
            ("estado", "Estado", 120)
        ]
        for col, heading, width in columnas:
            self.tabla.heading(col, text=heading)
            self.tabla.column(col, width=width, anchor="center")
        self.tabla.pack(side="left", fill="both", expand=True)
        scrollbar = ttk.Scrollbar(tabla_frame, orient="vertical", command=self.tabla.yview)
        self.tabla.configure(yscroll=scrollbar.set)
        scrollbar.pack(side="right", fill="y")

        # Barra inferior de acciones (botones alineados abajo)
        acciones_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        acciones_frame.pack(fill="x", pady=(0, 8))

        # Usar grid para control de distribución y permitir expansión
        botones = [
            {
                "texto": "🔍 Ver Detalles",
                "comando": self.ver_detalles_insumo,
                "color": None,
                "hover": None
            },
            {
                "texto": "✏️ Editar Seleccionado",
                "comando": self.editar_insumo,
                "color": "#1976D2",
                "hover": "#1565C0"
            },
            {
                "texto": "🗑️ Eliminar Seleccionado",
                "comando": self.eliminar_insumo,
                "color": "#D32F2F",
                "hover": "#C62828"
            },
            {
                "texto": "📥 Importar Excel",
                "comando": self.importar_excel,
                "color": "#388E3C",
                "hover": "#2E7D32"
            },
            {
                "texto": "📋 Descargar Plantilla",
                "comando": self.descargar_plantilla_excel,
                "color": "#F57C00",
                "hover": "#EF6C00"
            }
        ]

        for idx, btn in enumerate(botones):
            b = ctk.CTkButton(
                acciones_frame,
                text=btn["texto"],
                command=btn["comando"],
                fg_color=btn["color"] if btn["color"] else None,
                hover_color=btn["hover"] if btn["hover"] else None,
                width=160
            )
            b.grid(row=0, column=idx, padx=5, pady=5)
            acciones_frame.grid_columnconfigure(idx, weight=1)

    def crear_mantenimientos(self):
        """Crear el sistema de movimientos de stock"""
        main_frame = ctk.CTkFrame(self.frame_mantenimientos)
        main_frame.pack(fill="both", expand=True, padx=10, pady=10)

        ctk.CTkLabel(
            main_frame,
            text="� Movimientos de Stock de Insumos",
            font=("Segoe UI", 18, "bold")
        ).pack(pady=(0, 10))

        # Formulario de movimiento con scroll
        form_scroll = ctk.CTkScrollableFrame(main_frame, height=280, corner_radius=10)
        form_scroll.pack(fill="x", expand=False, pady=(0, 10))
        
        form_frame = ctk.CTkFrame(form_scroll, fg_color="transparent")
        form_frame.pack(fill="both", padx=5, pady=5)

        row1 = ctk.CTkFrame(form_frame, fg_color="transparent")
        row1.pack(fill="x", padx=10, pady=10)
        
        ctk.CTkLabel(row1, text="Insumo *:", width=150).pack(side="left", padx=5)
        self.combo_insumo_mant = ctk.CTkComboBox(row1, width=350)
        self.combo_insumo_mant.pack(side="left", padx=5)
        
        ctk.CTkButton(
            row1,
            text="🔄 Actualizar Lista",
            command=self.cargar_insumos_para_mantenimiento,
            width=150
        ).pack(side="left", padx=5)

        row2 = ctk.CTkFrame(form_frame, fg_color="transparent")
        row2.pack(fill="x", padx=10, pady=5)
        
        ctk.CTkLabel(row2, text="Tipo de Movimiento *:", width=150).pack(side="left", padx=5)
        self.combo_tipo_mant = ctk.CTkComboBox(
            row2,
            values=["Entrada", "Salida", "Ajuste"],
            width=200
        )
        self.combo_tipo_mant.set("Entrada")
        self.combo_tipo_mant.pack(side="left", padx=5)
        
        ctk.CTkLabel(row2, text="Fecha *:", width=80).pack(side="left", padx=5)
        self.entry_fecha_mant = ctk.CTkEntry(row2, width=150, placeholder_text="YYYY-MM-DD")
        self.entry_fecha_mant.insert(0, datetime.now().strftime("%Y-%m-%d"))
        self.entry_fecha_mant.pack(side="left", padx=5)

        row3 = ctk.CTkFrame(form_frame, fg_color="transparent")
        row3.pack(fill="x", padx=10, pady=5)
        
        ctk.CTkLabel(row3, text="Cantidad *:", width=150).pack(side="left", padx=5)
        self.entry_costo_mant = ctk.CTkEntry(row3, width=150, placeholder_text="Cantidad del movimiento")
        self.entry_costo_mant.pack(side="left", padx=5)
        
        ctk.CTkLabel(row3, text="Responsable:", width=120).pack(side="left", padx=5)
        self.entry_realizado_por = ctk.CTkEntry(row3, width=250, placeholder_text="Quien realizó el movimiento")
        self.entry_realizado_por.pack(side="left", padx=5)
        
        # Row 3B - Campos de precio (solo para Entrada/Compra)
        self.row3b_precios = ctk.CTkFrame(form_frame, fg_color="transparent")
        self.row3b_precios.pack(fill="x", padx=10, pady=5)
        
        ctk.CTkLabel(self.row3b_precios, text="Precio Unitario:", width=150).pack(side="left", padx=5)
        self.entry_precio_unitario_mov = ctk.CTkEntry(self.row3b_precios, width=150, placeholder_text="Precio por unidad")
        self.entry_precio_unitario_mov.pack(side="left", padx=5)
        
        ctk.CTkLabel(self.row3b_precios, text="Precio Total:", width=120).pack(side="left", padx=5)
        self.entry_precio_total_mov = ctk.CTkEntry(self.row3b_precios, width=150, placeholder_text="Calculado automático")
        self.entry_precio_total_mov.configure(state="readonly")
        self.entry_precio_total_mov.pack(side="left", padx=5)
        
        # Vincular eventos para cálculo automático de precio total
        self.entry_costo_mant.bind("<KeyRelease>", self._calcular_precio_total_mov)
        self.entry_precio_unitario_mov.bind("<KeyRelease>", self._calcular_precio_total_mov)
        
        # Vincular evento de cambio de tipo de movimiento
        self.combo_tipo_mant.configure(command=self._actualizar_campos_precio)
        
        # Ocultar por defecto (se muestra solo si es Entrada)
        self.row3b_precios.pack_forget()

        row4 = ctk.CTkFrame(form_frame, fg_color="transparent")
        row4.pack(fill="x", padx=10, pady=5)
        
        ctk.CTkLabel(row4, text="Finca Destino:", width=150).pack(side="left", padx=5)
        self.entry_proveedor_mant = ctk.CTkEntry(row4, width=250, placeholder_text="Para salidas/consumos")
        self.entry_proveedor_mant.pack(side="left", padx=5)
        
        ctk.CTkLabel(row4, text="Motivo:", width=100).pack(side="left", padx=5)
        self.entry_proximo_mant = ctk.CTkEntry(row4, width=250, placeholder_text="Compra, consumo, pérdida, etc.")
        self.entry_proximo_mant.pack(side="left", padx=5)

        row5 = ctk.CTkFrame(form_frame, fg_color="transparent")
        row5.pack(fill="x", padx=10, pady=5)
        
        ctk.CTkLabel(row5, text="Observaciones:", width=150).pack(side="left", padx=5, anchor="n")
        self.text_desc_mant = ctk.CTkTextbox(row5, width=500, height=100)
        self.text_desc_mant.pack(side="left", padx=5, fill="x", expand=True)

        # Frame de botones - EMPAQUETADO PRIMERO con side="bottom" para que se quede abajo
        btn_bottom_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        btn_bottom_frame.pack(side="bottom", fill="x", pady=(5, 10))
        
        # Columna izquierda: Acciones de movimiento
        btn_left = ctk.CTkFrame(btn_bottom_frame, fg_color="transparent")
        btn_left.pack(side="left", padx=5)
        
        ctk.CTkButton(
            btn_left,
            text="💾 Registrar Movimiento",
            command=self.guardar_mantenimiento,
            fg_color="green",
            width=200
        ).pack(side="left", padx=3)
        
        ctk.CTkButton(
            btn_left,
            text="🗑️ Eliminar del Historial",
            command=self.completar_mantenimiento,
            fg_color="#dc3545",
            hover_color="#c82333",
            width=200
        ).pack(side="left", padx=3)
        
        ctk.CTkButton(
            btn_left,
            text="� Actualizar Vista",
            command=self.cargar_mantenimientos,
            fg_color="#17a2b8",
            hover_color="#138496",
            width=150
        ).pack(side="left", padx=3)
        
        # Columna derecha: Acciones de insumo
        btn_right = ctk.CTkFrame(btn_bottom_frame, fg_color="transparent")
        btn_right.pack(side="right", padx=5)
        
        ctk.CTkButton(
            btn_right,
            text="📊 Ver Catálogo",
            command=lambda: self.notebook.select(1),
            fg_color="#6c757d",
            width=180
        ).pack(side="left", padx=3)
        
        ctk.CTkButton(
            btn_right,
            text="🗑️ Limpiar Campos",
            command=self.eliminar_insumo_desde_mantenimiento,
            fg_color="#dc3545",
            width=180
        ).pack(side="left", padx=3)

        # Historial de movimientos - EMPAQUETADO DESPUÉS de los botones
        hist_frame = ctk.CTkFrame(main_frame)
        hist_frame.pack(fill="both", expand=True, pady=(0, 5))
        
        ctk.CTkLabel(
            hist_frame,
            text="📅 Historial de Movimientos de Insumos",
            font=("Segoe UI", 14, "bold")
        ).pack(pady=5)

        # Frame contenedor para tabla
        tabla_container = ctk.CTkFrame(hist_frame, fg_color="transparent")
        tabla_container.pack(fill="both", expand=True)
        
        self.tabla_mant = ttk.Treeview(
            tabla_container,
            columns=("id", "fecha", "insumo", "tipo", "cantidad", "precio", "realizado_por", "finca"),
            show="headings",
            height=6
        )
        
        columnas_mant = [
            ("id", "ID", 50),
            ("fecha", "Fecha", 90),
            ("insumo", "Insumo", 220),
            ("tipo", "Tipo", 110),
            ("cantidad", "Cantidad", 100),
            ("precio", "Precio", 90),
            ("realizado_por", "Realizado Por", 140),
            ("finca", "Finca Destino", 120)
        ]
        
        for col, heading, width in columnas_mant:
            self.tabla_mant.heading(col, text=heading)
            self.tabla_mant.column(col, width=width, anchor="center")

        self.tabla_mant.pack(side="left", fill="both", expand=True)
        
        scrollbar_mant = ttk.Scrollbar(tabla_container, orient="vertical", command=self.tabla_mant.yview)
        self.tabla_mant.configure(yscroll=scrollbar_mant.set)
        scrollbar_mant.pack(side="right", fill="y")
        
        self.cargar_mantenimientos()

    def cargar_fincas(self):
        """Carga las fincas disponibles"""
        try:
            with db.get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT id, nombre FROM finca WHERE estado = 'Activa' OR estado = 'Activo'")
                fincas = [f"{row[0]}-{row[1]}" for row in cursor.fetchall()]
                
                fincas_con_todas = ["Todas"] + fincas
                
                if hasattr(self, 'combo_finca'):
                    self.combo_finca.configure(values=fincas)
                    if fincas:
                        self.combo_finca.set(fincas[0])
                
                if hasattr(self, 'combo_filtro_finca'):
                    self.combo_filtro_finca.configure(values=fincas_con_todas)
                    self.combo_filtro_finca.set("Todas")
                    
        except Exception as e:
            messagebox.showerror("Error", f"No se pudieron cargar las fincas:\n{e}")
    
    def cargar_trabajadores(self):
        """Carga los empleados activos para asignar como responsables"""
        try:
            with db.get_connection() as conn:
                cursor = conn.cursor()
                # Verificar primero si la tabla empleado existe
                cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='empleado'")
                if not cursor.fetchone():
                    # Tabla no existe, usar solo Bodega
                    if hasattr(self, 'combo_responsable'):
                        self.combo_responsable.configure(values=["Bodega"])
                        self.combo_responsable.set("Bodega")
                    return
                
                # Usar tabla empleado (de nómina) en lugar de trabajador
                cursor.execute("""
                    SELECT rowid, nombres || ' ' || apellidos as nombre_completo, cargo 
                    FROM empleado 
                    WHERE estado_actual = 'Activo' OR estado_actual IS NULL
                    ORDER BY nombres
                """)
                trabajadores = [f"{row[0]}-{row[1]} ({row[2] or 'Sin cargo'})" for row in cursor.fetchall()]
                
                # Agregar opción "Bodega" al inicio
                opciones = ["Bodega"] + trabajadores
                
                if hasattr(self, 'combo_responsable'):
                    self.combo_responsable.configure(values=opciones)
                    self.combo_responsable.set("Bodega")
                    
        except Exception:
            # Si ocurre cualquier error, usar solo "Bodega"
            if hasattr(self, 'combo_responsable'):
                self.combo_responsable.configure(values=["Bodega"])
                self.combo_responsable.set("Bodega")
    
    def seleccionar_imagen(self):
        """Selecciona una imagen para la insumo"""
        archivo = filedialog.askopenfilename(
            title="Seleccionar imagen de la insumo",
            filetypes=[
                ("Imágenes", "*.png *.jpg *.jpeg *.gif *.bmp"),
                ("Todos los archivos", "*.*")
            ]
        )
        if archivo:
            self.foto_path_actual = archivo
            nombre_archivo = os.path.basename(archivo)
            self.label_imagen_nombre.configure(text=nombre_archivo, text_color="green")
    
    def ver_imagen(self):
        """Muestra la imagen seleccionada"""
        if not self.foto_path_actual:
            messagebox.showinfo("Info", "No hay imagen seleccionada")
            return
        
        try:
            ventana = ctk.CTkToplevel(self)
            ventana.title("Vista previa de imagen")
            ventana.geometry("600x600")
            
            img = Image.open(self.foto_path_actual)
            # Redimensionar manteniendo aspecto
            img.thumbnail((580, 580), Image.Resampling.LANCZOS)
            
            ctk_img = ctk.CTkImage(light_image=img, dark_image=img, size=img.size)
            label = ctk.CTkLabel(ventana, image=ctk_img, text="")
            label.pack(padx=10, pady=10)
            
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo abrir la imagen:\n{e}")
    
    def quitar_imagen(self):
        """Quita la imagen seleccionada"""
        self.foto_path_actual = None
        self.label_imagen_nombre.configure(text="Sin imagen", text_color="gray")

    def guardar_insumo(self):
        """Guarda una nueva insumo o actualiza una existente"""
        # Validar campos obligatorios
        codigo = self.entry_codigo.get().strip()
        nombre = self.entry_nombre.get().strip()
        categoria = self.combo_categoria.get()
        unidad_medida = self.entry_unidad_medida.get().strip()
        proveedor = self.entry_proveedor_principal.get().strip()
        cantidad_inicial = self.entry_stock_actual.get().strip()
        valor_unitario = self.entry_valor.get().strip()
        observaciones = self.text_observaciones.get("1.0", "end-1c").strip()
        
        # Validar finca obligatoria
        if not self.combo_finca.get():
            messagebox.showwarning("Atención", "Debe seleccionar una Finca")
            return
            
        # Validar responsable obligatorio
        if not self.combo_responsable.get():
            messagebox.showwarning("Atención", "Debe asignar un Responsable")
            return
        
        if not codigo or not nombre or not categoria or not unidad_medida or not proveedor or not cantidad_inicial or not valor_unitario or not observaciones:
            messagebox.showwarning("Atención", 
                "Complete todos los campos obligatorios:\n" +
                "• Código\n• Nombre\n• Categoría\n• Unidad de Medida\n• Proveedor\n• Cantidad Inicial\n• Valor Unitario\n• Finca\n• Responsable\n• Observaciones")
            return
        
        try:
            # Obtener id_finca
            id_finca = int(self.combo_finca.get().split("-")[0])
            
            # Validar y parsear valor unitario
            try:
                valor = self._parse_valor(valor_unitario)
                if valor is None:
                    messagebox.showerror("Error", "El Valor Unitario no puede estar vacío")
                    return
                if valor < 0:
                    messagebox.showerror("Error", f"El Valor Unitario no puede ser negativo. Valor: {valor}")
                    return
            except ValueError as e:
                messagebox.showerror("Error", str(e))
                return

            # Validar cantidad inicial como entero positivo (pero permitir formato decimal y truncar)
            try:
                cantidad_norm = cantidad_inicial.strip().replace(',', '.')
                if cantidad_norm == "":
                    messagebox.showerror("Error", "La Cantidad Inicial no puede estar vacía")
                    return
                cantidad_float = float(cantidad_norm)
                if cantidad_float <= 0:
                    messagebox.showerror("Error", f"La Cantidad Inicial debe ser mayor a 0. Valor: {cantidad_float}")
                    return
                # Convertir a entero lógico (stock inicial sin fracciones)
                cantidad_inicial_val = int(cantidad_float)
            except ValueError:
                messagebox.showerror("Error", f"La Cantidad Inicial debe ser numérica. Valor ingresado: '{cantidad_inicial}'")
                return
            # Stock mínimo (opcional)
            vida_util = None
            if self.entry_vida_util.get().strip():
                try:
                    # Normalizar entrada: convertir comas a puntos y usar float
                    stock_min_txt = self.entry_vida_util.get().strip().replace(",", ".")
                    vida_util = float(stock_min_txt)
                    if vida_util < 0:
                        messagebox.showerror("Error", "El stock mínimo no puede ser negativo")
                        return
                except ValueError:
                    messagebox.showerror("Error", f"El stock mínimo debe ser un número válido.\nValor ingresado: '{self.entry_vida_util.get()}'")
                    return
            
            # Obtener id_trabajador del combo responsable (usa rowid de empleado)
            id_trabajador = None
            responsable_text = None
            responsable_sel = self.combo_responsable.get()
            if responsable_sel and responsable_sel != "Bodega":
                try:
                    id_trabajador = int(responsable_sel.split("-")[0])
                    # Obtener nombre del empleado para campo responsable
                    with db.get_connection() as conn:
                        cursor = conn.cursor()
                        cursor.execute("SELECT nombres || ' ' || apellidos FROM empleado WHERE rowid = ?", (id_trabajador,))
                        row = cursor.fetchone()
                        if row:
                            responsable_text = row[0]
                except:
                    responsable_text = responsable_sel
            elif responsable_sel == "Bodega":
                responsable_text = "Bodega"
            
            # Copiar imagen a carpeta uploads/insumos si se seleccionó una
            foto_final_path = None
            if self.foto_path_actual:
                try:
                    uploads_dir = os.path.join(os.path.dirname(__file__), '../../uploads/insumos')
                    os.makedirs(uploads_dir, exist_ok=True)
                    
                    # Generar nombre único
                    ext = os.path.splitext(self.foto_path_actual)[1]
                    nombre_archivo = f"herr_{codigo}_{datetime.now().strftime('%Y%m%d%H%M%S')}{ext}"
                    foto_final_path = os.path.join(uploads_dir, nombre_archivo)
                    
                    shutil.copy2(self.foto_path_actual, foto_final_path)
                except Exception as e:
                    messagebox.showwarning("Advertencia", f"No se pudo copiar la imagen:\n{e}")
            
            # Validaciones de stock
            stock_actual_input = self.entry_stock_actual.get().strip()
            stock_actual = self._get_stock_actual_validado(stock_actual_input)
            # Si estamos en modo creación y se ingresó cantidad inicial decimal, usar cantidad_inicial_val parseada
            if not self.insumo_editando_id and 'cantidad_inicial_val' in locals():
                # Priorizar cantidad inicial explícita sobre stock_actual calculado
                stock_actual = float(cantidad_inicial_val)
            # Al crear un insumo nuevo, stock_bodega = stock_actual (todo está en bodega inicialmente)
            stock_bodega = stock_actual

            with db.get_connection() as conn:
                cursor = conn.cursor()
                
                if self.insumo_editando_id:
                    # Modo edición
                    cursor.execute("""
                        UPDATE insumo SET
                            codigo = ?, nombre = ?, categoria = ?, descripcion = ?,
                            proveedor_principal = ?, unidad_medida = ?, lote_proveedor = ?, id_finca = ?,
                            ubicacion = ?, estado = ?, fecha_adquisicion = ?,
                            precio_unitario = ?, stock_minimo = ?, responsable = ?,
                            observaciones = ?, id_trabajador = ?, foto_path = ?, stock_actual = ?, stock_bodega = ?
                        WHERE id = ?
                    """, (
                        codigo, nombre, categoria,
                        self.text_descripcion.get("1.0", "end-1c").strip() or None,
                        self.entry_proveedor_principal.get().strip() or None,
                        self.entry_unidad_medida.get().strip() or None,
                        self.entry_serie.get().strip() or None,
                        id_finca,
                        self.entry_ubicacion.get().strip() or None,
                        self.combo_estado.get(),
                        self.entry_fecha_adq.get().strip() or None,
                        valor,
                        vida_util,
                        responsable_text,
                        self.text_observaciones.get("1.0", "end-1c").strip() or None,
                        id_trabajador,
                        foto_final_path or self.foto_path_actual,
                        stock_actual, stock_bodega,
                        self.insumo_editando_id
                    ))
                    messagebox.showinfo("Éxito", "✅ Insumo actualizada correctamente")
                else:
                    # Modo creación
                    cursor.execute("""
                        INSERT INTO insumo (
                            codigo, nombre, categoria, descripcion, proveedor_principal, unidad_medida, lote_proveedor,
                            id_finca, ubicacion, estado, fecha_adquisicion, precio_unitario,
                            stock_minimo, responsable, observaciones, id_trabajador, foto_path,
                            stock_actual, stock_bodega
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (
                        codigo, nombre, categoria,
                        self.text_descripcion.get("1.0", "end-1c").strip() or None,
                        self.entry_proveedor_principal.get().strip() or None,
                        self.entry_unidad_medida.get().strip() or None,
                        self.entry_serie.get().strip() or None,
                        id_finca,
                        self.entry_ubicacion.get().strip() or None,
                        self.combo_estado.get(),
                        self.entry_fecha_adq.get().strip() or None,
                        valor,
                        vida_util,
                        responsable_text,
                        self.text_observaciones.get("1.0", "end-1c").strip() or None,
                        id_trabajador,
                        foto_final_path,
                        stock_actual, stock_bodega
                    ))
                    messagebox.showinfo("Éxito", "✅ Insumo registrada correctamente")
                
                conn.commit()
            
            self.limpiar_formulario()
            self.cargar_insumos()
            self.notebook.select(1)  # Cambiar a pestaña de catálogo
            
        except sqlite3.IntegrityError:
            messagebox.showerror("Error", f"Ya existe una insumo con el código '{codigo}'")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo guardar la insumo:\n{e}")

    def limpiar_formulario(self):
        """Limpia el formulario"""
        self.entry_codigo.delete(0, "end")
        self.entry_nombre.delete(0, "end")
        self.entry_proveedor_principal.delete(0, "end")
        self.entry_unidad_medida.delete(0, "end")
        self.entry_serie.delete(0, "end")
        self.entry_ubicacion.delete(0, "end")
        self.entry_fecha_adq.delete(0, "end")
        self.entry_fecha_adq.insert(0, datetime.now().strftime("%Y-%m-%d"))
        self.entry_valor.delete(0, "end")
        self.entry_vida_util.delete(0, "end")
        if hasattr(self, 'entry_stock_actual'):
            self.entry_stock_actual.delete(0, "end")
            self.entry_stock_actual.insert(0, "1")
        if hasattr(self, 'entry_stock_bodega'):
            self.entry_stock_bodega.delete(0, "end")
            self.entry_stock_bodega.insert(0, "1")
        self.text_descripcion.delete("1.0", "end")
        self.text_observaciones.delete("1.0", "end")
        self.combo_categoria.set("Medicamento")
        self.combo_estado.set("Operativa")
        self.combo_responsable.set("Bodega")
        self.foto_path_actual = None
        self.label_imagen_nombre.configure(text="Sin imagen", text_color="gray")
        self.insumo_editando_id = None

    def cargar_insumos(self):
        """Carga insumos en Catálogo con filtrado y muestra stock_actual y stock_bodega"""
        for item in self.tabla.get_children():
            self.tabla.delete(item)

        try:
            with db.get_connection() as conn:
                cursor = conn.cursor()
                query = """
                    SELECT h.codigo, h.nombre, h.categoria, f.nombre as finca_nombre,
                           h.stock_actual, h.id_trabajador, h.responsable, h.stock_bodega,
                           h.estado, h.id_finca
                    FROM insumo h
                    LEFT JOIN finca f ON h.id_finca = f.id
                    WHERE 1=1
                """
                params = []

                # Filtro por finca
                if hasattr(self, 'combo_filtro_finca') and self.combo_filtro_finca.get() != "Todas":
                    try:
                        id_finca = int(self.combo_filtro_finca.get().split("-")[0])
                        query += " AND h.id_finca = ?"
                        params.append(id_finca)
                    except Exception:
                        pass

                # Filtro por estado
                if hasattr(self, 'combo_filtro_estado') and self.combo_filtro_estado.get() != "Todos":
                    query += " AND h.estado = ?"
                    params.append(self.combo_filtro_estado.get())

                query += " ORDER BY h.codigo"
                cursor.execute(query, params)

                for row in cursor.fetchall():
                    # row indices after join: 0 codigo,1 nombre,2 categoria,3 finca_nombre,4 stock_actual,5 id_trabajador,
                    # 6 responsable,7 stock_bodega,8 estado,9 id_finca
                    stock_actual = row[4] if row[4] is not None else 0
                    stock_bodega = row[7] if row[7] is not None else 0
                    responsable = row[6] or "Sin asignar"
                    finca_nombre = row[3] or 'Sin Finca'
                    
                    self.tabla.insert("", "end", values=(
                        row[0], row[1], row[2] or 'N/A', finca_nombre, stock_actual, stock_bodega, responsable, row[8]
                    ))
        except Exception as e:
            messagebox.showerror("Error", f"No se pudieron cargar las insumos:\n{e}")

    def ver_detalles_insumo(self):
        """Muestra los detalles de la insumo seleccionada con su foto"""
        seleccionado = self.tabla.selection()
        if not seleccionado:
            messagebox.showwarning("Atención", "Seleccione una insumo")
            return
        
        codigo = self.tabla.item(seleccionado[0])["values"][0]
        
        try:
            with db.get_connection() as conn:
                conn.row_factory = sqlite3.Row
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT h.*, f.nombre as finca_nombre
                    FROM insumo h
                    LEFT JOIN finca f ON h.id_finca = f.id
                    WHERE h.codigo = ?
                """, (codigo,))
                
                row = cursor.fetchone()
                if row:
                    h = dict(row)
                    detalles = f"""
📋 DETALLES DE LA INSUMO

🔧 Código: {h['codigo']}
📝 Nombre: {h['nombre']}
📦 Categoría: {h.get('categoria') or 'N/A'}
📋 Descripción: {h.get('descripcion') or 'No hay descripción'}

🏷️ IDENTIFICACIÓN
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
• Marca: {h.get('proveedor_principal') or 'N/A'}
• Modelo: {h.get('unidad_medida') or 'N/A'}
• N° Serie: {h.get('lote_proveedor') or 'N/A'}

🏞️ UBICACIÓN, ASIGNACIÓN Y STOCK
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
• Finca: {h.get('finca_nombre') or 'N/A'}
• Ubicación: {h.get('ubicacion') or 'N/A'}
• Responsable: {h.get('responsable') or 'No asignado'}
• Estado: {h.get('estado', 'N/A')}
• Stock Actual: {h.get('stock_actual', 'N/D')}
• Stock en Bodega: {h.get('stock_bodega', 'N/D')}

💰 INFORMACIÓN FINANCIERA
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
• Fecha Adquisición: {h.get('fecha_adquisicion') or 'N/A'}
• Valor: ${h.get('precio_unitario', 0):,.2f}
• Stock Mínimo: {h.get('stock_minimo') or 'N/A'}

📝 OBSERVACIONES
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
{h.get('observaciones') or 'No hay observaciones'}
                    """
                    
                    ventana = ctk.CTkToplevel(self)
                    ventana.title(f"Detalles - {h['nombre']}")
                    ventana.geometry("900x700")
                    
                    # Frame principal con dos columnas
                    main_frame = ctk.CTkFrame(ventana)
                    main_frame.pack(fill="both", expand=True, padx=10, pady=10)
                    
                    # Columna izquierda: Detalles de texto
                    text_frame = ctk.CTkFrame(main_frame)
                    text_frame.pack(side="left", fill="both", expand=True, padx=(0, 5))
                    
                    text = ctk.CTkTextbox(text_frame, width=500, height=680)
                    text.pack(padx=5, pady=5, fill="both", expand=True)
                    text.insert("1.0", detalles)
                    text.configure(state="disabled")
                    
                    # Columna derecha: Foto
                    foto_frame = ctk.CTkFrame(main_frame, width=350)
                    foto_frame.pack(side="right", fill="both", padx=(5, 0))
                    foto_frame.pack_propagate(False)
                    
                    # Título de la sección de foto
                    titulo_foto = ctk.CTkLabel(foto_frame, text="📷 FOTOGRAFÍA", 
                                              font=ctk.CTkFont(size=14, weight="bold"))
                    titulo_foto.pack(pady=(10, 5))
                    
                    # Cargar y mostrar la foto
                    foto_path = h.get('foto_path')
                    if foto_path and os.path.exists(foto_path):
                        try:
                            # Cargar imagen
                            img = Image.open(foto_path)
                            
                            # Calcular tamaño manteniendo aspecto (máximo 320x500)
                            max_width = 320
                            max_height = 500
                            img.thumbnail((max_width, max_height), Image.Resampling.LANCZOS)
                            
                            # Crear imagen para CTk
                            ctk_img = ctk.CTkImage(light_image=img, dark_image=img, 
                                                  size=(img.width, img.height))
                            
                            # Mostrar imagen
                            label_img = ctk.CTkLabel(foto_frame, image=ctk_img, text="")
                            label_img.pack(pady=10)
                            
                            # Botón para ver en tamaño completo
                            btn_ampliar = ctk.CTkButton(foto_frame, text="🔍 Ver tamaño completo",
                                                       command=lambda: self._mostrar_imagen_completa(foto_path, h['nombre']))
                            btn_ampliar.pack(pady=5)
                            
                        except Exception as e:
                            # Error al cargar la imagen
                            label_error = ctk.CTkLabel(foto_frame, 
                                                      text="❌ Error al cargar imagen",
                                                      font=ctk.CTkFont(size=12))
                            label_error.pack(pady=20)
                            
                            label_detalle = ctk.CTkLabel(foto_frame, 
                                                        text=f"{str(e)[:50]}...",
                                                        font=ctk.CTkFont(size=10),
                                                        text_color="gray")
                            label_detalle.pack(pady=5)
                    else:
                        # No hay imagen disponible
                        label_sin_img = ctk.CTkLabel(foto_frame, 
                                                    text="📷",
                                                    font=ctk.CTkFont(size=80))
                        label_sin_img.pack(pady=50)
                        
                        label_texto = ctk.CTkLabel(foto_frame, 
                                                  text="Sin imagen disponible",
                                                  font=ctk.CTkFont(size=14),
                                                  text_color="gray")
                        label_texto.pack(pady=10)
                    
        except Exception as e:
            messagebox.showerror("Error", f"No se pudieron cargar los detalles:\n{e}")
    
    def _mostrar_imagen_completa(self, foto_path, nombre_insumo):
        """Muestra la imagen en una ventana separada a tamaño completo"""
        try:
            ventana_img = ctk.CTkToplevel(self)
            ventana_img.title(f"Imagen - {nombre_insumo}")
            ventana_img.geometry("800x800")
            
            # Cargar imagen
            img = Image.open(foto_path)
            
            # Redimensionar si es muy grande (máximo 780x780)
            img.thumbnail((780, 780), Image.Resampling.LANCZOS)
            
            # Crear imagen para CTk
            ctk_img = ctk.CTkImage(light_image=img, dark_image=img, 
                                  size=(img.width, img.height))
            
            # Frame scrollable por si la imagen es muy grande
            scroll_frame = ctk.CTkScrollableFrame(ventana_img)
            scroll_frame.pack(fill="both", expand=True, padx=10, pady=10)
            
            # Mostrar imagen
            label_img = ctk.CTkLabel(scroll_frame, image=ctk_img, text="")
            label_img.pack(pady=10)
            
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo abrir la imagen:\n{e}")

    def editar_insumo(self):
        """Edita la insumo seleccionada en el catálogo"""
        seleccionado = self.tabla.selection()
        if not seleccionado:
            messagebox.showwarning("Atención", "Seleccione una insumo para editar")
            return
        codigo = self.tabla.item(seleccionado[0])["values"][0]
        try:
            with db.get_connection() as conn:
                conn.row_factory = sqlite3.Row
                cursor = conn.cursor()
                cursor.execute("SELECT * FROM insumo WHERE codigo = ?", (codigo,))
                row = cursor.fetchone()
                if not row:
                    messagebox.showerror("Error", "Insumo no encontrada")
                    return
                h = dict(row)
                self._cargar_insumo_en_form(h, cursor)
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo cargar la insumo:\n{e}")

    def _cargar_insumo_en_form(self, h: dict, cursor):
        """Carga los datos de una insumo (dict) en el formulario de registro/edición."""
        # Identificador para edición
        self.insumo_editando_id = h['id']
        # Campos básicos
        self.entry_codigo.delete(0, "end")
        self.entry_codigo.insert(0, h['codigo'])
        self.entry_nombre.delete(0, "end")
        self.entry_nombre.insert(0, h['nombre'])
        self.combo_categoria.set(h.get('categoria') or "Medicamento")
        # Descripción
        self.text_descripcion.delete("1.0", "end")
        if h.get('descripcion'):
            self.text_descripcion.insert("1.0", h['descripcion'])
        # Marca/Modelo/Serie
        self.entry_proveedor_principal.delete(0, "end")
        if h.get('proveedor_principal'):
            self.entry_proveedor_principal.insert(0, h['proveedor_principal'])
        self.entry_unidad_medida.delete(0, "end")
        if h.get('unidad_medida'):
            self.entry_unidad_medida.insert(0, h['unidad_medida'])
        self.entry_serie.delete(0, "end")
        if h.get('lote_proveedor'):
            self.entry_serie.insert(0, h['lote_proveedor'])
        # Finca
        if h.get('id_finca'):
            cursor.execute("SELECT id, nombre FROM finca WHERE id = ?", (h['id_finca'],))
            finca = cursor.fetchone()
            if finca:
                try:
                    self.combo_finca.set(f"{finca['id']}-{finca['nombre']}")
                except Exception:
                    self.combo_finca.set(f"{finca[0]}-{finca[1]}")
        # Ubicación y estado
        self.entry_ubicacion.delete(0, "end")
        if h.get('ubicacion'):
            self.entry_ubicacion.insert(0, h['ubicacion'])
        self.combo_estado.set(h.get('estado') or "Operativa")
        # Fechas y valores
        self.entry_fecha_adq.delete(0, "end")
        if h.get('fecha_adquisicion'):
            self.entry_fecha_adq.insert(0, h['fecha_adquisicion'])
        self.entry_valor.delete(0, "end")
        if h.get('precio_unitario'):
            self.entry_valor.insert(0, str(h['precio_unitario']))
        self.entry_vida_util.delete(0, "end")
        if h.get('stock_minimo'):
            self.entry_vida_util.insert(0, str(h['stock_minimo']))
        # Responsable (priorizar id_trabajador)
        if h.get('id_trabajador'):
            cursor.execute("SELECT rowid, nombres || ' ' || apellidos, cargo FROM empleado WHERE rowid = ?", (h['id_trabajador'],))
            trab = cursor.fetchone()
            if trab:
                try:
                    self.combo_responsable.set(f"{trab['rowid']}-{trab['nombres || ' ' || apellidos']} ({trab['cargo'] or 'Sin cargo'})")
                except Exception:
                    self.combo_responsable.set(f"{trab[0]}-{trab[1]} ({trab[2] or 'Sin cargo'})")
        elif h.get('responsable'):
            if h['responsable'] == "Bodega":
                self.combo_responsable.set("Bodega")
            else:
                self.combo_responsable.set(h['responsable'])
        # Observaciones
        self.text_observaciones.delete("1.0", "end")
        if h.get('observaciones'):
            self.text_observaciones.insert("1.0", h['observaciones'])
        # Stock
        if hasattr(self, 'entry_stock_actual'):
            self.entry_stock_actual.delete(0, "end")
            self.entry_stock_actual.insert(0, str(h.get('stock_actual', 1)))
        if hasattr(self, 'entry_stock_bodega'):
            self.entry_stock_bodega.delete(0, "end")
            valor_bodega = h.get('stock_bodega', 0 if h.get('id_trabajador') else 1)
            self.entry_stock_bodega.insert(0, str(valor_bodega))
        # Foto
        if h.get('foto_path'):
            self.foto_path_actual = h['foto_path']
            try:
                self.label_imagen_nombre.configure(
                    text=os.path.basename(h['foto_path']),
                    text_color="green"
                )
            except Exception:
                self.foto_path_actual = None
                self.label_imagen_nombre.configure(text="Sin imagen", text_color="gray")
        # Cambiar a pestaña de registro
        self.notebook.select(0)
        messagebox.showinfo("Edición", "✏️ Insumo cargada para edición.\nModifique los campos y presione 'Guardar'")

    def eliminar_insumo(self):
        """Elimina la insumo seleccionada en el catálogo"""
        seleccionado = self.tabla.selection()
        if not seleccionado:
            messagebox.showwarning("Atención", "Seleccione una insumo")
            return
        codigo = self.tabla.item(seleccionado[0])["values"][0]
        if not messagebox.askyesno("Confirmar", f"¿Eliminar la insumo '{codigo}'?"):
            return
        try:
            with db.get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("DELETE FROM insumo WHERE codigo = ?", (codigo,))
                conn.commit()
            messagebox.showinfo("Éxito", "Insumo eliminada correctamente")
            self.cargar_insumos()
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo eliminar:\n{e}")

    def cargar_insumos_para_mantenimiento(self):
        """Carga las insumos en el combo de mantenimiento"""
        try:
            with db.get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT id, codigo, nombre 
                    FROM insumo 
                    WHERE estado != 'Fuera de Servicio'
                    ORDER BY nombre
                """)
                insumos = [f"{row[0]}-{row[1]} - {row[2]}" for row in cursor.fetchall()]
                
                if hasattr(self, 'combo_insumo_mant'):
                    self.combo_insumo_mant.configure(values=insumos)
                    if insumos:
                        self.combo_insumo_mant.set(insumos[0])
                        
        except Exception as e:
            messagebox.showerror("Error", f"No se pudieron cargar las insumos:\n{e}")

    def actualizar_estado_insumo_mant(self):
        """Actualiza el estado de la insumo seleccionada desde la pestaña Movimientos"""
        if not hasattr(self, 'combo_insumo_mant') or not self.combo_insumo_mant.get():
            messagebox.showwarning("Atención", "Seleccione un insumo en el combo de Movimientos")
            return
        nuevo_estado = self.combo_estado_mant.get().strip() if hasattr(self, 'combo_estado_mant') else None
        if not nuevo_estado:
            messagebox.showwarning("Atención", "Seleccione un estado a aplicar")
            return
        try:
            insumo_id = int(self.combo_insumo_mant.get().split("-")[0])
            with db.get_connection() as conn:
                cursor = conn.cursor()
                try:
                    cursor.execute("UPDATE insumo SET estado = ? WHERE id = ?", (nuevo_estado, insumo_id))
                    conn.commit()
                except Exception as e:
                    # Manejo por si el estado 'En Revisión' no existe en el CHECK constraint
                    if "constraint" in str(e).lower() and nuevo_estado == "En Revisión":
                        messagebox.showerror(
                            "Error",
                            "El estado 'En Revisión' no está habilitado en la base de datos.\n"
                            "Por favor ejecute: aplicar_migraciones_mantenimiento.bat"
                        )
                        return
                    raise
            messagebox.showinfo("Éxito", f"Estado actualizado a: {nuevo_estado}")
            # Refrescar vistas
            self.cargar_insumos()
            self.cargar_insumos_para_mantenimiento()
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo actualizar el estado:\n{e}")
    
    def _actualizar_campos_precio(self, *args):
        """Muestra u oculta los campos de precio según el tipo de movimiento"""
        tipo = self.combo_tipo_mant.get()
        if tipo == "Entrada":
            self.row3b_precios.pack(fill="x", padx=10, pady=5, after=self.row3b_precios.master.winfo_children()[2])
        else:
            self.row3b_precios.pack_forget()
    
    def _calcular_precio_total_mov(self, event=None):
        """Calcula el precio total = cantidad * precio_unitario"""
        try:
            cantidad_txt = self.entry_costo_mant.get().strip().replace(",", ".")
            precio_unit_txt = self.entry_precio_unitario_mov.get().strip().replace(",", ".")
            
            if cantidad_txt and precio_unit_txt:
                cantidad = float(cantidad_txt)
                precio_unit = float(precio_unit_txt)
                total = cantidad * precio_unit
                
                # Actualizar el campo de precio total
                self.entry_precio_total_mov.configure(state="normal")
                self.entry_precio_total_mov.delete(0, "end")
                self.entry_precio_total_mov.insert(0, f"{total:,.2f}")
                self.entry_precio_total_mov.configure(state="readonly")
        except:
            pass

    def editar_insumo_desde_mantenimiento(self):
        """Abre la insumo seleccionada en la pestaña de mantenimiento para edición"""
        if not self.combo_insumo_mant.get():
            messagebox.showwarning("Atención", "Seleccione una insumo en el combo")
            return
        try:
            insumo_id = int(self.combo_insumo_mant.get().split("-")[0])
            with db.get_connection() as conn:
                conn.row_factory = sqlite3.Row
                cursor = conn.cursor()
                cursor.execute("SELECT * FROM insumo WHERE id = ?", (insumo_id,))
                row = cursor.fetchone()
                if not row:
                    messagebox.showerror("Error", "Insumo no encontrada")
                    return
                h = dict(row)
                self._cargar_insumo_en_form(h, cursor)
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo cargar la insumo:\n{e}")

    def eliminar_insumo_desde_mantenimiento(self):
        """Limpia los campos del formulario de movimientos"""
        # Limpiar todos los campos del formulario
        self.combo_insumo_mant.set("")
        self.combo_tipo_mant.set("Entrada")
        self.entry_fecha_mant.delete(0, "end")
        self.entry_fecha_mant.insert(0, datetime.now().strftime("%Y-%m-%d"))
        self.entry_costo_mant.delete(0, "end")
        self.entry_realizado_por.delete(0, "end")
        self.entry_proveedor_mant.delete(0, "end")
        self.entry_proximo_mant.delete(0, "end")
        self.text_desc_mant.delete("1.0", "end")
        self.entry_precio_unitario_mov.delete(0, "end")
        self.entry_precio_total_mov.configure(state="normal")
        self.entry_precio_total_mov.delete(0, "end")
        self.entry_precio_total_mov.configure(state="readonly")
        
        # Ocultar campos de precio
        self.row3b_precios.pack_forget()

    def guardar_mantenimiento(self):
        """Guarda un registro de mantenimiento"""
        if not self.combo_insumo_mant.get():
            messagebox.showwarning("Atención", "Seleccione una insumo")
            return
        
        try:
            # Adaptación: esta ventana registra movimientos de stock de insumos
            insumo_id = int(self.combo_insumo_mant.get().split("-")[0])
            tipo_mov = self.combo_tipo_mant.get()  # 'Entrada' | 'Salida' | 'Ajuste'
            fecha_mov = self.entry_fecha_mant.get().strip()
            responsable = self.entry_realizado_por.get().strip() or None
            observaciones = self.text_desc_mant.get("1.0", "end-1c").strip() or None
            destino_finca = self.entry_proveedor_mant.get().strip() or None  # Usamos proveedor_mant como destino finca textual
            motivo = self.entry_proximo_mant.get().strip() or None

            # VALIDACIÓN 1: Responsable debe ser texto, no número
            if responsable:
                test_value = responsable.replace('.', '').replace(',', '').replace(' ', '')
                if test_value.isdigit():
                    messagebox.showerror(
                        "Error de Validación", 
                        f"El campo 'Responsable' debe ser un nombre o texto, no un número.\n\n"
                        f"Valor ingresado: '{responsable}'"
                    )
                    return
            
            # VALIDACIÓN 2: Cantidad debe ser numérica y positiva
            try:
                cantidad_txt = (self.entry_costo_mant.get().strip() or "").replace(",", ".")
                if not cantidad_txt:
                    messagebox.showwarning("Datos incompletos", "Debe ingresar una cantidad")
                    return
                    
                cantidad = float(cantidad_txt)
                if cantidad <= 0:
                    messagebox.showerror(
                        "Error de Validación",
                        f"La cantidad debe ser un número mayor a 0\n\n"
                        f"Valor ingresado: {cantidad}"
                    )
                    return
            except ValueError:
                messagebox.showerror(
                    "Error de Validación",
                    f"La cantidad debe ser un número válido\n\n"
                    f"Valor ingresado: '{self.entry_costo_mant.get()}'"
                )
                return
            
            # VALIDACIÓN 3: Precio debe ser numérico si se ingresa
            costo_unitario = None
            costo_total = None
            if tipo_mov == "Entrada":
                precio_unit_txt = self.entry_precio_unitario_mov.get().strip().replace(",", ".")
                if precio_unit_txt:
                    try:
                        costo_unitario = float(precio_unit_txt)
                        if costo_unitario < 0:
                            messagebox.showerror(
                                "Error de Validación",
                                "El precio unitario no puede ser negativo"
                            )
                            return
                        costo_total = cantidad * costo_unitario
                    except ValueError:
                        messagebox.showerror(
                            "Error de Validación",
                            f"El precio unitario debe ser un número válido\n\n"
                            f"Valor ingresado: '{precio_unit_txt}'"
                        )
                        return

            with db.get_connection() as conn:
                cursor = conn.cursor()

                # Insertar movimiento
                cursor.execute(
                    """
                    INSERT INTO movimiento_insumo (
                        insumo_id, tipo_movimiento, cantidad, motivo, referencia,
                        usuario, observaciones, fecha_movimiento, costo_unitario, costo_total
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        insumo_id, tipo_mov, cantidad, motivo, destino_finca,
                        responsable, observaciones, fecha_mov, costo_unitario, costo_total
                    ),
                )
                
                # Obtener el ID del movimiento recién creado
                movimiento_id = cursor.lastrowid
                if not movimiento_id or movimiento_id <= 0:
                    raise Exception(
                        "Error al crear el movimiento: No se generó un ID válido.\n"
                        "Verifique que la tabla movimiento_insumo tenga PRIMARY KEY AUTOINCREMENT."
                    )

                # Actualizar stock del insumo
                cursor.execute("SELECT stock_actual, stock_bodega, responsable, estado, precio_unitario, stock_minimo FROM insumo WHERE id = ?", (insumo_id,))
                row = cursor.fetchone()
                if not row:
                    raise Exception("Insumo no encontrado")
                stock_actual, stock_bodega, responsable_actual, estado_actual, precio_anterior, stock_minimo = row
                stock_minimo = stock_minimo or 0

                if tipo_mov == "Entrada":
                    nuevo_stock_actual = (stock_actual or 0) + cantidad
                    nuevo_stock_bodega = (stock_bodega or 0) + cantidad
                    # Si estaba agotado y ahora tiene stock, cambiar a Disponible
                    nuevo_estado = "Disponible" if estado_actual == "Agotado" and nuevo_stock_actual > 0 else estado_actual
                    # Actualizar precio_unitario del insumo si se proporcionó
                    nuevo_precio = costo_unitario if costo_unitario is not None else precio_anterior
                elif tipo_mov == "Salida":
                    nuevo_stock_actual = max(0, (stock_actual or 0) - cantidad)
                    # Si salida desde bodega, reducir bodega; si asignada, reducir bodega en lo posible
                    reduce_bodega = cantidad
                    nuevo_stock_bodega = max(0, (stock_bodega or 0) - reduce_bodega)
                    # Si el stock llega a 0 o es menor/igual al mínimo, marcar como Agotado
                    nuevo_estado = "Agotado" if (nuevo_stock_actual <= 0 or nuevo_stock_actual <= stock_minimo) else estado_actual
                    nuevo_precio = precio_anterior
                else:  # Ajuste
                    # Para ajustes, solo tocar stock_actual; mantener stock_bodega si no se especifica
                    nuevo_stock_actual = max(0, (stock_actual or 0) + cantidad)
                    nuevo_stock_bodega = stock_bodega
                    nuevo_estado = estado_actual
                    nuevo_precio = precio_anterior

                cursor.execute(
                    "UPDATE insumo SET stock_actual = ?, stock_bodega = ?, estado = ?, precio_unitario = ? WHERE id = ?",
                    (nuevo_stock_actual, nuevo_stock_bodega, nuevo_estado, nuevo_precio, insumo_id),
                )

                conn.commit()

            messagebox.showinfo("Éxito", "✅ Movimiento registrado y stock actualizado")
            self.cargar_mantenimientos()
            self.cargar_insumos()  # Actualizar catálogo

            # Limpiar campos
            self.text_desc_mant.delete("1.0", "end")
            self.entry_costo_mant.delete(0, "end")
            self.entry_realizado_por.delete(0, "end")
            self.entry_proximo_mant.delete(0, "end")
            self.entry_proveedor_mant.delete(0, "end")
            self.entry_precio_unitario_mov.delete(0, "end")
            self.entry_precio_total_mov.configure(state="normal")
            self.entry_precio_total_mov.delete(0, "end")
            self.entry_precio_total_mov.configure(state="readonly")

        except Exception as e:
            messagebox.showerror("Error", f"No se pudo registrar el movimiento:\n{e}")

    def cargar_mantenimientos(self):
        """Carga el historial de movimientos de stock"""
        for item in self.tabla_mant.get_children():
            self.tabla_mant.delete(item)
        
        try:
            with db.get_connection() as conn:
                cursor = conn.cursor()
                
                cursor.execute(
                    """
                    SELECT m.id, m.fecha_movimiento, h.codigo || ' - ' || h.nombre,
                           m.tipo_movimiento, m.cantidad, 
                           COALESCE(m.costo_total, m.costo_unitario * m.cantidad, 0) as precio,
                           m.usuario, m.referencia
                    FROM movimiento_insumo m
                    JOIN insumo h ON m.insumo_id = h.id
                    WHERE m.id IS NOT NULL
                    ORDER BY m.fecha_movimiento DESC, m.id DESC
                    LIMIT 100
                    """
                )

                registros_cargados = 0
                registros_omitidos = 0
                
                for row in cursor.fetchall():
                    # Validar que el ID no sea None antes de agregar a la tabla
                    if row[0] is None:
                        registros_omitidos += 1
                        continue
                    
                    cantidad_fmt = f"{row[4]:,.2f}" if row[4] is not None else "N/A"
                    precio_fmt = f"${row[5]:,.2f}" if row[5] and row[5] > 0 else "N/A"
                    tipo = row[3]
                    tipo_display = "➕ Entrada" if tipo == "Entrada" else ("➖ Salida" if tipo == "Salida" else "⚙️ Ajuste")
                    usuario = row[6] or "N/A"
                    finca_destino = row[7] or "N/A"
                    
                    item_id = self.tabla_mant.insert(
                        "",
                        "end",
                        values=(row[0], row[1], row[2], tipo_display, cantidad_fmt, precio_fmt, usuario, finca_destino),
                    )
                    registros_cargados += 1
                    
                    # Colorear según tipo
                    if tipo == "Entrada":
                        self.tabla_mant.item(item_id, tags=("entrada",))
                    elif tipo == "Salida":
                        self.tabla_mant.item(item_id, tags=("salida",))
                    else:
                        self.tabla_mant.item(item_id, tags=("ajuste",))

                self.tabla_mant.tag_configure("entrada", background="#d4edda")  # Verde
                self.tabla_mant.tag_configure("salida", background="#f8d7da")   # Rojo claro
                self.tabla_mant.tag_configure("ajuste", background="#fff3cd")   # Amarillo
                
                # Advertir si se omitieron registros sin ID
                if registros_omitidos > 0:
                    messagebox.showwarning(
                        "Advertencia",
                        f"Se omitieron {registros_omitidos} movimientos sin ID válido.\n\n"
                        f"Ejecute la migración de base de datos para corregir este problema."
                    )
                    
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo cargar el historial:\n{e}")
    
    def completar_mantenimiento(self):
        """Elimina un movimiento del historial sin afectar el insumo en el catálogo"""
        seleccion = self.tabla_mant.selection()
        if not seleccion:
            messagebox.showwarning("Atención", "Seleccione un movimiento de la tabla")
            return
        
        try:
            # Obtener ID del movimiento (primera columna)
            item = self.tabla_mant.item(seleccion[0])
            mov_id = item['values'][0]
            
            # Validar que el ID no sea None o inválido
            if mov_id is None or mov_id == "" or mov_id == "None":
                messagebox.showerror(
                    "Error", 
                    "Movimiento sin ID válido.\n\n"
                    "Por favor, actualice la vista y vuelva a intentar.\n"
                    "Si el problema persiste, contacte al administrador del sistema."
                )
                return
            
            # Convertir a entero para asegurar tipo correcto
            try:
                mov_id = int(mov_id)
            except (ValueError, TypeError):
                messagebox.showerror(
                    "Error", 
                    f"ID de movimiento inválido: {mov_id}\n\n"
                    "Sincronice el historial y vuelva a intentar."
                )
                return
            
            # Confirmación del usuario
            if not messagebox.askyesno(
                "Confirmar Eliminación", 
                f"¿Eliminar este movimiento del historial?\n\n"
                f"ID del movimiento: {mov_id}\n\n"
                f"Nota: Esta acción no afecta el catálogo del insumo.\n"
                f"El stock no se modificará."
            ):
                return
            
            # Eliminar del historial
            with db.get_connection() as conn:
                cursor = conn.cursor()
                
                # Verificar que el movimiento existe
                cursor.execute("SELECT id FROM movimiento_insumo WHERE id = ?", (mov_id,))
                if not cursor.fetchone():
                    messagebox.showerror(
                        "Error", 
                        f"No se encontró el movimiento con ID {mov_id}.\n\n"
                        "El registro puede haber sido eliminado previamente."
                    )
                    self.cargar_mantenimientos()  # Actualizar vista
                    return
                
                # Eliminar registro
                cursor.execute("DELETE FROM movimiento_insumo WHERE id = ?", (mov_id,))
                rows_affected = cursor.rowcount

                # Si la tabla queda vacía, reiniciar secuencia de IDs (SQLite)
                cursor.execute("SELECT COUNT(*) FROM movimiento_insumo")
                restantes = cursor.fetchone()[0]
                if restantes == 0:
                    try:
                        cursor.execute("DELETE FROM sqlite_sequence WHERE name='movimiento_insumo'")
                    except Exception:
                        pass
                conn.commit()
                
                if rows_affected > 0:
                    # Eliminar de la tabla visual inmediatamente
                    self.tabla_mant.delete(seleccion[0])
                    messagebox.showinfo(
                        "Éxito", 
                        f"✅ Movimiento eliminado del historial\n\n"
                        f"ID eliminado: {mov_id}"
                    )
                else:
                    messagebox.showwarning(
                        "Advertencia", 
                        f"No se eliminó ningún registro con ID {mov_id}"
                    )
                    self.cargar_mantenimientos()  # Actualizar vista
                    
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo eliminar el movimiento:\n{e}")
            # Recargar vista en caso de error para mostrar estado real
            self.cargar_mantenimientos()
    
    def ver_detalles_mantenimiento(self):
        """Muestra los detalles completos de un movimiento de insumo"""
        seleccion = self.tabla_mant.selection()
        if not seleccion:
            messagebox.showwarning("Atención", "Seleccione un movimiento de la tabla")
            return
        
        try:
            # Obtener ID del movimiento
            item = self.tabla_mant.item(seleccion[0])
            mant_id = item['values'][0]
            
            with db.get_connection() as conn:
                cursor = conn.cursor()
                
                # Verificar si existe la columna estado_actual (migración 017)
                cursor.execute("PRAGMA table_info(mantenimiento_insumo)")
                columnas = [col[1] for col in cursor.fetchall()]
                tiene_estado_actual = 'estado_actual' in columnas
                
                cursor.execute("""
                    SELECT m.*, h.codigo || ' - ' || h.nombre as insumo_nombre
                    FROM mantenimiento_insumo m
                    JOIN insumo h ON m.insumo_id = h.id
                    WHERE m.id = ?
                """, (mant_id,))
                
                row = cursor.fetchone()
                if not row:
                    messagebox.showerror("Error", "No se encontró el movimiento")
                    return
                
                # Crear ventana de detalles
                ventana = ctk.CTkToplevel(self)
                ventana.title("Detalles del Movimiento de Insumo")
                ventana.geometry("600x500")
                ventana.transient(self.winfo_toplevel())
                ventana.grab_set()
                
                # Frame principal con scroll
                main_frame = ctk.CTkScrollableFrame(ventana)
                main_frame.pack(fill="both", expand=True, padx=10, pady=10)
                
                ctk.CTkLabel(
                    main_frame,
                    text="📋 Detalles del Mantenimiento",
                    font=("Segoe UI", 16, "bold")
                ).pack(pady=(0, 15))
                
                # Información general (ajustar índices según si tiene estado_actual)
                if tiene_estado_actual:
                    # Con migración 017: id, insumo_id, tipo, fecha, descripcion, costo, 
                    # proveedor, proximo, realizado_por, observaciones, fecha_creacion, 
                    # estado_actual, estado_previo, fecha_completado, insumo_nombre
                    info = [
                        ("ID:", row[0]),
                        ("Insumo:", row[-1]),  # última columna
                        ("Tipo:", row[2]),
                        ("Estado:", f"🔧 {row[11]}" if len(row) > 11 and row[11] == 'Activo' else f"✅ {row[11]}" if len(row) > 11 else "N/A"),
                        ("Fecha:", row[3]),
                        ("Costo:", f"${row[5]:,.2f}" if row[5] else "N/A"),
                        ("Realizado por:", row[8] or "N/A"),
                        ("Proveedor:", row[6] or "N/A"),
                        ("Próximo mant.:", row[7] or "N/A"),
                        ("Estado previo insumo:", row[12] if len(row) > 12 else "N/A"),
                        ("Fecha completado:", row[13] if len(row) > 13 else "N/A"),
                    ]
                else:
                    # Sin migración 017
                    info = [
                        ("ID:", row[0]),
                        ("Insumo:", row[-1]),  # última columna
                        ("Tipo:", row[2]),
                        ("Fecha:", row[3]),
                        ("Costo:", f"${row[5]:,.2f}" if row[5] else "N/A"),
                        ("Realizado por:", row[8] or "N/A"),
                        ("Proveedor:", row[6] or "N/A"),
                        ("Próximo mant.:", row[7] or "N/A"),
                    ]
                
                for label, valor in info:
                    row_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
                    row_frame.pack(fill="x", pady=3)
                    
                    ctk.CTkLabel(
                        row_frame,
                        text=label,
                        font=("Segoe UI", 11, "bold"),
                        width=180,
                        anchor="w"
                    ).pack(side="left", padx=5)
                    
                    ctk.CTkLabel(
                        row_frame,
                        text=str(valor),
                        font=("Segoe UI", 11),
                        anchor="w"
                    ).pack(side="left", padx=5, fill="x", expand=True)
                
                # Descripción
                if row[4]:
                    ctk.CTkLabel(
                        main_frame,
                        text="Descripción:",
                        font=("Segoe UI", 11, "bold"),
                        anchor="w"
                    ).pack(fill="x", padx=5, pady=(10, 5))
                    
                    text_desc = ctk.CTkTextbox(main_frame, height=100)
                    text_desc.pack(fill="x", padx=5, pady=5)
                    text_desc.insert("1.0", row[4])
                    text_desc.configure(state="disabled")
                
                # Observaciones
                if row[9]:
                    ctk.CTkLabel(
                        main_frame,
                        text="Observaciones:",
                        font=("Segoe UI", 11, "bold"),
                        anchor="w"
                    ).pack(fill="x", padx=5, pady=(10, 5))
                    
                    text_obs = ctk.CTkTextbox(main_frame, height=80)
                    text_obs.pack(fill="x", padx=5, pady=5)
                    text_obs.insert("1.0", row[9])
                    text_obs.configure(state="disabled")
                
                # Botón cerrar
                ctk.CTkButton(
                    ventana,
                    text="Cerrar",
                    command=ventana.destroy
                ).pack(pady=10)
                
        except Exception as e:
            messagebox.showerror("Error", f"No se pudieron cargar los detalles:\n{e}")
    
    def importar_excel(self):
        """Importa insumos desde un archivo Excel"""
        archivo = filedialog.askopenfilename(
            title="Seleccionar archivo Excel",
            filetypes=[("Excel", "*.xlsx *.xls"), ("Todos", "*.*")]
        )
        
        if not archivo:
            return
        
        try:
            wb = openpyxl.load_workbook(archivo)
            ws = wb.active
            
            # Verificar encabezados
            headers = [cell.value for cell in ws[1]]
            required = ["codigo", "nombre", "categoria"]
            
            if not all(h in [str(x).lower() if x else "" for x in headers] for h in required):
                messagebox.showerror("Error", 
                    "El archivo debe contener las columnas:\n" +
                    "codigo, nombre, categoria (obligatorias)\n" +
                    "Opcionales: finca, proveedor_principal, unidad_medida, lote_proveedor, estado, ubicacion, responsable, fecha_adquisicion,\n" +
                    "precio_unitario, stock_minimo, descripcion, observaciones, stock_actual, stock_bodega")
                return
            
            # Mapear índices de columnas
            col_map = {}
            for idx, h in enumerate(headers):
                if h:
                    col_map[str(h).lower().strip()] = idx
            
            importados = 0
            errores = []
            
            with db.get_connection() as conn:
                cursor = conn.cursor()
                
                # Obtener mapeo de fincas con normalización (nombre -> id)
                cursor.execute("SELECT id, nombre FROM finca")
                fincas_map = {}
                for fid, fname in cursor.fetchall():
                    norm = self._normalize_text(fname)
                    fincas_map[norm] = fid
                    # Agregar variante sin prefijo 'finca ' si existe
                    if norm.startswith('finca '):
                        fincas_map[norm[6:]] = fid
                
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    try:
                        codigo = row[col_map.get("codigo")]
                        nombre = row[col_map.get("nombre")]
                        categoria = row[col_map.get("categoria")]
                        
                        if not codigo or not nombre or not categoria:
                            errores.append(f"Fila {row_idx}: Faltan campos obligatorios")
                            continue
                        
                        # Obtener id_finca
                        id_finca = None
                        if "finca" in col_map and row[col_map["finca"]]:
                            finca_nombre_original = str(row[col_map["finca"]]).strip()
                            finca_key = self._normalize_text(finca_nombre_original)
                            id_finca = fincas_map.get(finca_key)
                            # Fallback quitar 'finca ' inicial
                            if id_finca is None and finca_key.startswith('finca '):
                                id_finca = fincas_map.get(finca_key[6:])
                            # Fallback búsqueda parcial si aún no se encuentra
                            if id_finca is None:
                                candidatos = [fid for k, fid in fincas_map.items() if finca_key in k]
                                if len(candidatos) == 1:
                                    id_finca = candidatos[0]
                                elif len(candidatos) > 1:
                                    errores.append(f"Fila {row_idx}: Nombre de finca '{finca_nombre_original}' ambiguo")
                            if id_finca is None:
                                errores.append(f"Fila {row_idx}: Finca '{finca_nombre_original}' no encontrada - verifique nombre")
                        
                        # Obtener otros campos opcionales
                        proveedor_principal = row[col_map.get("proveedor_principal")] if "proveedor_principal" in col_map else None
                        unidad_medida = row[col_map.get("unidad_medida")] if "unidad_medida" in col_map else None
                        lote_proveedor = row[col_map.get("lote_proveedor")] if "lote_proveedor" in col_map else None
                        estado = row[col_map.get("estado", "Operativa")] if "estado" in col_map else "Operativa"
                        ubicacion = row[col_map.get("ubicacion")] if "ubicacion" in col_map else None
                        responsable = row[col_map.get("responsable")] if "responsable" in col_map else None
                        fecha_adq = row[col_map.get("fecha_adquisicion")] if "fecha_adquisicion" in col_map else None
                        valor = None
                        if "precio_unitario" in col_map:
                            raw_valor = row[col_map.get("precio_unitario")]
                            try:
                                valor = self._parse_valor(raw_valor)
                            except ValueError:
                                errores.append(f"Fila {row_idx}: Valor inválido '{raw_valor}'")
                        vida_util = row[col_map.get("stock_minimo")] if "stock_minimo" in col_map else None
                        descripcion = row[col_map.get("descripcion")] if "descripcion" in col_map else None
                        observaciones = row[col_map.get("observaciones")] if "observaciones" in col_map else None
                        stock_actual_raw = row[col_map.get("stock_actual")] if "stock_actual" in col_map else None
                        stock_bodega_raw = row[col_map.get("stock_bodega")] if "stock_bodega" in col_map else None
                        stock_actual = self._get_stock_actual_validado(stock_actual_raw)
                        stock_bodega = self._get_stock_bodega_validado(stock_bodega_raw, stock_actual, responsable or "Bodega")
                        
                        # Buscar id_trabajador si hay responsable especificado
                        id_trabajador = None
                        if responsable:
                            responsable_norm = self._normalize_text(str(responsable))
                            cursor.execute("""
                                SELECT rowid FROM empleado 
                                WHERE estado_actual = 'Activo'
                            """)
                            for emp_id, in cursor.fetchall():
                                cursor.execute("""
                                    SELECT nombres, apellidos FROM empleado WHERE rowid = ?
                                """, (emp_id,))
                                nombres, apellidos = cursor.fetchone()
                                nombre_completo = f"{nombres} {apellidos}" if apellidos else nombres
                                if self._normalize_text(nombre_completo) == responsable_norm:
                                    id_trabajador = emp_id
                                    break
                        
                        cursor.execute("""
                            INSERT INTO insumo (
                                codigo, nombre, categoria, descripcion, proveedor_principal, unidad_medida, lote_proveedor,
                                id_finca, ubicacion, estado, fecha_adquisicion, precio_unitario,
                                stock_minimo, responsable, observaciones, stock_actual, stock_bodega, id_trabajador
                            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """, (
                            codigo, nombre, categoria, descripcion, proveedor_principal, unidad_medida, lote_proveedor,
                            id_finca, ubicacion, estado, fecha_adq, valor, vida_util, responsable, observaciones,
                            stock_actual, stock_bodega, id_trabajador
                        ))
                        
                        importados += 1
                        
                    except sqlite3.IntegrityError:
                        errores.append(f"Fila {row_idx}: Código '{codigo}' ya existe")
                    except Exception as e:
                        errores.append(f"Fila {row_idx}: {str(e)}")
                
                conn.commit()
            
            mensaje = f"✅ Importación completada:\n• {importados} insumos importadas"
            if errores:
                mensaje += f"\n• {len(errores)} errores:\n" + "\n".join(errores[:5])
                if len(errores) > 5:
                    mensaje += f"\n... y {len(errores) - 5} errores más"
            
            messagebox.showinfo("Importación", mensaje)
            self.cargar_insumos()
            
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo importar el archivo:\n{e}")
    
    def descargar_plantilla_excel(self):
        """Genera y descarga la plantilla Excel de insumos"""
        try:
            plantillas_dir = os.path.join(os.path.dirname(__file__), '../../plantillas de carga')
            os.makedirs(plantillas_dir, exist_ok=True)
            
            archivo_plantilla = os.path.join(plantillas_dir, 'plantilla_insumos.xlsx')
            
            # Crear workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Insumos"
            
            # Encabezados
            headers = [
                "codigo", "nombre", "categoria", "finca", "proveedor_principal", "unidad_medida", "lote_proveedor",
                "estado", "ubicacion", "responsable", "fecha_adquisicion", "precio_unitario",
                "stock_minimo", "descripcion", "observaciones", "stock_actual", "stock_bodega"
            ]
            ws.append(headers)
            
            # Formatear encabezados
            for cell in ws[1]:
                cell.font = openpyxl.styles.Font(bold=True)
                cell.fill = openpyxl.styles.PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.font = openpyxl.styles.Font(color="FFFFFF", bold=True)
            
            # Agregar filas de ejemplo
            ws.append([
                "HER-001", "Tractor John Deere", "Medicamento", "Finca El Prado", "John Deere",
                "5075E", "SN123456", "Operativa", "Bodega Principal", "Bodega",
                "2023-01-15", 45000.00, 10, "Tractor para labores agrícolas", "Mantenimiento al día", 1, 1
            ])
            ws.append([
                "HER-002", "Motosierra Husqvarna", "Insumo Alimento", "Finca El León", "Husqvarna",
                "450e", "HS789012", "Operativa", "Bodega Insumos", "Bodega",
                "2023-03-20", 850.00, 5, "Motosierra para poda y corte", "", 3, 3
            ])
            
            # Ajustar ancho de columnas
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
            
            # Guardar
            wb.save(archivo_plantilla)
            
            messagebox.showinfo("Éxito", 
                f"✅ Plantilla creada exitosamente en:\n{archivo_plantilla}\n\n" +
                "Columnas obligatorias: codigo, nombre, categoria\n" +
                "Opcionales (recomendado): stock_actual, stock_bodega\n" +
                "Estados válidos: Operativa, En Mantenimiento, Dañada, Fuera de Servicio\n" +
                "Categorías válidas: Medicamento, Insumo Alimento, Semilla Medico, Vehiculo, Semilla Oficina, Otro")
            
            # Abrir carpeta
            os.startfile(plantillas_dir)
            
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo crear la plantilla:\n{e}")

    # ---------------- Helpers de validación de stock -----------------
    def _get_stock_actual_validado(self, valor):
        """Valida stock_actual (float >=0). Si vacío o inválido retorna 0"""
        try:
            if valor is None:
                return 0
            if isinstance(valor, (int, float)):
                v = float(valor)
            else:
                txt = str(valor).strip().replace(",", ".")
                v = float(txt) if txt else 0
            if v < 0:
                v = 0
            return v
        except Exception:
            return 0

    def _get_stock_bodega_validado(self, valor, stock_actual, responsable_actual):
        """Valida stock_bodega (float >=0 <= stock_actual). Si asignada a trabajador y vacío -> 0"""
        asignada = responsable_actual and responsable_actual != "Bodega"
        try:
            if valor is None or (isinstance(valor, str) and not valor.strip()):
                return 0 if asignada else stock_actual
            if isinstance(valor, (int, float)):
                v = float(valor)
            else:
                v = float(str(valor).strip().replace(",", "."))
            if v < 0:
                v = 0
            if v > stock_actual:
                v = stock_actual
            # Heurística: si asignada y stock_bodega == stock_actual, disminuir 1 para reflejar uso
            if asignada and v == stock_actual:
                v = max(0, stock_actual - 1)
            return v
        except Exception:
            return 0 if asignada else stock_actual
