"""
Módulo de Gestión de Herramientas y Equipos
Permite registrar, actualizar y dar seguimiento a las herramientas de la finca
Incluye: Edición, Eliminación, Importación Excel, Imágenes, Asignación a trabajadores
"""
import customtkinter as ctk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
import sqlite3
import sys
import os
from PIL import Image
import shutil
import openpyxl
import unicodedata

sys.path.append(os.path.join(os.path.dirname(__file__), '../..'))
from database import db


class HerramientasModule(ctk.CTkFrame):
    def __init__(self, master):
        super().__init__(master)
        self.pack(fill="both", expand=True)
        self.foto_path_actual = None  # Ruta de la foto seleccionada
        self.herramienta_editando_id = None  # ID de herramienta en edición
        self.crear_widgets()
        self.cargar_fincas()
        self.cargar_trabajadores()
        self.cargar_herramientas()

    # ----------------------- Helpers de Normalización y Parsing -----------------------
    def _normalize_text(self, texto: str) -> str:
        """Normaliza texto para comparaciones (lowercase, sin acentos, espacios simples)."""
        if texto is None:
            return ""
        t = str(texto).strip().lower()
        # Remover acentos
        t = unicodedata.normalize('NFD', t)
        t = ''.join(c for c in t if unicodedata.category(c) != 'Mn')
        # Colapsar espacios múltiples
        while "  " in t:
            t = t.replace("  ", " ")
        return t

    def _parse_valor(self, valor_raw):
        """Convierte distintos formatos de valor monetario a float seguro.
        Acepta: 170000, 170.000, 170,000, $170000.50, 170000,50
        Retorna None si vacío. Lanza ValueError si imposible convertir.
        """
        if valor_raw is None:
            return None
        s = str(valor_raw).strip()
        if s == "":
            return None
        # Quitar símbolo de moneda y espacios
        s = s.replace('$', '').replace('₡', '').replace('€', '').replace('USD', '').strip()
        # Reemplazar separadores decimales comunes
        # Caso: si hay tanto ',' como '.' decidir último como decimal y quitar el otro
        if ',' in s and '.' in s:
            # Tomar el que aparece más a la derecha como decimal
            last_comma = s.rfind(',')
            last_dot = s.rfind('.')
            if last_comma > last_dot:  # coma es decimal
                s = s.replace('.', '')  # quitar puntos miles
                s = s.replace(',', '.')  # coma decimal a punto
            else:  # punto es decimal
                s = s.replace(',', '')  # quitar comas miles
        else:
            # Solo comas: asumir coma decimal si hay una y parte decimal corta, si más de una -> quitar todas como miles
            if ',' in s and s.count(',') == 1 and len(s.split(',')[-1]) <= 2:
                s = s.replace(',', '.')
            else:
                s = s.replace(',', '')
            # Solo puntos: si múltiples puntos -> quitar como miles excepto último
            if s.count('.') > 1:
                parts = s.split('.')
                s = ''.join(parts[:-1]) + '.' + parts[-1]
        # Quitar espacios finales
        s = s.strip()
        try:
            return float(s)
        except ValueError:
            raise ValueError(f"Formato de valor inválido: '{valor_raw}'")

    def crear_widgets(self):
        # Contenedor principal sin scroll para ocupar todo el espacio
        container = ctk.CTkFrame(self, fg_color="transparent")
        container.pack(fill="both", expand=True, padx=2, pady=5)
        
        # Título
        titulo = ctk.CTkLabel(
            container,
            text="🔧 Gestión de Herramientas y Equipos",
            font=("Segoe UI", 22, "bold")
        )
        titulo.pack(pady=(5, 3))

        # Frame principal con tabs
        self.notebook = ttk.Notebook(container)
        self.notebook.pack(fill="both", expand=True, padx=2, pady=(3, 10))

        # Tab: Registrar Herramienta
        self.frame_registro = ctk.CTkFrame(self.notebook)
        self.notebook.add(self.frame_registro, text="➕ Nueva Herramienta")

        # Tab: Catálogo
        self.frame_catalogo = ctk.CTkFrame(self.notebook)
        self.notebook.add(self.frame_catalogo, text="📋 Catálogo")

        # Tab: Mantenimientos
        self.frame_mantenimientos = ctk.CTkFrame(self.notebook)
        self.notebook.add(self.frame_mantenimientos, text="🔧 Mantenimientos")

        # Crear contenido
        self.crear_formulario_registro()
        self.crear_catalogo()
        self.crear_mantenimientos()

    def crear_formulario_registro(self):
        """Formulario para registrar nueva herramienta"""
        main_frame = ctk.CTkScrollableFrame(self.frame_registro)
        main_frame.pack(fill="both", expand=True, padx=10, pady=5)

        ctk.CTkLabel(
            main_frame,
            text="📝 Registrar Nueva Herramienta/Equipo",
            font=("Segoe UI", 18, "bold")
        ).pack(pady=(0, 5))

        # Información básica
        info_frame = ctk.CTkFrame(main_frame, corner_radius=10)
        info_frame.pack(fill="both", expand=True, pady=5)
        
        ctk.CTkLabel(info_frame, text="📋 INFORMACIÓN BÁSICA", 
                    font=("Segoe UI", 14, "bold")).pack(anchor="w", padx=10, pady=10)

        # Código y Nombre
        row1 = ctk.CTkFrame(info_frame, fg_color="transparent")
        row1.pack(fill="x", padx=10, pady=5)
        ctk.CTkLabel(row1, text="Código *:", width=150).pack(side="left", padx=5)
        self.entry_codigo = ctk.CTkEntry(row1, width=200)
        self.entry_codigo.pack(side="left", padx=5)
        
        ctk.CTkLabel(row1, text="Nombre *:", width=100).pack(side="left", padx=5)
        self.entry_nombre = ctk.CTkEntry(row1, width=300)
        self.entry_nombre.pack(side="left", padx=5)

        # Categoría y Finca
        row2 = ctk.CTkFrame(info_frame, fg_color="transparent")
        row2.pack(fill="x", padx=10, pady=5)
        ctk.CTkLabel(row2, text="Categoría *:", width=150).pack(side="left", padx=5)
        self.combo_categoria = ctk.CTkComboBox(
            row2, 
            values=["Maquinaria", "Herramienta Manual", "Equipo Medico", "Vehiculo", "Equipo Oficina", "Otro"],
            width=200
        )
        self.combo_categoria.pack(side="left", padx=5)
        
        ctk.CTkLabel(row2, text="Finca:", width=100).pack(side="left", padx=5)
        self.combo_finca = ctk.CTkComboBox(row2, width=300)
        self.combo_finca.pack(side="left", padx=5)

        # Marca y Modelo
        row3 = ctk.CTkFrame(info_frame, fg_color="transparent")
        row3.pack(fill="x", padx=10, pady=5)
        ctk.CTkLabel(row3, text="Marca:", width=150).pack(side="left", padx=5)
        self.entry_marca = ctk.CTkEntry(row3, width=200)
        self.entry_marca.pack(side="left", padx=5)
        
        ctk.CTkLabel(row3, text="Modelo:", width=100).pack(side="left", padx=5)
        self.entry_modelo = ctk.CTkEntry(row3, width=300)
        self.entry_modelo.pack(side="left", padx=5)

        # Número de Serie y Estado
        row4 = ctk.CTkFrame(info_frame, fg_color="transparent")
        row4.pack(fill="x", padx=10, pady=5)
        ctk.CTkLabel(row4, text="N° Serie:", width=150).pack(side="left", padx=5)
        self.entry_serie = ctk.CTkEntry(row4, width=200)
        self.entry_serie.pack(side="left", padx=5)
        
        ctk.CTkLabel(row4, text="Estado:", width=100).pack(side="left", padx=5)
        self.combo_estado = ctk.CTkComboBox(
            row4,
            values=["Operativa", "En Mantenimiento", "En Revisión", "Dañada", "Fuera de Servicio"],
            width=300
        )
        self.combo_estado.set("Operativa")
        self.combo_estado.pack(side="left", padx=5)

        # Ubicación y Responsable
        row5 = ctk.CTkFrame(info_frame, fg_color="transparent")
        row5.pack(fill="x", padx=10, pady=5)
        ctk.CTkLabel(row5, text="Ubicación:", width=150).pack(side="left", padx=5)
        self.entry_ubicacion = ctk.CTkEntry(row5, width=200)
        self.entry_ubicacion.pack(side="left", padx=5)
        
        ctk.CTkLabel(row5, text="Responsable:", width=100).pack(side="left", padx=5)
        self.combo_responsable = ctk.CTkComboBox(row5, width=300)
        self.combo_responsable.pack(side="left", padx=5)
        
        # Stock (Inventario básico)
        stock_frame = ctk.CTkFrame(info_frame, fg_color="transparent")
        stock_frame.pack(fill="x", padx=10, pady=5)
        ctk.CTkLabel(stock_frame, text="Stock Total:", width=150).pack(side="left", padx=5)
        self.entry_stock_total = ctk.CTkEntry(stock_frame, width=120)
        self.entry_stock_total.insert(0, "1")
        self.entry_stock_total.pack(side="left", padx=5)
        ctk.CTkLabel(stock_frame, text="Stock Bodega:", width=120).pack(side="left", padx=5)
        self.entry_stock_bodega = ctk.CTkEntry(stock_frame, width=120)
        self.entry_stock_bodega.insert(0, "1")
        self.entry_stock_bodega.pack(side="left", padx=5)
        
        # Imagen de la herramienta
        img_frame = ctk.CTkFrame(info_frame, corner_radius=10)
        img_frame.pack(fill="x", padx=10, pady=10)
        
        ctk.CTkLabel(img_frame, text="📷 Imagen:", width=150).pack(side="left", padx=5)
        ctk.CTkButton(
            img_frame,
            text="📁 Seleccionar Imagen",
            command=self.seleccionar_imagen,
            width=180
        ).pack(side="left", padx=5)
        
        self.label_imagen_nombre = ctk.CTkLabel(img_frame, text="Sin imagen", text_color="gray")
        self.label_imagen_nombre.pack(side="left", padx=10)
        
        ctk.CTkButton(
            img_frame,
            text="👁️ Ver",
            command=self.ver_imagen,
            width=80
        ).pack(side="left", padx=5)
        
        ctk.CTkButton(
            img_frame,
            text="🗑️ Quitar",
            command=self.quitar_imagen,
            width=80,
            fg_color="#D32F2F"
        ).pack(side="left", padx=5)

        # Información financiera
        fin_frame = ctk.CTkFrame(main_frame, corner_radius=10)
        fin_frame.pack(fill="x", pady=10)
        
        ctk.CTkLabel(fin_frame, text="💰 INFORMACIÓN FINANCIERA", 
                    font=("Segoe UI", 14, "bold")).pack(anchor="w", padx=10, pady=10)

        row6 = ctk.CTkFrame(fin_frame, fg_color="transparent")
        row6.pack(fill="x", padx=10, pady=5)
        ctk.CTkLabel(row6, text="Fecha Adquisición:", width=150).pack(side="left", padx=5)
        self.entry_fecha_adq = ctk.CTkEntry(row6, width=150)
        self.entry_fecha_adq.insert(0, datetime.now().strftime("%Y-%m-%d"))
        self.entry_fecha_adq.pack(side="left", padx=5)
        
        ctk.CTkLabel(row6, text="Valor $:", width=80).pack(side="left", padx=5)
        self.entry_valor = ctk.CTkEntry(row6, width=150)
        self.entry_valor.pack(side="left", padx=5)
        
        ctk.CTkLabel(row6, text="Vida Útil (años):", width=120).pack(side="left", padx=5)
        self.entry_vida_util = ctk.CTkEntry(row6, width=100)
        self.entry_vida_util.pack(side="left", padx=5)

        # Descripción y Observaciones
        row7 = ctk.CTkFrame(main_frame, fg_color="transparent")
        row7.pack(fill="both", expand=True, pady=5)
        ctk.CTkLabel(row7, text="Descripción:", width=150).pack(side="left", padx=5, anchor="n")
        self.text_descripcion = ctk.CTkTextbox(row7, width=600, height=100)
        self.text_descripcion.pack(side="left", padx=5, fill="both", expand=True)

        row8 = ctk.CTkFrame(main_frame, fg_color="transparent")
        row8.pack(fill="both", expand=True, pady=5)
        ctk.CTkLabel(row8, text="Observaciones:", width=150).pack(side="left", padx=5, anchor="n")
        self.text_observaciones = ctk.CTkTextbox(row8, width=600, height=100)
        self.text_observaciones.pack(side="left", padx=5, fill="both", expand=True)

        # Botones
        btn_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        btn_frame.pack(pady=5)
        ctk.CTkButton(
            btn_frame, 
            text="💾 Guardar Herramienta", 
            command=self.guardar_herramienta,
            fg_color="green",
            hover_color="#006400"
        ).pack(side="left", padx=5)
        ctk.CTkButton(
            btn_frame,
            text="🔄 Limpiar Formulario",
            command=self.limpiar_formulario
        ).pack(side="left", padx=5)

    def crear_catalogo(self):
        """Crea el catálogo de herramientas"""
        main_frame = ctk.CTkFrame(self.frame_catalogo)
        main_frame.pack(fill="both", expand=True, padx=10, pady=10)

        ctk.CTkLabel(
            main_frame,
            text="📋 Catálogo de Herramientas y Equipos",
            font=("Segoe UI", 18, "bold")
        ).pack(pady=(0, 10))

        # Filtros
        filtro_frame = ctk.CTkFrame(main_frame)
        filtro_frame.pack(fill="x", pady=10)
        
        ctk.CTkLabel(filtro_frame, text="Filtrar por Finca:").pack(side="left", padx=5)
        self.combo_filtro_finca = ctk.CTkComboBox(filtro_frame, width=200)
        self.combo_filtro_finca.pack(side="left", padx=5)
        
        ctk.CTkLabel(filtro_frame, text="Estado:").pack(side="left", padx=5)
        self.combo_filtro_estado = ctk.CTkComboBox(
            filtro_frame,
            values=["Todos", "Operativa", "En Mantenimiento", "En Revisión", "Dañada", "Fuera de Servicio"],
            width=180
        )
        self.combo_filtro_estado.set("Todos")
        self.combo_filtro_estado.pack(side="left", padx=5)
        
        ctk.CTkButton(
            filtro_frame,
            text="🔍 Filtrar",
            command=self.cargar_herramientas
        ).pack(side="left", padx=5)

        # Frame para tabla (centrar y permitir barra a la derecha)
        tabla_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        tabla_frame.pack(fill="both", expand=True, pady=(5, 5))

        self.tabla = ttk.Treeview(
            tabla_frame,
            columns=("codigo", "nombre", "categoria", "finca", "stock_total", "asignacion", "stock_bodega", "estado"),
            show="headings",
            height=15
        )
        columnas = [
            ("codigo", "Código", 100),
            ("nombre", "Nombre", 190),
            ("categoria", "Categoría", 140),
            ("finca", "Finca", 140),
            ("stock_total", "Stock Total", 95),
            ("asignacion", "Asignación", 100),
            ("stock_bodega", "Stock Bodega", 105),
            ("estado", "Estado", 120)
        ]
        for col, heading, width in columnas:
            self.tabla.heading(col, text=heading)
            self.tabla.column(col, width=width, anchor="center")
        self.tabla.pack(side="left", fill="both", expand=True)
        scrollbar = ttk.Scrollbar(tabla_frame, orient="vertical", command=self.tabla.yview)
        self.tabla.configure(yscroll=scrollbar.set)
        scrollbar.pack(side="right", fill="y")

        # Barra inferior de acciones (botones alineados abajo)
        acciones_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        acciones_frame.pack(fill="x", pady=(0, 8))

        # Usar grid para control de distribución y permitir expansión
        botones = [
            {
                "texto": "🔍 Ver Detalles",
                "comando": self.ver_detalles_herramienta,
                "color": None,
                "hover": None
            },
            {
                "texto": "✏️ Editar Seleccionado",
                "comando": self.editar_herramienta,
                "color": "#1976D2",
                "hover": "#1565C0"
            },
            {
                "texto": "🗑️ Eliminar Seleccionado",
                "comando": self.eliminar_herramienta,
                "color": "#D32F2F",
                "hover": "#C62828"
            },
            {
                "texto": "📥 Importar Excel",
                "comando": self.importar_excel,
                "color": "#388E3C",
                "hover": "#2E7D32"
            },
            {
                "texto": "📋 Descargar Plantilla",
                "comando": self.descargar_plantilla_excel,
                "color": "#F57C00",
                "hover": "#EF6C00"
            }
        ]

        for idx, btn in enumerate(botones):
            b = ctk.CTkButton(
                acciones_frame,
                text=btn["texto"],
                command=btn["comando"],
                fg_color=btn["color"] if btn["color"] else None,
                hover_color=btn["hover"] if btn["hover"] else None,
                width=160
            )
            b.grid(row=0, column=idx, padx=5, pady=5)
            acciones_frame.grid_columnconfigure(idx, weight=1)

    def crear_mantenimientos(self):
        """Crear el sistema de mantenimientos"""
        main_frame = ctk.CTkFrame(self.frame_mantenimientos)
        main_frame.pack(fill="both", expand=True, padx=10, pady=10)

        ctk.CTkLabel(
            main_frame,
            text="🔧 Registro de Mantenimientos",
            font=("Segoe UI", 18, "bold")
        ).pack(pady=(0, 10))

        # Formulario de mantenimiento con scroll
        form_scroll = ctk.CTkScrollableFrame(main_frame, height=280, corner_radius=10)
        form_scroll.pack(fill="x", expand=False, pady=(0, 10))
        
        form_frame = ctk.CTkFrame(form_scroll, fg_color="transparent")
        form_frame.pack(fill="both", padx=5, pady=5)

        row1 = ctk.CTkFrame(form_frame, fg_color="transparent")
        row1.pack(fill="x", padx=10, pady=10)
        
        ctk.CTkLabel(row1, text="Herramienta *:", width=150).pack(side="left", padx=5)
        self.combo_herramienta_mant = ctk.CTkComboBox(row1, width=300)
        self.combo_herramienta_mant.pack(side="left", padx=5)
        
        ctk.CTkButton(
            row1,
            text="🔄 Actualizar Lista",
            command=self.cargar_herramientas_para_mantenimiento
        ).pack(side="left", padx=5)

        row2 = ctk.CTkFrame(form_frame, fg_color="transparent")
        row2.pack(fill="x", padx=10, pady=5)
        
        ctk.CTkLabel(row2, text="Tipo *:", width=150).pack(side="left", padx=5)
        self.combo_tipo_mant = ctk.CTkComboBox(
            row2,
            values=["Preventivo", "Correctivo", "Calibración", "Inspección"],
            width=200
        )
        self.combo_tipo_mant.pack(side="left", padx=5)
        
        ctk.CTkLabel(row2, text="Fecha:", width=80).pack(side="left", padx=5)
        self.entry_fecha_mant = ctk.CTkEntry(row2, width=150)
        self.entry_fecha_mant.insert(0, datetime.now().strftime("%Y-%m-%d"))
        self.entry_fecha_mant.pack(side="left", padx=5)

        row3 = ctk.CTkFrame(form_frame, fg_color="transparent")
        row3.pack(fill="x", padx=10, pady=5)
        
        ctk.CTkLabel(row3, text="Costo $:", width=150).pack(side="left", padx=5)
        self.entry_costo_mant = ctk.CTkEntry(row3, width=150)
        self.entry_costo_mant.pack(side="left", padx=5)
        
        ctk.CTkLabel(row3, text="Realizado por:", width=120).pack(side="left", padx=5)
        self.entry_realizado_por = ctk.CTkEntry(row3, width=250)
        self.entry_realizado_por.pack(side="left", padx=5)

        row4 = ctk.CTkFrame(form_frame, fg_color="transparent")
        row4.pack(fill="x", padx=10, pady=5)
        
        ctk.CTkLabel(row4, text="Próximo Mant.:", width=150).pack(side="left", padx=5)
        self.entry_proximo_mant = ctk.CTkEntry(row4, width=150)
        self.entry_proximo_mant.pack(side="left", padx=5)
        
        ctk.CTkLabel(row4, text="Proveedor:", width=120).pack(side="left", padx=5)
        self.entry_proveedor_mant = ctk.CTkEntry(row4, width=250)
        self.entry_proveedor_mant.pack(side="left", padx=5)

        # Fila rápida para actualizar estado de la herramienta
        row_estado = ctk.CTkFrame(form_frame, fg_color="transparent")
        row_estado.pack(fill="x", padx=10, pady=5)
        ctk.CTkLabel(row_estado, text="Estado herramienta:", width=150).pack(side="left", padx=5)
        self.combo_estado_mant = ctk.CTkComboBox(
            row_estado,
            values=["Operativa", "En Mantenimiento", "En Revisión", "Dañada", "Fuera de Servicio"],
            width=220
        )
        self.combo_estado_mant.set("Operativa")
        self.combo_estado_mant.pack(side="left", padx=5)
        ctk.CTkButton(
            row_estado,
            text="Actualizar Estado",
            command=self.actualizar_estado_herramienta_mant,
            fg_color="#6c757d"
        ).pack(side="left", padx=8)

        row5 = ctk.CTkFrame(form_frame, fg_color="transparent")
        row5.pack(fill="x", padx=10, pady=5)
        
        ctk.CTkLabel(row5, text="Descripción:", width=150).pack(side="left", padx=5, anchor="n")
        self.text_desc_mant = ctk.CTkTextbox(row5, width=500, height=100)
        self.text_desc_mant.pack(side="left", padx=5, fill="x", expand=True)

        # Frame de botones - EMPAQUETADO PRIMERO con side="bottom" para que se quede abajo
        btn_bottom_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        btn_bottom_frame.pack(side="bottom", fill="x", pady=(5, 10))
        
        # Columna izquierda: Acciones de mantenimiento
        btn_left = ctk.CTkFrame(btn_bottom_frame, fg_color="transparent")
        btn_left.pack(side="left", padx=5)
        
        ctk.CTkButton(
            btn_left,
            text="💾 Registrar Mantenimiento",
            command=self.guardar_mantenimiento,
            fg_color="green",
            width=200
        ).pack(side="left", padx=3)
        
        ctk.CTkButton(
            btn_left,
            text="✅ Completar Mantenimiento",
            command=self.completar_mantenimiento,
            fg_color="#28a745",
            hover_color="#218838",
            width=200
        ).pack(side="left", padx=3)
        
        ctk.CTkButton(
            btn_left,
            text="📋 Ver Detalles",
            command=self.ver_detalles_mantenimiento,
            fg_color="#17a2b8",
            hover_color="#138496",
            width=150
        ).pack(side="left", padx=3)
        
        # Columna derecha: Acciones de herramienta
        btn_right = ctk.CTkFrame(btn_bottom_frame, fg_color="transparent")
        btn_right.pack(side="right", padx=5)
        
        ctk.CTkButton(
            btn_right,
            text="✏️ Editar Herramienta",
            command=self.editar_herramienta_desde_mantenimiento,
            fg_color="#0069d9",
            width=180
        ).pack(side="left", padx=3)
        
        ctk.CTkButton(
            btn_right,
            text="🗑️ Eliminar Registro",
            command=self.eliminar_herramienta_desde_mantenimiento,
            fg_color="#dc3545",
            width=180
        ).pack(side="left", padx=3)

        # Historial de mantenimientos - EMPAQUETADO DESPUÉS de los botones
        hist_frame = ctk.CTkFrame(main_frame)
        hist_frame.pack(fill="both", expand=True, pady=(0, 5))
        
        ctk.CTkLabel(
            hist_frame,
            text="📅 Historial de Mantenimientos",
            font=("Segoe UI", 14, "bold")
        ).pack(pady=5)

        # Frame contenedor para tabla
        tabla_container = ctk.CTkFrame(hist_frame, fg_color="transparent")
        tabla_container.pack(fill="both", expand=True)
        
        self.tabla_mant = ttk.Treeview(
            tabla_container,
            columns=("id", "fecha", "herramienta", "tipo", "estado", "costo", "realizado_por", "proximo"),
            show="headings",
            height=6
        )
        
        columnas_mant = [
            ("id", "ID", 50),
            ("fecha", "Fecha", 90),
            ("herramienta", "Herramienta", 220),
            ("tipo", "Tipo", 110),
            ("estado", "Estado", 100),
            ("costo", "Costo", 90),
            ("realizado_por", "Realizado Por", 140),
            ("proximo", "Próximo", 90)
        ]
        
        for col, heading, width in columnas_mant:
            self.tabla_mant.heading(col, text=heading)
            self.tabla_mant.column(col, width=width, anchor="center")

        self.tabla_mant.pack(side="left", fill="both", expand=True)
        
        scrollbar_mant = ttk.Scrollbar(tabla_container, orient="vertical", command=self.tabla_mant.yview)
        self.tabla_mant.configure(yscroll=scrollbar_mant.set)
        scrollbar_mant.pack(side="right", fill="y")
        
        self.cargar_mantenimientos()

    def cargar_fincas(self):
        """Carga las fincas disponibles"""
        try:
            with db.get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT id, nombre FROM finca WHERE estado = 'Activa' OR estado = 'Activo'")
                fincas = [f"{row[0]}-{row[1]}" for row in cursor.fetchall()]
                
                fincas_con_todas = ["Todas"] + fincas
                
                if hasattr(self, 'combo_finca'):
                    self.combo_finca.configure(values=fincas)
                    if fincas:
                        self.combo_finca.set(fincas[0])
                
                if hasattr(self, 'combo_filtro_finca'):
                    self.combo_filtro_finca.configure(values=fincas_con_todas)
                    self.combo_filtro_finca.set("Todas")
                    
        except Exception as e:
            messagebox.showerror("Error", f"No se pudieron cargar las fincas:\n{e}")
    
    def cargar_trabajadores(self):
        """Carga los empleados activos para asignar como responsables"""
        try:
            with db.get_connection() as conn:
                cursor = conn.cursor()
                # Verificar primero si la tabla empleado existe
                cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='empleado'")
                if not cursor.fetchone():
                    # Tabla no existe, usar solo Bodega
                    if hasattr(self, 'combo_responsable'):
                        self.combo_responsable.configure(values=["Bodega"])
                        self.combo_responsable.set("Bodega")
                    return
                
                # Usar tabla empleado (de nómina) en lugar de trabajador
                cursor.execute("""
                    SELECT rowid, nombres || ' ' || apellidos as nombre_completo, cargo 
                    FROM empleado 
                    WHERE estado_actual = 'Activo' OR estado_actual IS NULL
                    ORDER BY nombres
                """)
                trabajadores = [f"{row[0]}-{row[1]} ({row[2] or 'Sin cargo'})" for row in cursor.fetchall()]
                
                # Agregar opción "Bodega" al inicio
                opciones = ["Bodega"] + trabajadores
                
                if hasattr(self, 'combo_responsable'):
                    self.combo_responsable.configure(values=opciones)
                    self.combo_responsable.set("Bodega")
                    
        except Exception:
            # Si ocurre cualquier error, usar solo "Bodega"
            if hasattr(self, 'combo_responsable'):
                self.combo_responsable.configure(values=["Bodega"])
                self.combo_responsable.set("Bodega")
    
    def seleccionar_imagen(self):
        """Selecciona una imagen para la herramienta"""
        archivo = filedialog.askopenfilename(
            title="Seleccionar imagen de la herramienta",
            filetypes=[
                ("Imágenes", "*.png *.jpg *.jpeg *.gif *.bmp"),
                ("Todos los archivos", "*.*")
            ]
        )
        if archivo:
            self.foto_path_actual = archivo
            nombre_archivo = os.path.basename(archivo)
            self.label_imagen_nombre.configure(text=nombre_archivo, text_color="green")
    
    def ver_imagen(self):
        """Muestra la imagen seleccionada"""
        if not self.foto_path_actual:
            messagebox.showinfo("Info", "No hay imagen seleccionada")
            return
        
        try:
            ventana = ctk.CTkToplevel(self)
            ventana.title("Vista previa de imagen")
            ventana.geometry("600x600")
            
            img = Image.open(self.foto_path_actual)
            # Redimensionar manteniendo aspecto
            img.thumbnail((580, 580), Image.Resampling.LANCZOS)
            
            ctk_img = ctk.CTkImage(light_image=img, dark_image=img, size=img.size)
            label = ctk.CTkLabel(ventana, image=ctk_img, text="")
            label.pack(padx=10, pady=10)
            
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo abrir la imagen:\n{e}")
    
    def quitar_imagen(self):
        """Quita la imagen seleccionada"""
        self.foto_path_actual = None
        self.label_imagen_nombre.configure(text="Sin imagen", text_color="gray")

    def guardar_herramienta(self):
        """Guarda una nueva herramienta o actualiza una existente"""
        codigo = self.entry_codigo.get().strip()
        nombre = self.entry_nombre.get().strip()
        categoria = self.combo_categoria.get()
        
        if not codigo or not nombre or not categoria:
            messagebox.showwarning("Atención", "Complete los campos obligatorios (Código, Nombre, Categoría)")
            return
        
        try:
            id_finca = None
            if self.combo_finca.get():
                id_finca = int(self.combo_finca.get().split("-")[0])
            
            valor = None
            if self.entry_valor.get().strip():
                try:
                    valor = self._parse_valor(self.entry_valor.get())
                except ValueError as e:
                    messagebox.showerror("Error", str(e))
                    return
            
            vida_util = None
            if self.entry_vida_util.get().strip():
                vida_util = int(self.entry_vida_util.get())
            
            # Obtener id_trabajador del combo responsable (usa rowid de empleado)
            id_trabajador = None
            responsable_text = None
            responsable_sel = self.combo_responsable.get()
            if responsable_sel and responsable_sel != "Bodega":
                try:
                    id_trabajador = int(responsable_sel.split("-")[0])
                    # Obtener nombre del empleado para campo responsable
                    with db.get_connection() as conn:
                        cursor = conn.cursor()
                        cursor.execute("SELECT nombres || ' ' || apellidos FROM empleado WHERE rowid = ?", (id_trabajador,))
                        row = cursor.fetchone()
                        if row:
                            responsable_text = row[0]
                except:
                    responsable_text = responsable_sel
            elif responsable_sel == "Bodega":
                responsable_text = "Bodega"
            
            # Copiar imagen a carpeta uploads/herramientas si se seleccionó una
            foto_final_path = None
            if self.foto_path_actual:
                try:
                    uploads_dir = os.path.join(os.path.dirname(__file__), '../../uploads/herramientas')
                    os.makedirs(uploads_dir, exist_ok=True)
                    
                    # Generar nombre único
                    ext = os.path.splitext(self.foto_path_actual)[1]
                    nombre_archivo = f"herr_{codigo}_{datetime.now().strftime('%Y%m%d%H%M%S')}{ext}"
                    foto_final_path = os.path.join(uploads_dir, nombre_archivo)
                    
                    shutil.copy2(self.foto_path_actual, foto_final_path)
                except Exception as e:
                    messagebox.showwarning("Advertencia", f"No se pudo copiar la imagen:\n{e}")
            
            # Validaciones de stock
            stock_total = self._get_stock_total_validado(self.entry_stock_total.get().strip())
            stock_bodega = self._get_stock_bodega_validado(self.entry_stock_bodega.get().strip(), stock_total, self.combo_responsable.get())
            # Regla explícita adicional: si está asignada y stock_total == 1 -> stock_bodega = 0
            asignada_flag = (self.combo_responsable.get() and self.combo_responsable.get() != "Bodega") or id_trabajador is not None
            if asignada_flag and stock_total == 1:
                stock_bodega = 0

            with db.get_connection() as conn:
                cursor = conn.cursor()
                
                if self.herramienta_editando_id:
                    # Modo edición
                    cursor.execute("""
                        UPDATE herramienta SET
                            codigo = ?, nombre = ?, categoria = ?, descripcion = ?,
                            marca = ?, modelo = ?, numero_serie = ?, id_finca = ?,
                            ubicacion = ?, estado = ?, fecha_adquisicion = ?,
                            valor_adquisicion = ?, vida_util_anos = ?, responsable = ?,
                            observaciones = ?, id_trabajador = ?, foto_path = ?, stock_total = ?, stock_bodega = ?
                        WHERE id = ?
                    """, (
                        codigo, nombre, categoria,
                        self.text_descripcion.get("1.0", "end-1c").strip() or None,
                        self.entry_marca.get().strip() or None,
                        self.entry_modelo.get().strip() or None,
                        self.entry_serie.get().strip() or None,
                        id_finca,
                        self.entry_ubicacion.get().strip() or None,
                        self.combo_estado.get(),
                        self.entry_fecha_adq.get().strip() or None,
                        valor,
                        vida_util,
                        responsable_text,
                        self.text_observaciones.get("1.0", "end-1c").strip() or None,
                        id_trabajador,
                        foto_final_path or self.foto_path_actual,
                        stock_total, stock_bodega,
                        self.herramienta_editando_id
                    ))
                    messagebox.showinfo("Éxito", "✅ Herramienta actualizada correctamente")
                else:
                    # Modo creación
                    cursor.execute("""
                        INSERT INTO herramienta (
                            codigo, nombre, categoria, descripcion, marca, modelo, numero_serie,
                            id_finca, ubicacion, estado, fecha_adquisicion, valor_adquisicion,
                            vida_util_anos, responsable, observaciones, id_trabajador, foto_path,
                            stock_total, stock_bodega
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (
                        codigo, nombre, categoria,
                        self.text_descripcion.get("1.0", "end-1c").strip() or None,
                        self.entry_marca.get().strip() or None,
                        self.entry_modelo.get().strip() or None,
                        self.entry_serie.get().strip() or None,
                        id_finca,
                        self.entry_ubicacion.get().strip() or None,
                        self.combo_estado.get(),
                        self.entry_fecha_adq.get().strip() or None,
                        valor,
                        vida_util,
                        responsable_text,
                        self.text_observaciones.get("1.0", "end-1c").strip() or None,
                        id_trabajador,
                        foto_final_path,
                        stock_total, stock_bodega
                    ))
                    messagebox.showinfo("Éxito", "✅ Herramienta registrada correctamente")
                
                conn.commit()
            
            self.limpiar_formulario()
            self.cargar_herramientas()
            self.notebook.select(1)  # Cambiar a pestaña de catálogo
            
        except sqlite3.IntegrityError:
            messagebox.showerror("Error", f"Ya existe una herramienta con el código '{codigo}'")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo guardar la herramienta:\n{e}")

    def limpiar_formulario(self):
        """Limpia el formulario"""
        self.entry_codigo.delete(0, "end")
        self.entry_nombre.delete(0, "end")
        self.entry_marca.delete(0, "end")
        self.entry_modelo.delete(0, "end")
        self.entry_serie.delete(0, "end")
        self.entry_ubicacion.delete(0, "end")
        self.entry_fecha_adq.delete(0, "end")
        self.entry_fecha_adq.insert(0, datetime.now().strftime("%Y-%m-%d"))
        self.entry_valor.delete(0, "end")
        self.entry_vida_util.delete(0, "end")
        if hasattr(self, 'entry_stock_total'):
            self.entry_stock_total.delete(0, "end")
            self.entry_stock_total.insert(0, "1")
        if hasattr(self, 'entry_stock_bodega'):
            self.entry_stock_bodega.delete(0, "end")
            self.entry_stock_bodega.insert(0, "1")
        self.text_descripcion.delete("1.0", "end")
        self.text_observaciones.delete("1.0", "end")
        self.combo_categoria.set("Maquinaria")
        self.combo_estado.set("Operativa")
        self.combo_responsable.set("Bodega")
        self.foto_path_actual = None
        self.label_imagen_nombre.configure(text="Sin imagen", text_color="gray")
        self.herramienta_editando_id = None

    def cargar_herramientas(self):
        """Carga herramientas en Catálogo con filtrado y muestra stock_total y stock_bodega"""
        for item in self.tabla.get_children():
            self.tabla.delete(item)

        try:
            with db.get_connection() as conn:
                cursor = conn.cursor()
                query = """
                    SELECT h.codigo, h.nombre, h.categoria, f.nombre as finca_nombre,
                           h.stock_total, h.id_trabajador, h.responsable, h.stock_bodega,
                           h.estado, h.id_finca
                    FROM herramienta h
                    LEFT JOIN finca f ON h.id_finca = f.id
                    WHERE 1=1
                """
                params = []

                # Filtro por finca
                if hasattr(self, 'combo_filtro_finca') and self.combo_filtro_finca.get() != "Todas":
                    try:
                        id_finca = int(self.combo_filtro_finca.get().split("-")[0])
                        query += " AND h.id_finca = ?"
                        params.append(id_finca)
                    except Exception:
                        pass

                # Filtro por estado
                if hasattr(self, 'combo_filtro_estado') and self.combo_filtro_estado.get() != "Todos":
                    query += " AND h.estado = ?"
                    params.append(self.combo_filtro_estado.get())

                query += " ORDER BY h.codigo"
                cursor.execute(query, params)

                for row in cursor.fetchall():
                    # row indices after join: 0 codigo,1 nombre,2 categoria,3 finca_nombre,4 stock_total,5 id_trabajador,
                    # 6 responsable,7 stock_bodega,8 estado,9 id_finca
                    # Mostrar Asignada si hay id_trabajador o si hay responsable textual distinto de Bodega
                    responsable_txt = (row[6] or "").strip().lower() if row[6] else ""
                    asignada_por_texto = bool(responsable_txt and responsable_txt != "bodega")
                    asignacion = "Asignada" if (row[5] or asignada_por_texto) else "En Bodega"
                    stock_total = row[4] if row[4] is not None else 1
                    # Si no hay dato de stock_bodega, estimar: 0 si asignada, total si en bodega
                    stock_bodega = row[7] if row[7] is not None else (0 if (row[5] or asignada_por_texto) else stock_total)
                    finca_nombre = row[3] or 'Sin Finca'
                    self.tabla.insert("", "end", values=(
                        row[0], row[1], row[2] or 'N/A', finca_nombre, stock_total, asignacion, stock_bodega, row[8]
                    ))
        except Exception as e:
            messagebox.showerror("Error", f"No se pudieron cargar las herramientas:\n{e}")

    def ver_detalles_herramienta(self):
        """Muestra los detalles de la herramienta seleccionada con su foto"""
        seleccionado = self.tabla.selection()
        if not seleccionado:
            messagebox.showwarning("Atención", "Seleccione una herramienta")
            return
        
        codigo = self.tabla.item(seleccionado[0])["values"][0]
        
        try:
            with db.get_connection() as conn:
                conn.row_factory = sqlite3.Row
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT h.*, f.nombre as finca_nombre
                    FROM herramienta h
                    LEFT JOIN finca f ON h.id_finca = f.id
                    WHERE h.codigo = ?
                """, (codigo,))
                
                row = cursor.fetchone()
                if row:
                    h = dict(row)
                    detalles = f"""
📋 DETALLES DE LA HERRAMIENTA

🔧 Código: {h['codigo']}
📝 Nombre: {h['nombre']}
📦 Categoría: {h.get('categoria') or 'N/A'}
📋 Descripción: {h.get('descripcion') or 'No hay descripción'}

🏷️ IDENTIFICACIÓN
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
• Marca: {h.get('marca') or 'N/A'}
• Modelo: {h.get('modelo') or 'N/A'}
• N° Serie: {h.get('numero_serie') or 'N/A'}

🏞️ UBICACIÓN, ASIGNACIÓN Y STOCK
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
• Finca: {h.get('finca_nombre') or 'N/A'}
• Ubicación: {h.get('ubicacion') or 'N/A'}
• Responsable: {h.get('responsable') or 'No asignado'}
• Estado: {h.get('estado', 'N/A')}
• Stock Total: {h.get('stock_total', 'N/D')}
• Stock en Bodega: {h.get('stock_bodega', 'N/D')}

💰 INFORMACIÓN FINANCIERA
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
• Fecha Adquisición: {h.get('fecha_adquisicion') or 'N/A'}
• Valor: ${h.get('valor_adquisicion', 0):,.2f}
• Vida Útil: {h.get('vida_util_anos') or 'N/A'} años

📝 OBSERVACIONES
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
{h.get('observaciones') or 'No hay observaciones'}
                    """
                    
                    ventana = ctk.CTkToplevel(self)
                    ventana.title(f"Detalles - {h['nombre']}")
                    ventana.geometry("900x700")
                    
                    # Frame principal con dos columnas
                    main_frame = ctk.CTkFrame(ventana)
                    main_frame.pack(fill="both", expand=True, padx=10, pady=10)
                    
                    # Columna izquierda: Detalles de texto
                    text_frame = ctk.CTkFrame(main_frame)
                    text_frame.pack(side="left", fill="both", expand=True, padx=(0, 5))
                    
                    text = ctk.CTkTextbox(text_frame, width=500, height=680)
                    text.pack(padx=5, pady=5, fill="both", expand=True)
                    text.insert("1.0", detalles)
                    text.configure(state="disabled")
                    
                    # Columna derecha: Foto
                    foto_frame = ctk.CTkFrame(main_frame, width=350)
                    foto_frame.pack(side="right", fill="both", padx=(5, 0))
                    foto_frame.pack_propagate(False)
                    
                    # Título de la sección de foto
                    titulo_foto = ctk.CTkLabel(foto_frame, text="📷 FOTOGRAFÍA", 
                                              font=ctk.CTkFont(size=14, weight="bold"))
                    titulo_foto.pack(pady=(10, 5))
                    
                    # Cargar y mostrar la foto
                    foto_path = h.get('foto_path')
                    if foto_path and os.path.exists(foto_path):
                        try:
                            # Cargar imagen
                            img = Image.open(foto_path)
                            
                            # Calcular tamaño manteniendo aspecto (máximo 320x500)
                            max_width = 320
                            max_height = 500
                            img.thumbnail((max_width, max_height), Image.Resampling.LANCZOS)
                            
                            # Crear imagen para CTk
                            ctk_img = ctk.CTkImage(light_image=img, dark_image=img, 
                                                  size=(img.width, img.height))
                            
                            # Mostrar imagen
                            label_img = ctk.CTkLabel(foto_frame, image=ctk_img, text="")
                            label_img.pack(pady=10)
                            
                            # Botón para ver en tamaño completo
                            btn_ampliar = ctk.CTkButton(foto_frame, text="🔍 Ver tamaño completo",
                                                       command=lambda: self._mostrar_imagen_completa(foto_path, h['nombre']))
                            btn_ampliar.pack(pady=5)
                            
                        except Exception as e:
                            # Error al cargar la imagen
                            label_error = ctk.CTkLabel(foto_frame, 
                                                      text="❌ Error al cargar imagen",
                                                      font=ctk.CTkFont(size=12))
                            label_error.pack(pady=20)
                            
                            label_detalle = ctk.CTkLabel(foto_frame, 
                                                        text=f"{str(e)[:50]}...",
                                                        font=ctk.CTkFont(size=10),
                                                        text_color="gray")
                            label_detalle.pack(pady=5)
                    else:
                        # No hay imagen disponible
                        label_sin_img = ctk.CTkLabel(foto_frame, 
                                                    text="📷",
                                                    font=ctk.CTkFont(size=80))
                        label_sin_img.pack(pady=50)
                        
                        label_texto = ctk.CTkLabel(foto_frame, 
                                                  text="Sin imagen disponible",
                                                  font=ctk.CTkFont(size=14),
                                                  text_color="gray")
                        label_texto.pack(pady=10)
                    
        except Exception as e:
            messagebox.showerror("Error", f"No se pudieron cargar los detalles:\n{e}")
    
    def _mostrar_imagen_completa(self, foto_path, nombre_herramienta):
        """Muestra la imagen en una ventana separada a tamaño completo"""
        try:
            ventana_img = ctk.CTkToplevel(self)
            ventana_img.title(f"Imagen - {nombre_herramienta}")
            ventana_img.geometry("800x800")
            
            # Cargar imagen
            img = Image.open(foto_path)
            
            # Redimensionar si es muy grande (máximo 780x780)
            img.thumbnail((780, 780), Image.Resampling.LANCZOS)
            
            # Crear imagen para CTk
            ctk_img = ctk.CTkImage(light_image=img, dark_image=img, 
                                  size=(img.width, img.height))
            
            # Frame scrollable por si la imagen es muy grande
            scroll_frame = ctk.CTkScrollableFrame(ventana_img)
            scroll_frame.pack(fill="both", expand=True, padx=10, pady=10)
            
            # Mostrar imagen
            label_img = ctk.CTkLabel(scroll_frame, image=ctk_img, text="")
            label_img.pack(pady=10)
            
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo abrir la imagen:\n{e}")

    def editar_herramienta(self):
        """Edita la herramienta seleccionada en el catálogo"""
        seleccionado = self.tabla.selection()
        if not seleccionado:
            messagebox.showwarning("Atención", "Seleccione una herramienta para editar")
            return
        codigo = self.tabla.item(seleccionado[0])["values"][0]
        try:
            with db.get_connection() as conn:
                conn.row_factory = sqlite3.Row
                cursor = conn.cursor()
                cursor.execute("SELECT * FROM herramienta WHERE codigo = ?", (codigo,))
                row = cursor.fetchone()
                if not row:
                    messagebox.showerror("Error", "Herramienta no encontrada")
                    return
                h = dict(row)
                self._cargar_herramienta_en_form(h, cursor)
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo cargar la herramienta:\n{e}")

    def _cargar_herramienta_en_form(self, h: dict, cursor):
        """Carga los datos de una herramienta (dict) en el formulario de registro/edición."""
        # Identificador para edición
        self.herramienta_editando_id = h['id']
        # Campos básicos
        self.entry_codigo.delete(0, "end")
        self.entry_codigo.insert(0, h['codigo'])
        self.entry_nombre.delete(0, "end")
        self.entry_nombre.insert(0, h['nombre'])
        self.combo_categoria.set(h.get('categoria') or "Maquinaria")
        # Descripción
        self.text_descripcion.delete("1.0", "end")
        if h.get('descripcion'):
            self.text_descripcion.insert("1.0", h['descripcion'])
        # Marca/Modelo/Serie
        self.entry_marca.delete(0, "end")
        if h.get('marca'):
            self.entry_marca.insert(0, h['marca'])
        self.entry_modelo.delete(0, "end")
        if h.get('modelo'):
            self.entry_modelo.insert(0, h['modelo'])
        self.entry_serie.delete(0, "end")
        if h.get('numero_serie'):
            self.entry_serie.insert(0, h['numero_serie'])
        # Finca
        if h.get('id_finca'):
            cursor.execute("SELECT id, nombre FROM finca WHERE id = ?", (h['id_finca'],))
            finca = cursor.fetchone()
            if finca:
                try:
                    self.combo_finca.set(f"{finca['id']}-{finca['nombre']}")
                except Exception:
                    self.combo_finca.set(f"{finca[0]}-{finca[1]}")
        # Ubicación y estado
        self.entry_ubicacion.delete(0, "end")
        if h.get('ubicacion'):
            self.entry_ubicacion.insert(0, h['ubicacion'])
        self.combo_estado.set(h.get('estado') or "Operativa")
        # Fechas y valores
        self.entry_fecha_adq.delete(0, "end")
        if h.get('fecha_adquisicion'):
            self.entry_fecha_adq.insert(0, h['fecha_adquisicion'])
        self.entry_valor.delete(0, "end")
        if h.get('valor_adquisicion'):
            self.entry_valor.insert(0, str(h['valor_adquisicion']))
        self.entry_vida_util.delete(0, "end")
        if h.get('vida_util_anos'):
            self.entry_vida_util.insert(0, str(h['vida_util_anos']))
        # Responsable (priorizar id_trabajador)
        if h.get('id_trabajador'):
            cursor.execute("SELECT rowid, nombres || ' ' || apellidos, cargo FROM empleado WHERE rowid = ?", (h['id_trabajador'],))
            trab = cursor.fetchone()
            if trab:
                try:
                    self.combo_responsable.set(f"{trab['rowid']}-{trab['nombres || ' ' || apellidos']} ({trab['cargo'] or 'Sin cargo'})")
                except Exception:
                    self.combo_responsable.set(f"{trab[0]}-{trab[1]} ({trab[2] or 'Sin cargo'})")
        elif h.get('responsable'):
            if h['responsable'] == "Bodega":
                self.combo_responsable.set("Bodega")
            else:
                self.combo_responsable.set(h['responsable'])
        # Observaciones
        self.text_observaciones.delete("1.0", "end")
        if h.get('observaciones'):
            self.text_observaciones.insert("1.0", h['observaciones'])
        # Stock
        if hasattr(self, 'entry_stock_total'):
            self.entry_stock_total.delete(0, "end")
            self.entry_stock_total.insert(0, str(h.get('stock_total', 1)))
        if hasattr(self, 'entry_stock_bodega'):
            self.entry_stock_bodega.delete(0, "end")
            valor_bodega = h.get('stock_bodega', 0 if h.get('id_trabajador') else 1)
            self.entry_stock_bodega.insert(0, str(valor_bodega))
        # Foto
        if h.get('foto_path'):
            self.foto_path_actual = h['foto_path']
            try:
                self.label_imagen_nombre.configure(
                    text=os.path.basename(h['foto_path']),
                    text_color="green"
                )
            except Exception:
                self.foto_path_actual = None
                self.label_imagen_nombre.configure(text="Sin imagen", text_color="gray")
        # Cambiar a pestaña de registro
        self.notebook.select(0)
        messagebox.showinfo("Edición", "✏️ Herramienta cargada para edición.\nModifique los campos y presione 'Guardar'")

    def eliminar_herramienta(self):
        """Elimina la herramienta seleccionada en el catálogo"""
        seleccionado = self.tabla.selection()
        if not seleccionado:
            messagebox.showwarning("Atención", "Seleccione una herramienta")
            return
        codigo = self.tabla.item(seleccionado[0])["values"][0]
        if not messagebox.askyesno("Confirmar", f"¿Eliminar la herramienta '{codigo}'?"):
            return
        try:
            with db.get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("DELETE FROM herramienta WHERE codigo = ?", (codigo,))
                conn.commit()
            messagebox.showinfo("Éxito", "Herramienta eliminada correctamente")
            self.cargar_herramientas()
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo eliminar:\n{e}")

    def cargar_herramientas_para_mantenimiento(self):
        """Carga las herramientas en el combo de mantenimiento"""
        try:
            with db.get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT id, codigo, nombre 
                    FROM herramienta 
                    WHERE estado != 'Fuera de Servicio'
                    ORDER BY nombre
                """)
                herramientas = [f"{row[0]}-{row[1]} - {row[2]}" for row in cursor.fetchall()]
                
                if hasattr(self, 'combo_herramienta_mant'):
                    self.combo_herramienta_mant.configure(values=herramientas)
                    if herramientas:
                        self.combo_herramienta_mant.set(herramientas[0])
                        
        except Exception as e:
            messagebox.showerror("Error", f"No se pudieron cargar las herramientas:\n{e}")

    def actualizar_estado_herramienta_mant(self):
        """Actualiza el estado de la herramienta seleccionada desde la pestaña Mantenimientos"""
        if not hasattr(self, 'combo_herramienta_mant') or not self.combo_herramienta_mant.get():
            messagebox.showwarning("Atención", "Seleccione una herramienta en el combo de Mantenimientos")
            return
        nuevo_estado = self.combo_estado_mant.get().strip() if hasattr(self, 'combo_estado_mant') else None
        if not nuevo_estado:
            messagebox.showwarning("Atención", "Seleccione un estado a aplicar")
            return
        try:
            herramienta_id = int(self.combo_herramienta_mant.get().split("-")[0])
            with db.get_connection() as conn:
                cursor = conn.cursor()
                try:
                    cursor.execute("UPDATE herramienta SET estado = ? WHERE id = ?", (nuevo_estado, herramienta_id))
                    conn.commit()
                except Exception as e:
                    # Manejo por si el estado 'En Revisión' no existe en el CHECK constraint
                    if "constraint" in str(e).lower() and nuevo_estado == "En Revisión":
                        messagebox.showerror(
                            "Error",
                            "El estado 'En Revisión' no está habilitado en la base de datos.\n"
                            "Por favor ejecute: aplicar_migraciones_mantenimiento.bat"
                        )
                        return
                    raise
            messagebox.showinfo("Éxito", f"Estado actualizado a: {nuevo_estado}")
            # Refrescar vistas
            self.cargar_herramientas()
            self.cargar_herramientas_para_mantenimiento()
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo actualizar el estado:\n{e}")

    def editar_herramienta_desde_mantenimiento(self):
        """Abre la herramienta seleccionada en la pestaña de mantenimiento para edición"""
        if not self.combo_herramienta_mant.get():
            messagebox.showwarning("Atención", "Seleccione una herramienta en el combo")
            return
        try:
            herramienta_id = int(self.combo_herramienta_mant.get().split("-")[0])
            with db.get_connection() as conn:
                conn.row_factory = sqlite3.Row
                cursor = conn.cursor()
                cursor.execute("SELECT * FROM herramienta WHERE id = ?", (herramienta_id,))
                row = cursor.fetchone()
                if not row:
                    messagebox.showerror("Error", "Herramienta no encontrada")
                    return
                h = dict(row)
                self._cargar_herramienta_en_form(h, cursor)
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo cargar la herramienta:\n{e}")

    def eliminar_herramienta_desde_mantenimiento(self):
        """Elimina el registro de mantenimiento seleccionado en el historial"""
        seleccion = self.tabla_mant.selection()
        if not seleccion:
            messagebox.showwarning("Atención", "Seleccione un mantenimiento del historial para eliminar")
            return
        try:
            # Obtener ID del mantenimiento (primera columna)
            item = self.tabla_mant.item(seleccion[0])
            mant_id = item['values'][0]
            
            if not messagebox.askyesno("Confirmar", 
                "¿Eliminar este registro de mantenimiento?\n\n"
                "La herramienta seguirá en el catálogo, solo se elimina el registro de mantenimiento."):
                return
            
            with db.get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("DELETE FROM mantenimiento_herramienta WHERE id = ?", (mant_id,))
                conn.commit()
            
            messagebox.showinfo("Éxito", "Registro de mantenimiento eliminado correctamente")
            self.cargar_mantenimientos()
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo eliminar el registro:\n{e}")

    def guardar_mantenimiento(self):
        """Guarda un registro de mantenimiento"""
        if not self.combo_herramienta_mant.get():
            messagebox.showwarning("Atención", "Seleccione una herramienta")
            return
        
        try:
            herramienta_id = int(self.combo_herramienta_mant.get().split("-")[0])
            tipo = self.combo_tipo_mant.get()
            fecha = self.entry_fecha_mant.get().strip()
            
            costo = None
            if self.entry_costo_mant.get().strip():
                costo = float(self.entry_costo_mant.get())
            
            with db.get_connection() as conn:
                cursor = conn.cursor()
                
                # Verificar si existe la columna estado_actual (migración 017)
                cursor.execute("PRAGMA table_info(mantenimiento_herramienta)")
                columnas = [col[1] for col in cursor.fetchall()]
                tiene_estado_actual = 'estado_actual' in columnas
                
                if tiene_estado_actual:
                    # Obtener estado actual de la herramienta antes de cambiarlo
                    cursor.execute("SELECT estado FROM herramienta WHERE id = ?", (herramienta_id,))
                    estado_previo = cursor.fetchone()
                    estado_previo = estado_previo[0] if estado_previo else "Operativa"
                    
                    # Determinar nuevo estado según el tipo de mantenimiento
                    if tipo in ["Correctivo", "Calibración"]:
                        nuevo_estado = "En Mantenimiento"
                    else:  # Preventivo, Inspección
                        nuevo_estado = "En Revisión"
                    
                    # Insertar registro de mantenimiento con estado Activo
                    cursor.execute("""
                        INSERT INTO mantenimiento_herramienta (
                            herramienta_id, tipo_mantenimiento, fecha_mantenimiento,
                            descripcion, costo, proveedor_servicio, proximo_mantenimiento,
                            realizado_por, estado_actual, estado_previo_herramienta
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'Activo', ?)
                    """, (
                        herramienta_id, tipo, fecha,
                        self.text_desc_mant.get("1.0", "end-1c").strip() or None,
                        costo,
                        self.entry_proveedor_mant.get().strip() or None,
                        self.entry_proximo_mant.get().strip() or None,
                        self.entry_realizado_por.get().strip() or None,
                        estado_previo
                    ))
                    
                    # Actualizar estado de la herramienta
                    cursor.execute("""
                        UPDATE herramienta 
                        SET estado = ? 
                        WHERE id = ?
                    """, (nuevo_estado, herramienta_id))
                    
                    mensaje_extra = f"\nEstado de herramienta actualizado a: {nuevo_estado}"
                else:
                    # Versión sin estado_actual (antes de migración 017)
                    cursor.execute("""
                        INSERT INTO mantenimiento_herramienta (
                            herramienta_id, tipo_mantenimiento, fecha_mantenimiento,
                            descripcion, costo, proveedor_servicio, proximo_mantenimiento,
                            realizado_por
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """, (
                        herramienta_id, tipo, fecha,
                        self.text_desc_mant.get("1.0", "end-1c").strip() or None,
                        costo,
                        self.entry_proveedor_mant.get().strip() or None,
                        self.entry_proximo_mant.get().strip() or None,
                        self.entry_realizado_por.get().strip() or None
                    ))
                    
                    mensaje_extra = "\n\n⚠️ Para gestión automática de estados, aplique las migraciones:\naplicar_migraciones_mantenimiento.bat"
                
                conn.commit()
            
            messagebox.showinfo("Éxito", f"✅ Mantenimiento registrado correctamente{mensaje_extra}")
            self.cargar_mantenimientos()
            self.cargar_herramientas()  # Actualizar catálogo
            
            # Limpiar campos
            self.text_desc_mant.delete("1.0", "end")
            self.entry_costo_mant.delete(0, "end")
            self.entry_realizado_por.delete(0, "end")
            self.entry_proximo_mant.delete(0, "end")
            self.entry_proveedor_mant.delete(0, "end")
            
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo guardar el mantenimiento:\n{e}")

    def cargar_mantenimientos(self):
        """Carga el historial de mantenimientos"""
        for item in self.tabla_mant.get_children():
            self.tabla_mant.delete(item)
        
        try:
            with db.get_connection() as conn:
                cursor = conn.cursor()
                
                # Verificar si existe la columna estado_actual (migración 017)
                cursor.execute("PRAGMA table_info(mantenimiento_herramienta)")
                columnas = [col[1] for col in cursor.fetchall()]
                tiene_estado_actual = 'estado_actual' in columnas
                
                if tiene_estado_actual:
                    # Versión con estado_actual (después de migración 017)
                    # Solo mostrar mantenimientos activos (los completados ya no aparecen en el historial)
                    cursor.execute("""
                        SELECT m.id, m.fecha_mantenimiento, h.codigo || ' - ' || h.nombre,
                               m.tipo_mantenimiento, 
                               COALESCE(m.estado_actual, 'Activo') as estado,
                               m.costo, m.realizado_por, m.proximo_mantenimiento
                        FROM mantenimiento_herramienta m
                        JOIN herramienta h ON m.herramienta_id = h.id
                        WHERE COALESCE(m.estado_actual, 'Activo') = 'Activo'
                        ORDER BY m.fecha_mantenimiento DESC
                        LIMIT 100
                    """)
                    
                    for row in cursor.fetchall():
                        costo_fmt = f"${row[5]:,.2f}" if row[5] else "N/A"
                        # Formato visual para el estado
                        estado_display = f"🔧 {row[4]}" if row[4] == "Activo" else f"✅ {row[4]}"
                        
                        # Color diferenciado para mantenimientos activos
                        item_id = self.tabla_mant.insert("", "end", values=(
                            row[0], row[1], row[2], row[3], estado_display, costo_fmt,
                            row[6] or "N/A", row[7] or "N/A"
                        ))
                        
                        # Tag para color según estado
                        if row[4] == "Activo":
                            self.tabla_mant.item(item_id, tags=("activo",))
                        else:
                            self.tabla_mant.item(item_id, tags=("completado",))
                    
                    # Configurar colores
                    self.tabla_mant.tag_configure("activo", background="#fff3cd")  # Amarillo claro
                    self.tabla_mant.tag_configure("completado", background="#d4edda")  # Verde claro
                else:
                    # Versión sin estado_actual (antes de migración 017)
                    cursor.execute("""
                        SELECT m.id, m.fecha_mantenimiento, h.codigo || ' - ' || h.nombre,
                               m.tipo_mantenimiento, 
                               m.costo, m.realizado_por, m.proximo_mantenimiento
                        FROM mantenimiento_herramienta m
                        JOIN herramienta h ON m.herramienta_id = h.id
                        ORDER BY m.fecha_mantenimiento DESC
                        LIMIT 100
                    """)
                    
                    for row in cursor.fetchall():
                        costo_fmt = f"${row[4]:,.2f}" if row[4] else "N/A"
                        # Sin columna de estado
                        self.tabla_mant.insert("", "end", values=(
                            row[0], row[1], row[2], row[3], "N/A", costo_fmt,
                            row[5] or "N/A", row[6] or "N/A"
                        ))
                    
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo cargar el historial:\n{e}")
    
    def completar_mantenimiento(self):
        """Marca un mantenimiento como completado y restaura el estado de la herramienta"""
        seleccion = self.tabla_mant.selection()
        if not seleccion:
            messagebox.showwarning("Atención", "Seleccione un mantenimiento de la tabla")
            return
        
        try:
            with db.get_connection() as conn:
                cursor = conn.cursor()
                
                # Verificar si existe la columna estado_actual (migración 017)
                cursor.execute("PRAGMA table_info(mantenimiento_herramienta)")
                columnas = [col[1] for col in cursor.fetchall()]
                tiene_estado_actual = 'estado_actual' in columnas
                
                if not tiene_estado_actual:
                    messagebox.showwarning("Función no disponible", 
                        "Esta función requiere aplicar las migraciones.\n\n"
                        "Ejecute: aplicar_migraciones_mantenimiento.bat")
                    return
            
            # Obtener ID del mantenimiento (primera columna)
            item = self.tabla_mant.item(seleccion[0])
            mant_id = item['values'][0]
            estado_actual = item['values'][4]
            
            # Verificar que esté activo
            if "Completado" in estado_actual or estado_actual == "N/A":
                messagebox.showinfo("Información", "Este mantenimiento ya está completado o no tiene estado")
                return
            
            # Confirmar acción
            if not messagebox.askyesno("Confirmar", 
                "¿Está seguro de que desea marcar este mantenimiento como completado?\n\n"
                "Esto restaurará el estado anterior de la herramienta."):
                return
            
            with db.get_connection() as conn:
                cursor = conn.cursor()
                
                # Obtener datos del mantenimiento
                cursor.execute("""
                    SELECT herramienta_id, estado_previo_herramienta 
                    FROM mantenimiento_herramienta 
                    WHERE id = ?
                """, (mant_id,))
                
                result = cursor.fetchone()
                if not result:
                    messagebox.showerror("Error", "No se encontró el mantenimiento")
                    return
                
                herramienta_id, estado_previo = result
                
                # Marcar mantenimiento como completado
                cursor.execute("""
                    UPDATE mantenimiento_herramienta 
                    SET estado_actual = 'Completado',
                        fecha_completado = date('now')
                    WHERE id = ?
                """, (mant_id,))
                
                # Restaurar estado de la herramienta
                estado_restaurar = estado_previo if estado_previo else "Operativa"
                cursor.execute("""
                    UPDATE herramienta 
                    SET estado = ? 
                    WHERE id = ?
                """, (estado_restaurar, herramienta_id))
                
                conn.commit()
            
            messagebox.showinfo("Éxito", f"✅ Mantenimiento completado\n"
                              f"Estado de herramienta restaurado a: {estado_restaurar}")
            self.cargar_mantenimientos()
            self.cargar_herramientas()  # Actualizar catálogo
            
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo completar el mantenimiento:\n{e}")
    
    def ver_detalles_mantenimiento(self):
        """Muestra los detalles completos de un mantenimiento"""
        seleccion = self.tabla_mant.selection()
        if not seleccion:
            messagebox.showwarning("Atención", "Seleccione un mantenimiento de la tabla")
            return
        
        try:
            # Obtener ID del mantenimiento
            item = self.tabla_mant.item(seleccion[0])
            mant_id = item['values'][0]
            
            with db.get_connection() as conn:
                cursor = conn.cursor()
                
                # Verificar si existe la columna estado_actual (migración 017)
                cursor.execute("PRAGMA table_info(mantenimiento_herramienta)")
                columnas = [col[1] for col in cursor.fetchall()]
                tiene_estado_actual = 'estado_actual' in columnas
                
                cursor.execute("""
                    SELECT m.*, h.codigo || ' - ' || h.nombre as herramienta_nombre
                    FROM mantenimiento_herramienta m
                    JOIN herramienta h ON m.herramienta_id = h.id
                    WHERE m.id = ?
                """, (mant_id,))
                
                row = cursor.fetchone()
                if not row:
                    messagebox.showerror("Error", "No se encontró el mantenimiento")
                    return
                
                # Crear ventana de detalles
                ventana = ctk.CTkToplevel(self)
                ventana.title("Detalles del Mantenimiento")
                ventana.geometry("600x500")
                ventana.transient(self.winfo_toplevel())
                ventana.grab_set()
                
                # Frame principal con scroll
                main_frame = ctk.CTkScrollableFrame(ventana)
                main_frame.pack(fill="both", expand=True, padx=10, pady=10)
                
                ctk.CTkLabel(
                    main_frame,
                    text="📋 Detalles del Mantenimiento",
                    font=("Segoe UI", 16, "bold")
                ).pack(pady=(0, 15))
                
                # Información general (ajustar índices según si tiene estado_actual)
                if tiene_estado_actual:
                    # Con migración 017: id, herramienta_id, tipo, fecha, descripcion, costo, 
                    # proveedor, proximo, realizado_por, observaciones, fecha_creacion, 
                    # estado_actual, estado_previo, fecha_completado, herramienta_nombre
                    info = [
                        ("ID:", row[0]),
                        ("Herramienta:", row[-1]),  # última columna
                        ("Tipo:", row[2]),
                        ("Estado:", f"🔧 {row[11]}" if len(row) > 11 and row[11] == 'Activo' else f"✅ {row[11]}" if len(row) > 11 else "N/A"),
                        ("Fecha:", row[3]),
                        ("Costo:", f"${row[5]:,.2f}" if row[5] else "N/A"),
                        ("Realizado por:", row[8] or "N/A"),
                        ("Proveedor:", row[6] or "N/A"),
                        ("Próximo mant.:", row[7] or "N/A"),
                        ("Estado previo herramienta:", row[12] if len(row) > 12 else "N/A"),
                        ("Fecha completado:", row[13] if len(row) > 13 else "N/A"),
                    ]
                else:
                    # Sin migración 017
                    info = [
                        ("ID:", row[0]),
                        ("Herramienta:", row[-1]),  # última columna
                        ("Tipo:", row[2]),
                        ("Fecha:", row[3]),
                        ("Costo:", f"${row[5]:,.2f}" if row[5] else "N/A"),
                        ("Realizado por:", row[8] or "N/A"),
                        ("Proveedor:", row[6] or "N/A"),
                        ("Próximo mant.:", row[7] or "N/A"),
                    ]
                
                for label, valor in info:
                    row_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
                    row_frame.pack(fill="x", pady=3)
                    
                    ctk.CTkLabel(
                        row_frame,
                        text=label,
                        font=("Segoe UI", 11, "bold"),
                        width=180,
                        anchor="w"
                    ).pack(side="left", padx=5)
                    
                    ctk.CTkLabel(
                        row_frame,
                        text=str(valor),
                        font=("Segoe UI", 11),
                        anchor="w"
                    ).pack(side="left", padx=5, fill="x", expand=True)
                
                # Descripción
                if row[4]:
                    ctk.CTkLabel(
                        main_frame,
                        text="Descripción:",
                        font=("Segoe UI", 11, "bold"),
                        anchor="w"
                    ).pack(fill="x", padx=5, pady=(10, 5))
                    
                    text_desc = ctk.CTkTextbox(main_frame, height=100)
                    text_desc.pack(fill="x", padx=5, pady=5)
                    text_desc.insert("1.0", row[4])
                    text_desc.configure(state="disabled")
                
                # Observaciones
                if row[9]:
                    ctk.CTkLabel(
                        main_frame,
                        text="Observaciones:",
                        font=("Segoe UI", 11, "bold"),
                        anchor="w"
                    ).pack(fill="x", padx=5, pady=(10, 5))
                    
                    text_obs = ctk.CTkTextbox(main_frame, height=80)
                    text_obs.pack(fill="x", padx=5, pady=5)
                    text_obs.insert("1.0", row[9])
                    text_obs.configure(state="disabled")
                
                # Botón cerrar
                ctk.CTkButton(
                    ventana,
                    text="Cerrar",
                    command=ventana.destroy
                ).pack(pady=10)
                
        except Exception as e:
            messagebox.showerror("Error", f"No se pudieron cargar los detalles:\n{e}")
    
    def importar_excel(self):
        """Importa herramientas desde un archivo Excel"""
        archivo = filedialog.askopenfilename(
            title="Seleccionar archivo Excel",
            filetypes=[("Excel", "*.xlsx *.xls"), ("Todos", "*.*")]
        )
        
        if not archivo:
            return
        
        try:
            wb = openpyxl.load_workbook(archivo)
            ws = wb.active
            
            # Verificar encabezados
            headers = [cell.value for cell in ws[1]]
            required = ["codigo", "nombre", "categoria"]
            
            if not all(h in [str(x).lower() if x else "" for x in headers] for h in required):
                messagebox.showerror("Error", 
                    "El archivo debe contener las columnas:\n" +
                    "codigo, nombre, categoria (obligatorias)\n" +
                    "Opcionales: finca, marca, modelo, numero_serie, estado, ubicacion, responsable, fecha_adquisicion,\n" +
                    "valor_adquisicion, vida_util_anos, descripcion, observaciones, stock_total, stock_bodega")
                return
            
            # Mapear índices de columnas
            col_map = {}
            for idx, h in enumerate(headers):
                if h:
                    col_map[str(h).lower().strip()] = idx
            
            importados = 0
            errores = []
            
            with db.get_connection() as conn:
                cursor = conn.cursor()
                
                # Obtener mapeo de fincas con normalización (nombre -> id)
                cursor.execute("SELECT id, nombre FROM finca")
                fincas_map = {}
                for fid, fname in cursor.fetchall():
                    norm = self._normalize_text(fname)
                    fincas_map[norm] = fid
                    # Agregar variante sin prefijo 'finca ' si existe
                    if norm.startswith('finca '):
                        fincas_map[norm[6:]] = fid
                
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    try:
                        codigo = row[col_map.get("codigo")]
                        nombre = row[col_map.get("nombre")]
                        categoria = row[col_map.get("categoria")]
                        
                        if not codigo or not nombre or not categoria:
                            errores.append(f"Fila {row_idx}: Faltan campos obligatorios")
                            continue
                        
                        # Obtener id_finca
                        id_finca = None
                        if "finca" in col_map and row[col_map["finca"]]:
                            finca_nombre_original = str(row[col_map["finca"]]).strip()
                            finca_key = self._normalize_text(finca_nombre_original)
                            id_finca = fincas_map.get(finca_key)
                            # Fallback quitar 'finca ' inicial
                            if id_finca is None and finca_key.startswith('finca '):
                                id_finca = fincas_map.get(finca_key[6:])
                            # Fallback búsqueda parcial si aún no se encuentra
                            if id_finca is None:
                                candidatos = [fid for k, fid in fincas_map.items() if finca_key in k]
                                if len(candidatos) == 1:
                                    id_finca = candidatos[0]
                                elif len(candidatos) > 1:
                                    errores.append(f"Fila {row_idx}: Nombre de finca '{finca_nombre_original}' ambiguo")
                            if id_finca is None:
                                errores.append(f"Fila {row_idx}: Finca '{finca_nombre_original}' no encontrada - verifique nombre")
                        
                        # Obtener otros campos opcionales
                        marca = row[col_map.get("marca")] if "marca" in col_map else None
                        modelo = row[col_map.get("modelo")] if "modelo" in col_map else None
                        numero_serie = row[col_map.get("numero_serie")] if "numero_serie" in col_map else None
                        estado = row[col_map.get("estado", "Operativa")] if "estado" in col_map else "Operativa"
                        ubicacion = row[col_map.get("ubicacion")] if "ubicacion" in col_map else None
                        responsable = row[col_map.get("responsable")] if "responsable" in col_map else None
                        fecha_adq = row[col_map.get("fecha_adquisicion")] if "fecha_adquisicion" in col_map else None
                        valor = None
                        if "valor_adquisicion" in col_map:
                            raw_valor = row[col_map.get("valor_adquisicion")]
                            try:
                                valor = self._parse_valor(raw_valor)
                            except ValueError:
                                errores.append(f"Fila {row_idx}: Valor inválido '{raw_valor}'")
                        vida_util = row[col_map.get("vida_util_anos")] if "vida_util_anos" in col_map else None
                        descripcion = row[col_map.get("descripcion")] if "descripcion" in col_map else None
                        observaciones = row[col_map.get("observaciones")] if "observaciones" in col_map else None
                        stock_total_raw = row[col_map.get("stock_total")] if "stock_total" in col_map else None
                        stock_bodega_raw = row[col_map.get("stock_bodega")] if "stock_bodega" in col_map else None
                        stock_total = self._get_stock_total_validado(stock_total_raw)
                        stock_bodega = self._get_stock_bodega_validado(stock_bodega_raw, stock_total, responsable or "Bodega")
                        
                        # Buscar id_trabajador si hay responsable especificado
                        id_trabajador = None
                        if responsable:
                            responsable_norm = self._normalize_text(str(responsable))
                            cursor.execute("""
                                SELECT rowid FROM empleado 
                                WHERE estado_actual = 'Activo'
                            """)
                            for emp_id, in cursor.fetchall():
                                cursor.execute("""
                                    SELECT nombres, apellidos FROM empleado WHERE rowid = ?
                                """, (emp_id,))
                                nombres, apellidos = cursor.fetchone()
                                nombre_completo = f"{nombres} {apellidos}" if apellidos else nombres
                                if self._normalize_text(nombre_completo) == responsable_norm:
                                    id_trabajador = emp_id
                                    break
                        
                        cursor.execute("""
                            INSERT INTO herramienta (
                                codigo, nombre, categoria, descripcion, marca, modelo, numero_serie,
                                id_finca, ubicacion, estado, fecha_adquisicion, valor_adquisicion,
                                vida_util_anos, responsable, observaciones, stock_total, stock_bodega, id_trabajador
                            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """, (
                            codigo, nombre, categoria, descripcion, marca, modelo, numero_serie,
                            id_finca, ubicacion, estado, fecha_adq, valor, vida_util, responsable, observaciones,
                            stock_total, stock_bodega, id_trabajador
                        ))
                        
                        importados += 1
                        
                    except sqlite3.IntegrityError:
                        errores.append(f"Fila {row_idx}: Código '{codigo}' ya existe")
                    except Exception as e:
                        errores.append(f"Fila {row_idx}: {str(e)}")
                
                conn.commit()
            
            mensaje = f"✅ Importación completada:\n• {importados} herramientas importadas"
            if errores:
                mensaje += f"\n• {len(errores)} errores:\n" + "\n".join(errores[:5])
                if len(errores) > 5:
                    mensaje += f"\n... y {len(errores) - 5} errores más"
            
            messagebox.showinfo("Importación", mensaje)
            self.cargar_herramientas()
            
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo importar el archivo:\n{e}")
    
    def descargar_plantilla_excel(self):
        """Genera y descarga la plantilla Excel de herramientas"""
        try:
            plantillas_dir = os.path.join(os.path.dirname(__file__), '../../plantillas de carga')
            os.makedirs(plantillas_dir, exist_ok=True)
            
            archivo_plantilla = os.path.join(plantillas_dir, 'plantilla_herramientas.xlsx')
            
            # Crear workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Herramientas"
            
            # Encabezados
            headers = [
                "codigo", "nombre", "categoria", "finca", "marca", "modelo", "numero_serie",
                "estado", "ubicacion", "responsable", "fecha_adquisicion", "valor_adquisicion",
                "vida_util_anos", "descripcion", "observaciones", "stock_total", "stock_bodega"
            ]
            ws.append(headers)
            
            # Formatear encabezados
            for cell in ws[1]:
                cell.font = openpyxl.styles.Font(bold=True)
                cell.fill = openpyxl.styles.PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.font = openpyxl.styles.Font(color="FFFFFF", bold=True)
            
            # Agregar filas de ejemplo
            ws.append([
                "HER-001", "Tractor John Deere", "Maquinaria", "Finca El Prado", "John Deere",
                "5075E", "SN123456", "Operativa", "Bodega Principal", "Bodega",
                "2023-01-15", 45000.00, 10, "Tractor para labores agrícolas", "Mantenimiento al día", 1, 1
            ])
            ws.append([
                "HER-002", "Motosierra Husqvarna", "Herramienta Manual", "Finca El León", "Husqvarna",
                "450e", "HS789012", "Operativa", "Bodega Herramientas", "Bodega",
                "2023-03-20", 850.00, 5, "Motosierra para poda y corte", "", 3, 3
            ])
            
            # Ajustar ancho de columnas
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
            
            # Guardar
            wb.save(archivo_plantilla)
            
            messagebox.showinfo("Éxito", 
                f"✅ Plantilla creada exitosamente en:\n{archivo_plantilla}\n\n" +
                "Columnas obligatorias: codigo, nombre, categoria\n" +
                "Opcionales (recomendado): stock_total, stock_bodega\n" +
                "Estados válidos: Operativa, En Mantenimiento, Dañada, Fuera de Servicio\n" +
                "Categorías válidas: Maquinaria, Herramienta Manual, Equipo Medico, Vehiculo, Equipo Oficina, Otro")
            
            # Abrir carpeta
            os.startfile(plantillas_dir)
            
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo crear la plantilla:\n{e}")

    # ---------------- Helpers de validación de stock -----------------
    def _get_stock_total_validado(self, valor):
        """Valida stock_total (int >=1). Si vacío o inválido retorna 1"""
        try:
            if valor is None:
                return 1
            if isinstance(valor, (int, float)):
                v = int(valor)
            else:
                txt = str(valor).strip()
                v = int(txt) if txt else 1
            if v < 1:
                v = 1
            return v
        except Exception:
            return 1

    def _get_stock_bodega_validado(self, valor, stock_total, responsable_actual):
        """Valida stock_bodega (int >=0 <= stock_total). Si asignada a trabajador y vacío -> 0"""
        asignada = responsable_actual and responsable_actual != "Bodega"
        try:
            if valor is None or (isinstance(valor, str) and not valor.strip()):
                return 0 if asignada else stock_total
            if isinstance(valor, (int, float)):
                v = int(valor)
            else:
                v = int(str(valor).strip())
            if v < 0:
                v = 0
            if v > stock_total:
                v = stock_total
            # Heurística: si asignada y stock_bodega == stock_total, disminuir 1 para reflejar uso
            if asignada and v == stock_total:
                v = max(0, stock_total - 1)
            return v
        except Exception:
            return 0 if asignada else stock_total
