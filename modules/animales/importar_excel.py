"""
Módulo para importar animales desde Excel
"""
import customtkinter as ctk
from tkinter import messagebox, filedialog
import sys
import os
import sqlite3
from datetime import datetime

sys.path.append(os.path.join(os.path.dirname(__file__), '../..'))
from database.conexion import db

try:
    import openpyxl
except ImportError:
    messagebox.showerror(
        "Error",
        "openpyxl no está instalado.\n\n"
        "Instale con: pip install openpyxl"
    )
    openpyxl = None


def importar_animales_desde_excel():
    """Importa animales desde un archivo Excel"""
    if openpyxl is None:
        messagebox.showerror(
            "Error",
            "openpyxl no está instalado.\n\n"
            "Instale con: pip install openpyxl\n\n"
            "O ejecute: python crear_plantilla_excel.py"
        )
        return
    
    # Seleccionar archivo
    archivo = filedialog.askopenfilename(
        title="Seleccionar archivo Excel",
        filetypes=[("Excel files", "*.xlsx *.xls"), ("Todos los archivos", "*.*")]
    )
    
    if not archivo:
        return
    
    try:
        # Leer Excel
        wb = openpyxl.load_workbook(archivo)
        ws = wb.active
        
        # Leer encabezados
        headers = [cell.value for cell in ws[1]]
        
        # Mapeo de columnas
        col_map = {}
        for i, header in enumerate(headers, 1):
            if header:
                header_clean = str(header).strip().lower()
                if "código" in header_clean or "codigo" in header_clean:
                    col_map['codigo'] = i
                elif "nombre" in header_clean:
                    col_map['nombre'] = i
                elif "tipo" in header_clean and "ingreso" in header_clean:
                    col_map['tipo_ingreso'] = i
                elif "sexo" in header_clean:
                    col_map['sexo'] = i
                elif "fecha" in header_clean and "nacimiento" in header_clean:
                    col_map['fecha_nacimiento'] = i
                elif "fecha" in header_clean and "compra" in header_clean:
                    col_map['fecha_compra'] = i
                elif "finca" in header_clean:
                    col_map['finca'] = i
                elif "raza" in header_clean:
                    col_map['raza'] = i
                elif "potrero" in header_clean:
                    col_map['potrero'] = i
                elif "peso" in header_clean and "nacimiento" in header_clean:
                    col_map['peso_nacimiento'] = i
                elif "peso" in header_clean and "compra" in header_clean:
                    col_map['peso_compra'] = i
                elif "precio" in header_clean:
                    col_map['precio_compra'] = i
                elif "salud" in header_clean:
                    col_map['salud'] = i
                elif "color" in header_clean:
                    col_map['color'] = i
                elif "hierro" in header_clean:
                    col_map['hierro'] = i
                elif "comentario" in header_clean:
                    col_map['comentarios'] = i
        
        # Validar campos obligatorios
        if 'codigo' not in col_map or 'tipo_ingreso' not in col_map or 'sexo' not in col_map or 'finca' not in col_map:
            messagebox.showerror(
                "Error",
                "Faltan columnas obligatorias en el Excel.\n\n"
                "Requeridas: Código, Tipo Ingreso, Sexo, Finca"
            )
            return
        
        # Procesar filas
        animales_importados = 0
        animales_errores = []
        
        with db.get_connection() as conn:
            cursor = conn.cursor()
            
            # Obtener IDs de referencia
            cursor.execute("SELECT id, nombre FROM finca WHERE estado = 'Activa' OR estado = 'Activo'")
            fincas_dict = {row[1]: row[0] for row in cursor.fetchall()}
            
            cursor.execute("SELECT id, nombre FROM raza WHERE estado = 'Activa' OR estado = 'Activo'")
            razas_dict = {row[1]: row[0] for row in cursor.fetchall()}
            
            cursor.execute("SELECT id, nombre FROM potrero WHERE estado = 'Activo' OR estado = 'Activa'")
            potreros_dict = {row[1]: row[0] for row in cursor.fetchall()}
            
            # Procesar cada fila (empezando desde la 2)
            for row_num in range(2, ws.max_row + 1):
                try:
                    # Leer valores
                    codigo = str(ws.cell(row=row_num, column=col_map['codigo']).value or "").strip()
                    if not codigo:
                        continue  # Fila vacía
                    
                    nombre = str(ws.cell(row=row_num, column=col_map.get('nombre', 0)).value or "").strip() or None
                    tipo_ingreso = str(ws.cell(row=row_num, column=col_map['tipo_ingreso']).value or "").strip()
                    sexo = str(ws.cell(row=row_num, column=col_map['sexo']).value or "").strip()
                    finca_nombre = str(ws.cell(row=row_num, column=col_map['finca']).value or "").strip()
                    
                    # Validaciones básicas
                    if tipo_ingreso not in ["Nacimiento", "Compra"]:
                        animales_errores.append(f"{codigo}: Tipo ingreso inválido")
                        continue
                    
                    if sexo not in ["Macho", "Hembra"]:
                        animales_errores.append(f"{codigo}: Sexo inválido")
                        continue
                    
                    if finca_nombre not in fincas_dict:
                        animales_errores.append(f"{codigo}: Finca '{finca_nombre}' no encontrada")
                        continue
                    
                    id_finca = fincas_dict[finca_nombre]
                    
                    # Obtener otros valores
                    fecha_nacimiento = None
                    if 'fecha_nacimiento' in col_map:
                        fecha_val = ws.cell(row=row_num, column=col_map['fecha_nacimiento']).value
                        if fecha_val:
                            if hasattr(fecha_val, 'strftime'):
                                fecha_nacimiento = fecha_val.strftime("%Y-%m-%d")
                            else:
                                fecha_nacimiento = str(fecha_val).strip()
                    
                    fecha_compra = None
                    if 'fecha_compra' in col_map:
                        fecha_val = ws.cell(row=row_num, column=col_map['fecha_compra']).value
                        if fecha_val:
                            if hasattr(fecha_val, 'strftime'):
                                fecha_compra = fecha_val.strftime("%Y-%m-%d")
                            else:
                                fecha_compra = str(fecha_val).strip()
                    
                    # Raza: almacenar nombre (la BD actual tiene columna 'raza' como TEXT)
                    raza_nombre = None
                    if 'raza' in col_map:
                        raza_nombre = str(ws.cell(row=row_num, column=col_map['raza']).value or "").strip() or None
                    
                    # Potrero
                    id_potrero = None
                    if 'potrero' in col_map:
                        potrero_nombre = str(ws.cell(row=row_num, column=col_map['potrero']).value or "").strip()
                        if potrero_nombre and potrero_nombre in potreros_dict:
                            id_potrero = potreros_dict[potrero_nombre]
                    
                    # Pesos
                    peso_nacimiento = None
                    if 'peso_nacimiento' in col_map:
                        peso_val = ws.cell(row=row_num, column=col_map['peso_nacimiento']).value
                        if peso_val:
                            try:
                                peso_nacimiento = float(peso_val)
                            except:
                                pass
                    
                    peso_compra = None
                    if 'peso_compra' in col_map:
                        peso_val = ws.cell(row=row_num, column=col_map['peso_compra']).value
                        if peso_val:
                            try:
                                peso_compra = float(peso_val)
                            except:
                                pass
                    
                    # Precio
                    precio_compra = None
                    if 'precio_compra' in col_map:
                        precio_val = ws.cell(row=row_num, column=col_map['precio_compra']).value
                        if precio_val:
                            try:
                                precio_compra = float(precio_val)
                            except:
                                pass
                    
                    # Salud
                    salud = "Sano"
                    if 'salud' in col_map:
                        salud_val = str(ws.cell(row=row_num, column=col_map['salud']).value or "").strip()
                        if salud_val:
                            salud = salud_val
                    
                    # Otros campos
                    color = None
                    if 'color' in col_map:
                        color_val = str(ws.cell(row=row_num, column=col_map['color']).value or "").strip()
                        if color_val:
                            color = color_val
                    
                    hierro = None
                    if 'hierro' in col_map:
                        hierro_val = str(ws.cell(row=row_num, column=col_map['hierro']).value or "").strip()
                        if hierro_val:
                            hierro = hierro_val
                    
                    comentarios = None
                    if 'comentarios' in col_map:
                        comentarios_val = str(ws.cell(row=row_num, column=col_map['comentarios']).value or "").strip()
                        if comentarios_val:
                            comentarios = comentarios_val
                    
                    # Insertar en BD
                    cursor.execute("""
                        INSERT INTO animal (
                            id_finca, codigo, nombre, tipo_ingreso, sexo, raza, id_potrero,
                            fecha_nacimiento, fecha_compra, peso_nacimiento, peso_compra,
                            precio_compra, salud, estado, inventariado, color, hierro, comentarios,
                            fecha_registro
                        )
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'Activo', 0, ?, ?, ?, ?)
                    """, (
                        id_finca, codigo, nombre, tipo_ingreso, sexo, raza_nombre, id_potrero,
                        fecha_nacimiento, fecha_compra, peso_nacimiento, peso_compra,
                        precio_compra, salud, color, hierro, comentarios,
                        datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    ))
                    
                    animales_importados += 1
                    
                except sqlite3.IntegrityError:
                    animales_errores.append(f"{codigo}: Código duplicado")
                except Exception as e:
                    animales_errores.append(f"{codigo}: {str(e)}")
            
            conn.commit()
        
        # Mostrar resultado
        mensaje = f"✅ Importación completada\n\n"
        mensaje += f"Animales importados: {animales_importados}\n"
        if animales_errores:
            mensaje += f"Errores: {len(animales_errores)}\n\n"
            mensaje += "Primeros errores:\n"
            for error in animales_errores[:10]:
                mensaje += f"  • {error}\n"
            if len(animales_errores) > 10:
                mensaje += f"  ... y {len(animales_errores) - 10} más"
        else:
            mensaje += "✅ Todos los animales se importaron correctamente"
        
        messagebox.showinfo("Importación", mensaje)
        
    except Exception as e:
        messagebox.showerror("Error", f"Error al importar:\n{e}")

