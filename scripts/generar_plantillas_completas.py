"""
Script para generar todas las plantillas de importación Excel
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path

def crear_plantilla_excel(nombre_archivo, columnas, descripcion=""):
    """Crea una plantilla Excel con las columnas especificadas"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Datos"
    
    # Estilo para encabezados
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Escribir encabezados
    for col_num, columna in enumerate(columnas, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = columna
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Ajustar ancho de columnas
    for col_num in range(1, len(columnas) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 20
    
    # Guardar archivo
    ruta = Path("plantillas de carga") / nombre_archivo
    wb.save(ruta)
    # Evitar caracteres Unicode que pueden fallar en cp1252 (✓ / ✅)
    print(f"Creada: {nombre_archivo}")

def main():
    """Genera todas las plantillas"""
    print("Generando plantillas de importación Excel...\n")
    
    # 1. Animales
    crear_plantilla_excel(
        "plantilla_animales.xlsx",
        ["Código", "Nombre", "Tipo Ingreso", "Sexo", "Fecha Nacimiento", "Fecha Compra",
         "Finca", "Raza", "Potrero", "Lote", "Sector", "Grupo", "Peso Nacimiento", "Peso Compra", 
         "Precio Compra", "Vendedor", "Procedencia", "Salud", "Estado", "Inventariado", 
         "Color", "Hierro", "Condición Corporal", "Comentario"]
    )
    
    # 2. Tipo Explotación
    crear_plantilla_excel(
        "plantilla_tipo_explotacion.xlsx",
        ["Código", "Descripción", "Categoría", "Comentario"]
    )
    
    # 3. Finca
    crear_plantilla_excel(
        "plantilla_finca.xlsx",
        ["Código", "Nombre", "Propietario", "Ubicación", "Área Hectáreas", 
         "Teléfono", "Email", "Descripción", "Comentario"]
    )
    
    # 4. Sector
    crear_plantilla_excel(
        "plantilla_sector.xlsx",
        ["Código", "Nombre", "Descripción", "Comentario"]
    )
    
    # 5. Lote
    crear_plantilla_excel(
        "plantilla_lote.xlsx",
        ["Código", "Nombre", "Descripción", "Criterio", "Comentario"]
    )
    
    # 6. Condición Corporal
    crear_plantilla_excel(
        "plantilla_condicion_corporal.xlsx",
        ["Condición Corporal", "Rango Inferior", "Rango Superior", 
         "Descripción", "Recomendación", "Comentario"]
    )
    
    # 7. Razas
    crear_plantilla_excel(
        "plantilla_razas.xlsx",
        ["Código", "Nombre", "Tipo Ganado", "Especie", "Descripción", "Comentario"]
    )
    
    # 8. Potreros
    crear_plantilla_excel(
        "plantilla_potreros.xlsx",
        ["Código", "Finca", "Nombre", "Sector", "Área Hectáreas", 
         "Capacidad Máxima", "Tipo Pasto", "Descripción", "Estado", "Comentario"]
    )
    
    # 9. Empleados
    crear_plantilla_excel(
        "plantilla_empleados.xlsx",
        ["Cédula", "Nombre", "Cargo", "Teléfono", "Email", 
         "Fecha Ingreso", "Salario", "Dirección", "Comentario"]
    )
    
    # 10. Proveedores
    crear_plantilla_excel(
        "plantilla_proveedores.xlsx",
        ["Código", "Nombre", "NIT", "Teléfono", "Email", 
         "Dirección", "Ciudad", "Contacto", "Comentario"]
    )
    
    # 11. Procedencia
    crear_plantilla_excel(
        "plantilla_procedencia.xlsx",
        ["Código", "Nombre", "Tipo", "Ubicación", "Contacto", 
         "Teléfono", "Descripción", "Comentario"]
    )
    
    # 12. Motivos Venta
    crear_plantilla_excel(
        "plantilla_motivos_venta.xlsx",
        ["Código", "Descripción", "Comentario"]
    )
    
    # 13. Destino Venta
    crear_plantilla_excel(
        "plantilla_destino_venta.xlsx",
        ["Código", "Nombre", "Tipo", "NIT", "Dirección", 
         "Teléfono", "Email", "Comentario"]
    )
    
    # 14. Diagnósticos
    crear_plantilla_excel(
        "plantilla_diagnosticos.xlsx",
        ["Código", "Nombre", "Categoría", "Descripción", 
         "Tratamiento Sugerido", "Comentario"]
    )
    
    # 15. Causa Muerte
    crear_plantilla_excel(
        "plantilla_causa_muerte.xlsx",
        ["Código", "Descripción", "Tipo Causa", "Comentario"]
    )
    
    # 16. Calidad Animal
    crear_plantilla_excel(
        "plantilla_calidad_animal.xlsx",
        ["Código", "Descripción", "Comentario"]
    )
    
    # 17. Animales - Carga Masiva Completa
    crear_plantilla_excel(
        "plantilla_animales_masiva.xlsx",
        ["codigo", "nombre", "tipo_ingreso", "sexo", "fecha_nacimiento", "fecha_compra",
         "finca", "raza", "madre_codigo", "padre_codigo", "potrero", "lote", "sector", "grupo",
         "peso_nacimiento", "peso_compra", "precio_compra", "procedencia", "vendedor",
         "color", "hierro", "numero_hierros", "composicion_racial", "condicion_corporal", 
         "calidad", "salud", "estado", "inventariado", "comentarios"]
    )
    
    # 18. Tratamientos
    crear_plantilla_excel(
        "plantilla_tratamientos.xlsx",
        ["animal_codigo", "fecha", "tipo_tratamiento", "producto", "dosis", 
         "veterinario", "comentario", "fecha_proxima"]
    )
    
    # 19. Reproducción - Servicios
    crear_plantilla_excel(
        "plantilla_servicios.xlsx",
        ["animal_codigo_hembra", "fecha_cubricion", "tipo_cubricion", 
         "toro_semen", "observaciones"]
    )
    
    # 20. Ventas
    crear_plantilla_excel(
        "plantilla_ventas.xlsx",
        ["animal_codigo", "fecha_venta", "precio_total", "motivo_venta", 
         "destino_venta", "observaciones"]
    )
    
    # 21. Salud - Diagnósticos
    crear_plantilla_excel(
        "plantilla_diagnosticos_eventos.xlsx",
        ["animal_codigo", "fecha", "tipo", "diagnostico_detalle", 
         "severidad", "estado", "observaciones"]
    )
    
    # 22. Producción Leche
    crear_plantilla_excel(
        "plantilla_produccion_leche.xlsx",
        ["animal_codigo", "fecha", "cantidad_litros", "numero_ordeno", 
         "calidad", "observaciones"]
    )
    
    # 23. Pesajes
    crear_plantilla_excel(
        "plantilla_pesajes.xlsx",
        ["animal_codigo", "fecha_pesaje", "peso_kg", "condicion_corporal", 
         "observaciones"]
    )
    
    print(f"\nSe generaron 23 plantillas exitosamente en 'plantillas de carga/'")

if __name__ == "__main__":
    main()
