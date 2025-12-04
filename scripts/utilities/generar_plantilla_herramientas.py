"""
Script para generar la plantilla Excel de herramientas
"""
import openpyxl
from openpyxl.styles import Font, PatternFill
import os

def generar_plantilla_herramientas():
    """Genera la plantilla Excel para importación de herramientas"""
    
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Herramientas"
    
    # Encabezados
    headers = [
        "codigo", "nombre", "categoria", "finca", "marca", "modelo", "numero_serie",
        "estado", "ubicacion", "responsable", "fecha_adquisicion", "valor_adquisicion",
        "vida_util_anos", "descripcion", "observaciones"
    ]
    ws.append(headers)
    
    # Formatear encabezados
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    # Agregar filas de ejemplo con datos realistas
    ejemplos = [
        [
            "HER-001", "Tractor John Deere 5075E", "Maquinaria", "Finca El Prado", 
            "John Deere", "5075E", "SN123456789", "Operativa", "Bodega Principal", 
            "Bodega", "2023-01-15", 45000.00, 10, 
            "Tractor de 75HP para labores agrícolas pesadas", 
            "Requiere mantenimiento cada 100 horas de uso"
        ],
        [
            "HER-002", "Motosierra Husqvarna 450e", "Herramienta Manual", "Finca El León",
            "Husqvarna", "450e", "HS789012345", "Operativa", "Bodega Herramientas",
            "Bodega", "2023-03-20", 850.00, 5,
            "Motosierra profesional para poda y corte de árboles",
            "Revisar cadena y filtro mensualmente"
        ],
        [
            "HER-003", "Fumigadora Stihl SR 450", "Equipo Medico", "Finca El Prado",
            "Stihl", "SR 450", "ST456789012", "En Mantenimiento", "Taller",
            "Bodega", "2022-08-10", 1200.00, 7,
            "Fumigadora de mochila para aplicación de tratamientos",
            "Actualmente en revisión técnica"
        ],
        [
            "HER-004", "Camioneta Toyota Hilux", "Vehiculo", "Finca El León",
            "Toyota", "Hilux 2020", "VIN1234567890ABC", "Operativa", "Parqueadero",
            "Bodega", "2020-05-15", 28000.00, 15,
            "Camioneta para transporte de personal y materiales",
            "Placas: ABC123 - SOAT vigente hasta 2025-05"
        ],
        [
            "HER-005", "Ordeñadora Mecánica DeLaval", "Maquinaria", "Finca El Prado",
            "DeLaval", "MMU-300", "DV987654321", "Operativa", "Sala de Ordeño",
            "Bodega", "2021-11-20", 8500.00, 12,
            "Sistema de ordeño mecánico para 4 vacas simultáneas",
            "Limpieza diaria obligatoria"
        ]
    ]
    
    for ejemplo in ejemplos:
        ws.append(ejemplo)
    
    # Ajustar ancho de columnas automáticamente
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Agregar hoja de instrucciones
    ws_inst = wb.create_sheet("Instrucciones")
    
    instrucciones = [
        ["INSTRUCCIONES DE USO - PLANTILLA HERRAMIENTAS"],
        [""],
        ["COLUMNAS OBLIGATORIAS (*)"],
        ["• codigo: Código único de la herramienta (ej: HER-001)"],
        ["• nombre: Nombre descriptivo de la herramienta"],
        ["• categoria: Debe ser una de las siguientes:"],
        ["  - Maquinaria"],
        ["  - Herramienta Manual"],
        ["  - Equipo Medico"],
        ["  - Vehiculo"],
        ["  - Equipo Oficina"],
        ["  - Otro"],
        [""],
        ["COLUMNAS OPCIONALES"],
        ["• finca: Nombre exacto de la finca (debe existir en el sistema)"],
        ["• marca: Marca del equipo o herramienta"],
        ["• modelo: Modelo específico"],
        ["• numero_serie: Número de serie del fabricante"],
        ["• estado: Operativa / En Mantenimiento / Dañada / Fuera de Servicio"],
        ["• ubicacion: Ubicación física (ej: Bodega Principal, Potrero 1, etc.)"],
        ["• responsable: Nombre del trabajador asignado o 'Bodega'"],
        ["• fecha_adquisicion: Formato AAAA-MM-DD (ej: 2023-01-15)"],
        ["• valor_adquisicion: Valor numérico sin símbolos (ej: 45000.00)"],
        ["• vida_util_anos: Años de vida útil estimada (número entero)"],
        ["• descripcion: Descripción detallada de la herramienta"],
        ["• observaciones: Notas adicionales"],
        [""],
        ["NOTAS IMPORTANTES"],
        ["1. No modifique los nombres de las columnas del encabezado"],
        ["2. Los códigos deben ser únicos (no duplicados)"],
        ["3. Las fincas deben existir previamente en el sistema"],
        ["4. El estado por defecto es 'Operativa' si no se especifica"],
        ["5. Si no asigna responsable, quedará como 'Bodega'"],
        ["6. Elimine las filas de ejemplo antes de importar sus datos"],
        [""],
        ["FORMATO DE FECHAS"],
        ["• Use siempre el formato: AAAA-MM-DD"],
        ["• Ejemplo válido: 2023-01-15"],
        ["• Ejemplo inválido: 15/01/2023 o 01-15-2023"],
    ]
    
    for row in instrucciones:
        ws_inst.append(row)
    
    # Formatear título de instrucciones
    ws_inst['A1'].font = Font(bold=True, size=14, color="1F4E78")
    ws_inst['A1'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    # Ajustar ancho de columna
    ws_inst.column_dimensions['A'].width = 80
    
    # Guardar archivo
    plantillas_dir = os.path.join(os.path.dirname(__file__), '../../plantillas de carga')
    os.makedirs(plantillas_dir, exist_ok=True)
    
    archivo_salida = os.path.join(plantillas_dir, 'plantilla_herramientas.xlsx')
    wb.save(archivo_salida)
    
    print(f"✅ Plantilla creada exitosamente en: {archivo_salida}")
    return archivo_salida

if __name__ == "__main__":
    try:
        generar_plantilla_herramientas()
    except Exception as e:
        print(f"❌ Error: {e}")
