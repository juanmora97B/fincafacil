"""
Script para crear la plantilla Excel para importar animales
Ejecuta: python crear_plantilla_excel.py
"""
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Instalando openpyxl...")
    import subprocess
    import sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

def crear_plantilla():
    """Crea la plantilla Excel para importar animales"""
    
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Animales"
    
    # Encabezados con formato
    headers = [
        "Código*",
        "Nombre",
        "Tipo Ingreso*",
        "Sexo*",
        "Fecha Nacimiento",
        "Fecha Compra",
        "Finca*",
        "Raza",
        "Potrero",
        "Peso Nacimiento (kg)",
        "Peso Compra (kg)",
        "Precio Compra",
        "Salud",
        "Color",
        "Hierro",
        "Comentarios"
    ]
    
    # Estilo para encabezados
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    # Escribir encabezados
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Ajustar ancho de columnas
    column_widths = [15, 20, 15, 10, 18, 18, 20, 20, 20, 15, 15, 15, 12, 15, 15, 30]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Agregar fila de ejemplo
    ejemplo = [
        "001",
        "Toro 1",
        "Nacimiento",
        "Macho",
        "2023-01-15",
        "",
        "Finca El Prado",
        "Holstein",
        "Potrero 1",
        "35.5",
        "",
        "",
        "Sano",
        "Negro y Blanco",
        "HIERRO-001",
        "Animal de buena genética"
    ]
    
    for col, value in enumerate(ejemplo, 1):
        cell = ws.cell(row=2, column=col, value=value)
        if col == 3:  # Tipo Ingreso
            cell.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        if col == 4:  # Sexo
            cell.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    
    # Agregar hoja de instrucciones
    ws2 = wb.create_sheet("INSTRUCCIONES")
    
    instrucciones = """
INSTRUCCIONES PARA IMPORTAR ANIMALES
====================================

CAMPOS OBLIGATORIOS (*):
- Código: Código único del animal (ej: 001, VACA-001)
- Tipo Ingreso: "Nacimiento" o "Compra"
- Sexo: "Macho" o "Hembra"
- Finca: Nombre de la finca (debe existir en el sistema)

CAMPOS OPCIONALES:
- Nombre: Nombre del animal
- Fecha Nacimiento: Formato YYYY-MM-DD (ej: 2023-01-15)
- Fecha Compra: Solo si Tipo Ingreso = "Compra"
- Raza: Nombre de la raza (debe existir en el sistema)
- Potrero: Nombre del potrero (debe existir en el sistema)
- Peso: En kilogramos
- Precio Compra: Solo si es compra
- Salud: "Sano", "Enfermo", etc. (por defecto: "Sano")
- Color: Color del animal
- Hierro: Número o código del hierro
- Comentarios: Notas adicionales

IMPORTANTE:
1. No modifique la primera fila (encabezados)
2. Complete al menos los campos obligatorios
3. Use los nombres exactos de Finca, Raza y Potrero que están en el sistema
4. Para Tipo Ingreso use exactamente: "Nacimiento" o "Compra"
5. Para Sexo use exactamente: "Macho" o "Hembra"
6. Las fechas deben estar en formato YYYY-MM-DD

EJEMPLOS:
- Tipo Ingreso: "Nacimiento" o "Compra"
- Sexo: "Macho" o "Hembra"
- Fecha: "2023-01-15" (año-mes-día)

Después de completar el Excel:
1. Guarde el archivo
2. Vaya al módulo de Animales > Registro Animal
3. Use el botón "Importar desde Excel"
    """
    
    ws2['A1'] = instrucciones
    ws2['A1'].alignment = Alignment(wrap_text=True, vertical="top")
    ws2.column_dimensions['A'].width = 80
    ws2.row_dimensions[1].height = 400
    
    # Guardar
    nombre_archivo = "plantilla_importar_animales.xlsx"
    wb.save(nombre_archivo)
    print(f"✅ Plantilla creada: {nombre_archivo}")
    print(f"📝 Ubicación: {os.path.abspath(nombre_archivo)}")
    print()
    print("Ahora puede completar la plantilla con sus 70 animales.")
    print("Recuerde usar los nombres exactos de Finca, Raza y Potrero del sistema.")

if __name__ == "__main__":
    import os
    crear_plantilla()
    input("\nPresiona Enter para salir...")

