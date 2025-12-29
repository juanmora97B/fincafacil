"""
╔══════════════════════════════════════════════════════════════════════════╗
║                     EXPORTADOR EXCEL - FASE 3                            ║
╚══════════════════════════════════════════════════════════════════════════╝

Exporta reportes a formato Excel con formato profesional.
Requiere: openpyxl (pip install openpyxl)
"""

from datetime import datetime
from pathlib import Path
from typing import Dict, Any
import logging
from src.core.audit_service import log_event


class ExcelExporter:
    """Exportador de reportes a formato Excel"""
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
        self._Workbook: Any = None
        self._Font: Any = None
        self._verificar_dependencias()
    
    def _verificar_dependencias(self):
        """Verifica que openpyxl esté instalado"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
            self._Workbook = Workbook
            self._Font = Font
        except ImportError:
            self.logger.warning(
                "openpyxl no está instalado. "
                "Instalar con: pip install openpyxl"
            )
            self._Workbook = None
            self._Font = None
    
    def exportar(self, reporte: Dict[str, Any], ruta_salida: str) -> str:
        """
        Exporta un reporte a Excel.
        
        Args:
            reporte: Diccionario con datos del reporte
            ruta_salida: Ruta del archivo Excel de salida
        
        Returns:
            Ruta del archivo generado
        
        Raises:
            Exception: Si hay error en la exportación o falta openpyxl
        """
        if not self._Workbook or not self._Font:
            raise ImportError(
                "openpyxl no está instalado. "
                "Instalar con: pip install openpyxl"
            )
        
        try:
            # Asegurar que la ruta existe
            Path(ruta_salida).parent.mkdir(parents=True, exist_ok=True)
            
            tipo = reporte.get('tipo', 'desconocido')
            self.logger.info(f"Exportando reporte '{tipo}' a Excel: {ruta_salida}")
            
            # Crear workbook
            wb = self._Workbook()
            ws = wb.active
            ws.title = tipo.capitalize()
            
            # Delegar según tipo
            if tipo == 'animales':
                self._exportar_animales(ws, reporte)
            elif tipo == 'reproduccion':
                self._exportar_reproduccion(ws, reporte)
            elif tipo == 'produccion':
                self._exportar_produccion(ws, reporte)
            elif tipo == 'finanzas':
                self._exportar_finanzas(ws, reporte)
            elif tipo == 'completo':
                self._exportar_completo(wb, reporte)
            else:
                raise ValueError(f"Tipo de reporte no soportado: {tipo}")
            
            # Guardar archivo
            wb.save(ruta_salida)
            
            self.logger.info(f"✓ Excel exportado correctamente: {ruta_salida}")
            try:
                log_event(usuario=None, modulo="export", accion="EXPORTAR", entidad=tipo, resultado="OK", mensaje=f"Excel -> {ruta_salida}")
            except Exception:
                pass
            return ruta_salida
        
        except Exception as e:
            self.logger.error(f"Error exportando Excel: {e}", exc_info=True)
            try:
                tipo = reporte.get('tipo', 'desconocido')
                log_event(usuario=None, modulo="export", accion="EXPORTAR", entidad=tipo, resultado="ERROR", mensaje=str(e))
            except Exception:
                pass
            raise

    def exportar_analitico(self, datos: Dict[str, Any], destino: Path) -> Path:
        if not self._Workbook or not self._Font:
            raise ImportError("openpyxl no está instalado. Instalar con: pip install openpyxl")
        destino.parent.mkdir(parents=True, exist_ok=True)
        wb = self._Workbook()
        ws = wb.active
        ws.title = "Analytics"
        ws["A1"] = "Analytics Export"
        ws["A1"].font = self._Font(bold=True)
        row = 3
        fk = datos.get('kpis_financieros') or {}
        if fk:
            ws[f"A{row}"] = "KPIs Financieros"; ws[f"A{row}"].font = self._Font(bold=True); row += 1
            for label, key in [
                ("Ingresos totales", "ingresos_totales"),
                ("Ingresos animales", "ingresos_animales"),
                ("Ingresos leche", "ingresos_leche"),
                ("Costos totales", "costos_totales"),
                ("Nómina", "costos_nomina"),
                ("Insumos", "costos_insumos"),
                ("Margen bruto", "margen_bruto"),
                ("Rentabilidad mensual (%)", "rentabilidad_mensual"),
            ]:
                ws[f"A{row}"] = label; ws[f"B{row}"] = fk.get(key); row += 1
            row += 1
        pk = datos.get('kpis_productivos') or {}
        if pk:
            ws[f"A{row}"] = "KPIs Productivos"; ws[f"A{row}"].font = self._Font(bold=True); row += 1
            for label, key in [
                ("Producción diaria (L)", "produccion_diaria"),
                ("Producción mensual (L)", "produccion_mensual"),
                ("Promedio por animal (L)", "produccion_promedio_por_animal"),
                ("Tasa gestación (%)", "tasa_gestacion"),
                ("Intervalo partos (días)", "intervalo_promedio_partos_dias"),
                ("Mortalidad (n)", "mortalidad"),
            ]:
                ws[f"A{row}"] = label; ws[f"B{row}"] = pk.get(key); row += 1
            row += 1
        tendencias = datos.get('tendencias') or []
        if tendencias:
            ws[f"A{row}"] = "Tendencias"; ws[f"A{row}"].font = self._Font(bold=True); row += 1
            for serie in tendencias:
                ws[f"A{row}"] = serie.get('nombre', 'serie'); row += 1
                ws[f"A{row}"] = "Periodo"; ws[f"B{row}"] = "Valor"; row += 1
                for p in serie.get('puntos', []):
                    ws[f"A{row}"] = p.get('periodo'); ws[f"B{row}"] = p.get('valor'); row += 1
                row += 1
        insights = datos.get('insights') or []
        if insights:
            ws[f"A{row}"] = "Insights"; ws[f"A{row}"].font = self._Font(bold=True); row += 1
            ws[f"A{row}"] = "Nivel"; ws[f"B{row}"] = "Categoria"; ws[f"C{row}"] = "Mensaje"; ws[f"D{row}"] = "Recomendacion"; row += 1
            for ins in insights:
                ws[f"A{row}"] = ins.get('nivel'); ws[f"B{row}"] = ins.get('categoria'); ws[f"C{row}"] = ins.get('mensaje'); ws[f"D{row}"] = ins.get('recomendacion'); row += 1
        wb.save(str(destino))
        try:
            log_event(usuario=None, modulo="analytics", accion="EXPORT_ANALYTICS", entidad=destino.name, resultado="OK")
        except Exception:
            pass
        return destino
    
    def _exportar_animales(self, ws, reporte: Dict[str, Any]):
        """Exporta reporte de animales a worksheet"""
        # Encabezado
        ws['A1'] = 'REPORTE DE ANIMALES - FINCAFÁCIL'
        ws['A1'].font = self._Font(bold=True, size=14)
        
        ws['A2'] = f"Generado: {reporte['generado_en']}"
        ws['A3'] = f"Período: {reporte['periodo']['inicio']} → {reporte['periodo']['fin']}"
        
        row = 5
        
        # Inventario actual
        ws[f'A{row}'] = 'INVENTARIO ACTUAL'
        ws[f'A{row}'].font = self._Font(bold=True)
        row += 1
        
        ws[f'A{row}'] = 'Estado'
        ws[f'B{row}'] = 'Cantidad'
        ws[f'C{row}'] = 'Machos'
        ws[f'D{row}'] = 'Hembras'
        row += 1
        
        inventario = reporte['datos']['inventario_actual']
        for estado, info in inventario['por_estado'].items():
            ws[f'A{row}'] = estado
            ws[f'B{row}'] = info['cantidad']
            ws[f'C{row}'] = info['machos']
            ws[f'D{row}'] = info['hembras']
            row += 1
        
        row += 1
        ws[f'A{row}'] = f"Gestantes actuales: {inventario['gestantes']}"
        row += 2
        
        # Distribución por raza
        ws[f'A{row}'] = 'DISTRIBUCIÓN POR RAZA'
        ws[f'A{row}'].font = self._Font(bold=True)
        row += 1
        
        ws[f'A{row}'] = 'Raza'
        ws[f'B{row}'] = 'Cantidad'
        row += 1
        
        for item in inventario['por_raza']:
            ws[f'A{row}'] = item['raza']
            ws[f'B{row}'] = item['cantidad']
            row += 1
        
        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
    
    def _exportar_reproduccion(self, ws, reporte: Dict[str, Any]):
        """Exporta reporte de reproducción a worksheet"""
        # Encabezado
        ws['A1'] = 'REPORTE DE REPRODUCCIÓN - FINCAFÁCIL'
        ws['A1'].font = self._Font(bold=True, size=14)
        
        ws['A2'] = f"Generado: {reporte['generado_en']}"
        ws['A3'] = f"Período: {reporte['periodo']['inicio']} → {reporte['periodo']['fin']}"
        
        row = 5
        datos = reporte['datos']
        
        # Servicios
        ws[f'A{row}'] = 'SERVICIOS REALIZADOS'
        ws[f'A{row}'].font = self._Font(bold=True)
        row += 1
        ws[f'A{row}'] = f"Total: {datos['servicios']['total']}"
        row += 2
        
        ws[f'A{row}'] = 'Tipo'
        ws[f'B{row}'] = 'Cantidad'
        row += 1
        
        for item in datos['servicios']['por_tipo']:
            ws[f'A{row}'] = item['tipo']
            ws[f'B{row}'] = item['cantidad']
            row += 1
        
        row += 2
        
        # Indicadores
        ws[f'A{row}'] = 'INDICADORES REPRODUCTIVOS'
        ws[f'A{row}'].font = self._Font(bold=True)
        row += 1
        
        ind = datos['indicadores']
        ws[f'A{row}'] = 'Hembras aptas'
        ws[f'B{row}'] = ind['hembras_aptas']
        row += 1
        ws[f'A{row}'] = 'Gestantes'
        ws[f'B{row}'] = ind['gestantes']
        row += 1
        ws[f'A{row}'] = 'Tasa de preñez (%)'
        ws[f'B{row}'] = ind['tasa_prenez_pct']
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
    
    def _exportar_produccion(self, ws, reporte: Dict[str, Any]):
        """Exporta reporte de producción a worksheet"""
        # Encabezado
        ws['A1'] = 'REPORTE DE PRODUCCIÓN - FINCAFÁCIL'
        ws['A1'].font = self._Font(bold=True, size=14)
        
        ws['A2'] = f"Generado: {reporte['generado_en']}"
        ws['A3'] = f"Período: {reporte['periodo']['inicio']} → {reporte['periodo']['fin']}"
        
        row = 5
        datos = reporte['datos']
        prod = datos['produccion_periodo']
        
        # Resumen
        ws[f'A{row}'] = 'RESUMEN DE PRODUCCIÓN'
        ws[f'A{row}'].font = self._Font(bold=True)
        row += 1
        
        ws[f'A{row}'] = 'Litros totales'
        ws[f'B{row}'] = prod['litros_totales']
        row += 1
        ws[f'A{row}'] = 'Días de producción'
        ws[f'B{row}'] = prod['dias_produccion']
        row += 1
        ws[f'A{row}'] = 'Vacas productivas'
        ws[f'B{row}'] = prod['vacas_productivas']
        row += 2
        
        # Promedios
        prom = datos['promedios']
        ws[f'A{row}'] = 'Promedio litros/día'
        ws[f'B{row}'] = prom['promedio_litros_dia']
        row += 1
        ws[f'A{row}'] = 'Promedio litros/vaca'
        ws[f'B{row}'] = prom['promedio_litros_vaca']
        row += 2
        
        # Por animal
        ws[f'A{row}'] = 'PRODUCCIÓN POR ANIMAL (TOP 20)'
        ws[f'A{row}'].font = self._Font(bold=True)
        row += 1
        
        ws[f'A{row}'] = 'Código'
        ws[f'B{row}'] = 'Nombre'
        ws[f'C{row}'] = 'Litros Totales'
        ws[f'D{row}'] = 'Días'
        ws[f'E{row}'] = 'Promedio Diario'
        row += 1
        
        for item in datos['por_animal']:
            ws[f'A{row}'] = item['codigo']
            ws[f'B{row}'] = item['nombre']
            ws[f'C{row}'] = item['litros_totales']
            ws[f'D{row}'] = item['dias_registrados']
            ws[f'E{row}'] = item['promedio_diario']
            row += 1
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 18
    
    def _exportar_finanzas(self, ws, reporte: Dict[str, Any]):
        """Exporta reporte financiero a worksheet"""
        # Encabezado
        ws['A1'] = 'REPORTE FINANCIERO - FINCAFÁCIL'
        ws['A1'].font = self._Font(bold=True, size=14)
        
        ws['A2'] = f"Generado: {reporte['generado_en']}"
        ws['A3'] = f"Período: {reporte['periodo']['inicio']} → {reporte['periodo']['fin']}"
        
        row = 5
        datos = reporte['datos']
        
        # Ingresos
        ws[f'A{row}'] = 'INGRESOS'
        ws[f'A{row}'].font = self._Font(bold=True)
        row += 1
        
        ws[f'A{row}'] = 'Ventas de animales'
        ws[f'B{row}'] = datos['ingresos']['ventas_animales']
        ws[f'B{row}'].number_format = '$#,##0'
        row += 1
        
        ws[f'A{row}'] = 'Ventas de leche'
        ws[f'B{row}'] = datos['ingresos']['ventas_leche']
        ws[f'B{row}'].number_format = '$#,##0'
        row += 1
        
        ws[f'A{row}'] = 'TOTAL INGRESOS'
        ws[f'A{row}'].font = self._Font(bold=True)
        ws[f'B{row}'] = datos['ingresos']['total']
        ws[f'B{row}'].number_format = '$#,##0'
        ws[f'B{row}'].font = self._Font(bold=True)
        row += 2
        
        # Costos
        ws[f'A{row}'] = 'COSTOS'
        ws[f'A{row}'].font = self._Font(bold=True)
        row += 1
        
        ws[f'A{row}'] = 'Nómina'
        ws[f'B{row}'] = datos['costos']['nomina']
        ws[f'B{row}'].number_format = '$#,##0'
        row += 1
        
        ws[f'A{row}'] = 'Tratamientos'
        ws[f'B{row}'] = datos['costos']['tratamientos']
        ws[f'B{row}'].number_format = '$#,##0'
        row += 1
        
        ws[f'A{row}'] = 'Insumos'
        ws[f'B{row}'] = datos['costos']['insumos']
        ws[f'B{row}'].number_format = '$#,##0'
        row += 1
        
        ws[f'A{row}'] = 'TOTAL COSTOS'
        ws[f'A{row}'].font = self._Font(bold=True)
        ws[f'B{row}'] = datos['costos']['total']
        ws[f'B{row}'].number_format = '$#,##0'
        ws[f'B{row}'].font = self._Font(bold=True)
        row += 2
        
        # Margen
        ws[f'A{row}'] = 'MARGEN BRUTO'
        ws[f'A{row}'].font = self._Font(bold=True)
        row += 1
        
        ws[f'A{row}'] = 'Valor'
        ws[f'B{row}'] = datos['margen']['valor']
        ws[f'B{row}'].number_format = '$#,##0'
        row += 1
        
        ws[f'A{row}'] = 'Porcentaje'
        ws[f'B{row}'] = datos['margen']['porcentaje'] / 100
        ws[f'B{row}'].number_format = '0.0%'
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
    
    def _exportar_completo(self, wb, reporte: Dict[str, Any]):
        """Exporta reporte completo (múltiples hojas)"""
        # Eliminar hoja por defecto
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
        
        secciones = reporte.get('secciones', {})
        
        for nombre_seccion, datos_seccion in secciones.items():
            # Crear hoja para cada sección
            ws = wb.create_sheet(title=nombre_seccion.capitalize())
            
            # Crear mini-reporte
            mini_reporte = {
                'tipo': nombre_seccion,
                'periodo': reporte['periodo'],
                'generado_en': reporte['generado_en'],
                'datos': datos_seccion.get('datos', {}),
                'totales': datos_seccion.get('totales', {})
            }
            
            # Exportar a la hoja
            if nombre_seccion == 'animales':
                self._exportar_animales(ws, mini_reporte)
            elif nombre_seccion == 'reproduccion':
                self._exportar_reproduccion(ws, mini_reporte)
            elif nombre_seccion == 'produccion':
                self._exportar_produccion(ws, mini_reporte)
            elif nombre_seccion == 'finanzas':
                self._exportar_finanzas(ws, mini_reporte)


# Singleton
excel_exporter = ExcelExporter()
