"""
Utilidad para exportar datos a diferentes formatos (PDF, Excel, CSV)
Centraliza la funcionalidad de exportación para todos los módulos
"""
from typing import List, Dict, Any, Optional
from pathlib import Path
import logging

logger = logging.getLogger("ExportadorDatos")


class ExportadorDatos:
    """Clase para exportar datos a diferentes formatos"""
    
    @staticmethod
    def exportar_a_excel(datos: List[Dict[str, Any]], columnas: List[str], 
                         ruta_archivo: Path, nombre_hoja: str = "Datos") -> bool:
        """
        Exporta datos a formato Excel
        
        Args:
            datos: Lista de diccionarios con los datos
            columnas: Lista de nombres de columnas
            ruta_archivo: Ruta donde guardar el archivo
            nombre_hoja: Nombre de la hoja de Excel
            
        Returns:
            True si se exportó exitosamente
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # Crear workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = nombre_hoja
            
            # Estilo de encabezados
            header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Escribir encabezados
            for col_idx, columna in enumerate(columnas, start=1):
                celda = ws.cell(row=1, column=col_idx, value=columna)
                celda.fill = header_fill
                celda.font = header_font
                celda.alignment = header_alignment
            
            # Escribir datos
            for row_idx, fila in enumerate(datos, start=2):
                for col_idx, columna in enumerate(columnas, start=1):
                    valor = fila.get(columna, "")
                    ws.cell(row=row_idx, column=col_idx, value=valor)
            
            # Ajustar ancho de columnas
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Guardar archivo
            wb.save(ruta_archivo)
            logger.info(f"Datos exportados a Excel: {ruta_archivo}")
            return True
            
        except ImportError:
            logger.error("Módulo openpyxl no está instalado. Instale con: pip install openpyxl")
            return False
        except Exception as e:
            logger.error(f"Error exportando a Excel: {e}")
            return False
    
    @staticmethod
    def exportar_a_csv(datos: List[Dict[str, Any]], columnas: List[str], 
                       ruta_archivo: Path) -> bool:
        """
        Exporta datos a formato CSV
        
        Args:
            datos: Lista de diccionarios con los datos
            columnas: Lista de nombres de columnas
            ruta_archivo: Ruta donde guardar el archivo
            
        Returns:
            True si se exportó exitosamente
        """
        try:
            import csv
            
            with open(ruta_archivo, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.DictWriter(csvfile, fieldnames=columnas)
                writer.writeheader()
                writer.writerows(datos)
            
            logger.info(f"Datos exportados a CSV: {ruta_archivo}")
            return True
            
        except Exception as e:
            logger.error(f"Error exportando a CSV: {e}")
            return False
    
    @staticmethod
    def exportar_a_pdf(datos: List[Dict[str, Any]], columnas: List[str], 
                       ruta_archivo: Path, titulo: str = "Reporte") -> bool:
        """
        Exporta datos a formato PDF
        
        Args:
            datos: Lista de diccionarios con los datos
            columnas: Lista de nombres de columnas
            ruta_archivo: Ruta donde guardar el archivo
            titulo: Título del documento
            
        Returns:
            True si se exportó exitosamente
        """
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from datetime import datetime
            
            # Crear documento
            doc = SimpleDocTemplate(str(ruta_archivo), pagesize=landscape(letter))
            elements = []
            
            # Estilos
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                textColor=colors.HexColor('#1976D2'),
                spaceAfter=30,
                alignment=1  # Centrado
            )
            
            # Título
            elements.append(Paragraph(titulo, title_style))
            elements.append(Paragraph(f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles['Normal']))
            elements.append(Spacer(1, 0.3*inch))
            
            # Preparar datos para la tabla
            tabla_datos = [columnas]
            for fila in datos:
                tabla_datos.append([str(fila.get(col, "")) for col in columnas])
            
            # Crear tabla
            tabla = Table(tabla_datos)
            
            # Estilo de tabla
            tabla.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1976D2')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
            ]))
            
            elements.append(tabla)
            
            # Construir PDF
            doc.build(elements)
            logger.info(f"Datos exportados a PDF: {ruta_archivo}")
            return True
            
        except ImportError:
            logger.error("Módulo reportlab no está instalado. Instale con: pip install reportlab")
            return False
        except Exception as e:
            logger.error(f"Error exportando a PDF: {e}")
            return False
    
    @staticmethod
    def exportar_datos(datos: List[Dict[str, Any]], columnas: List[str], 
                      ruta_archivo: Path, formato: str = "excel", 
                      titulo: Optional[str] = None) -> bool:
        """
        Exporta datos al formato especificado
        
        Args:
            datos: Lista de diccionarios con los datos
            columnas: Lista de nombres de columnas
            ruta_archivo: Ruta donde guardar el archivo
            formato: Formato de exportación (excel, csv, pdf)
            titulo: Título para el documento (usado en PDF)
            
        Returns:
            True si se exportó exitosamente
        """
        formato = formato.lower()
        
        if formato == "excel":
            return ExportadorDatos.exportar_a_excel(datos, columnas, ruta_archivo)
        elif formato == "csv":
            return ExportadorDatos.exportar_a_csv(datos, columnas, ruta_archivo)
        elif formato == "pdf":
            titulo_doc = titulo or "Reporte"
            return ExportadorDatos.exportar_a_pdf(datos, columnas, ruta_archivo, titulo_doc)
        else:
            logger.error(f"Formato no soportado: {formato}")
            return False


def exportar_tabla_treeview(treeview, ruta_archivo: Path, formato: str = "excel", 
                           titulo: Optional[str] = None) -> bool:
    """
    Exporta el contenido de un Treeview directamente
    
    Args:
        treeview: Widget Treeview de tkinter
        ruta_archivo: Ruta donde guardar el archivo
        formato: Formato de exportación
        titulo: Título para el documento
        
    Returns:
        True si se exportó exitosamente
    """
    try:
        # Obtener columnas
        columnas = list(treeview["columns"])
        
        # Obtener datos
        datos = []
        for item in treeview.get_children():
            valores = treeview.item(item)["values"]
            fila = {col: val for col, val in zip(columnas, valores)}
            datos.append(fila)
        
        # Exportar
        return ExportadorDatos.exportar_datos(datos, columnas, ruta_archivo, formato, titulo)
        
    except Exception as e:
        logger.error(f"Error exportando Treeview: {e}")
        return False
