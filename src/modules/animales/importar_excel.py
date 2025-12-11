"""
Módulo para importar animales desde Excel
"""
import customtkinter as ctk
from tkinter import messagebox, filedialog
import sys
import os
import sqlite3
from datetime import datetime

sys.path.append(os.path.join(os.path.dirname(__file__), '../..'))
from database import db
from modules.utils.constants_ui import truncate

try:
    import openpyxl
except ImportError:
    messagebox.showerror(
        "Error",
        "openpyxl no está instalado.\n\n"
        "Instale con: pip install openpyxl"
    )
    openpyxl = None


def importar_animales_desde_excel():
    """Importa animales desde un archivo Excel"""
    if openpyxl is None:
        messagebox.showerror(
            "Error",
            "openpyxl no está instalado.\n\n"
            "Instale con: pip install openpyxl\n\n"
            "O ejecute: python crear_plantilla_excel.py"
        )
        return
    
    # Seleccionar archivo
    archivo = filedialog.askopenfilename(
        title="Seleccionar archivo Excel",
        filetypes=[("Excel files", "*.xlsx *.xls"), ("Todos los archivos", "*.*")]
    )
    
    if not archivo:
        return
    
    try:
        # Leer Excel
        wb = openpyxl.load_workbook(archivo)
        ws = wb.active
        
        # Leer encabezados
        headers = [cell.value for cell in ws[1]]
        
        # Mapeo de columnas
        col_map = {}
        for i, header in enumerate(headers, 1):
            if header:
                header_clean = str(header).strip().lower()
                if "código" in header_clean or "codigo" in header_clean:
                    col_map['codigo'] = i
                elif "nombre" in header_clean:
                    col_map['nombre'] = i
                elif "tipo" in header_clean and "ingreso" in header_clean:
                    col_map['tipo_ingreso'] = i
                elif "sexo" in header_clean:
                    col_map['sexo'] = i
                elif "fecha" in header_clean and "nacimiento" in header_clean:
                    col_map['fecha_nacimiento'] = i
                elif "fecha" in header_clean and "compra" in header_clean:
                    col_map['fecha_compra'] = i
                elif "finca" in header_clean:
                    col_map['finca'] = i
                elif "raza" in header_clean:
                    col_map['raza'] = i
                elif "potrero" in header_clean:
                    col_map['potrero'] = i
                elif "lote" in header_clean:
                    col_map['lote'] = i
                elif "sector" in header_clean:
                    col_map['sector'] = i
                elif "grupo" in header_clean:
                    col_map['grupo'] = i
                elif "condición" in header_clean and "corporal" in header_clean:
                    col_map['condicion_corporal'] = i
                elif "condicion" in header_clean and "corporal" in header_clean:
                    col_map['condicion_corporal'] = i
                elif "peso" in header_clean and "nacimiento" in header_clean:
                    col_map['peso_nacimiento'] = i
                elif "peso" in header_clean and "compra" in header_clean:
                    col_map['peso_compra'] = i
                elif "precio" in header_clean:
                    col_map['precio_compra'] = i
                elif "salud" in header_clean:
                    col_map['salud'] = i
                elif "color" in header_clean:
                    col_map['color'] = i
                elif "hierro" in header_clean:
                    col_map['hierro'] = i
                elif "comentario" in header_clean:
                    col_map['comentarios'] = i
        
        # Validar campos obligatorios
        if 'codigo' not in col_map or 'tipo_ingreso' not in col_map or 'sexo' not in col_map or 'finca' not in col_map:
            messagebox.showerror(
                "Error",
                "Faltan columnas obligatorias en el Excel.\n\n"
                "Requeridas: Código, Tipo Ingreso, Sexo, Finca"
            )
            return
        
        # Procesar filas
        animales_importados = 0
        animales_errores = []
        
        # Importar helpers para búsquedas case-insensitive
        from modules.utils.database_helpers import (
            obtener_diccionario_normalizado,
            buscar_finca_id,
            buscar_raza_id,
            buscar_potrero_id,
            buscar_lote_id,
            buscar_sector_id
        )
        
        with db.get_connection() as conn:
            cursor = conn.cursor()
            
            # Obtener IDs de referencia usando helpers case-insensitive
            fincas_dict = obtener_diccionario_normalizado(
                cursor, "finca", condicion="estado = 'Activa' OR estado = 'Activo'"
            )
            
            razas_dict = obtener_diccionario_normalizado(
                cursor, "raza", condicion="estado = 'Activa' OR estado = 'Activo'"
            )
            
            potreros_dict = obtener_diccionario_normalizado(
                cursor, "potrero", condicion="estado = 'Activo' OR estado = 'Activa'"
            )
            
            lotes_dict = obtener_diccionario_normalizado(
                cursor, "lote", condicion="estado = 'Activo'"
            )
            
            sectores_dict = obtener_diccionario_normalizado(
                cursor, "sector", condicion="estado = 'Activo'"
            )
            
            # Procesar cada fila (empezando desde la 2)
            for row_num in range(2, ws.max_row + 1):
                try:
                    # Leer valores
                    codigo = truncate(str(ws.cell(row=row_num, column=col_map['codigo']).value or "").strip().upper(), "codigo_animal")
                    if not codigo:
                        continue  # Fila vacía
                    
                    nombre = str(ws.cell(row=row_num, column=col_map.get('nombre', 0)).value or "").strip() or None
                    # Normalización de tipo_ingreso y sexo
                    tipo_ingreso_raw = str(ws.cell(row=row_num, column=col_map['tipo_ingreso']).value or "").strip().lower()
                    mapping_tipo = {
                        'nacimiento': 'Nacimiento', 'nac': 'Nacimiento', 'nacido': 'Nacimiento',
                        'compra': 'Compra', 'comprado': 'Compra', 'comp': 'Compra'
                    }
                    tipo_ingreso = mapping_tipo.get(tipo_ingreso_raw)
                    sexo_raw = str(ws.cell(row=row_num, column=col_map['sexo']).value or "").strip().lower()
                    mapping_sexo = {'macho': 'Macho', 'hembra': 'Hembra', 'm': 'Macho', 'h': 'Hembra'}
                    sexo = mapping_sexo.get(sexo_raw)
                    from modules.utils.database_helpers import normalizar_texto
                    finca_nombre = normalizar_texto(ws.cell(row=row_num, column=col_map['finca']).value)
                    
                    # Validaciones básicas
                    if not tipo_ingreso:
                        animales_errores.append(f"{codigo}: Tipo ingreso inválido ('{tipo_ingreso_raw}')")
                        continue
                    
                    if not sexo:
                        animales_errores.append(f"{codigo}: Sexo inválido ('{sexo_raw}')")
                        continue
                    
                    if finca_nombre not in fincas_dict:
                        animales_errores.append(f"{codigo}: Finca '{finca_nombre}' no encontrada")
                        continue
                    
                    id_finca = fincas_dict[finca_nombre]
                    
                    # Obtener otros valores
                    fecha_nacimiento = None
                    if 'fecha_nacimiento' in col_map:
                        fecha_val = ws.cell(row=row_num, column=col_map['fecha_nacimiento']).value
                        if fecha_val:
                            if hasattr(fecha_val, 'strftime'):
                                fecha_nacimiento = fecha_val.strftime("%Y-%m-%d")
                            else:
                                fecha_nacimiento = str(fecha_val).strip()
                    
                    fecha_compra = None
                    if 'fecha_compra' in col_map:
                        fecha_val = ws.cell(row=row_num, column=col_map['fecha_compra']).value
                        if fecha_val:
                            if hasattr(fecha_val, 'strftime'):
                                fecha_compra = fecha_val.strftime("%Y-%m-%d")
                            else:
                                fecha_compra = str(fecha_val).strip()
                    
                    # Raza: ahora normalizado a raza_id. Tomar nombre y mapear a id.
                    raza_id = None
                    if 'raza' in col_map:
                        raza_nombre_tmp = str(ws.cell(row=row_num, column=col_map['raza']).value or "").strip().lower() or None
                        if raza_nombre_tmp:
                            raza_id = razas_dict.get(raza_nombre_tmp)
                    
                    # Potrero
                    id_potrero = None
                    if 'potrero' in col_map:
                        potrero_nombre = str(ws.cell(row=row_num, column=col_map['potrero']).value or "").strip().lower()
                        if potrero_nombre and potrero_nombre in potreros_dict:
                            id_potrero = potreros_dict[potrero_nombre]
                    
                    # Lote
                    lote_id = None
                    if 'lote' in col_map:
                        lote_nombre = str(ws.cell(row=row_num, column=col_map['lote']).value or "").strip().lower()
                        if lote_nombre and lote_nombre in lotes_dict:
                            lote_id = lotes_dict[lote_nombre]
                    
                    # Sector
                    id_sector = None
                    if 'sector' in col_map:
                        sector_nombre = str(ws.cell(row=row_num, column=col_map['sector']).value or "").strip().lower()
                        if sector_nombre and sector_nombre in sectores_dict:
                            id_sector = sectores_dict[sector_nombre]
                    
                    # Grupo
                    grupo_compra = None
                    if 'grupo' in col_map:
                        grupo_val = str(ws.cell(row=row_num, column=col_map['grupo']).value or "").strip()
                        if grupo_val and grupo_val in ["Toros", "Vacas", "Terneros", "Novillos"]:
                            grupo_compra = grupo_val
                    
                    # Condición Corporal
                    condicion_corporal = None
                    if 'condicion_corporal' in col_map:
                        condicion_val = str(ws.cell(row=row_num, column=col_map['condicion_corporal']).value or "").strip()
                        if condicion_val:
                            condicion_corporal = condicion_val
                    
                    # Pesos
                    peso_nacimiento = None
                    if 'peso_nacimiento' in col_map:
                        peso_val = ws.cell(row=row_num, column=col_map['peso_nacimiento']).value
                        if peso_val:
                            try:
                                peso_nacimiento = float(peso_val)
                            except:
                                pass
                    
                    peso_compra = None
                    if 'peso_compra' in col_map:
                        peso_val = ws.cell(row=row_num, column=col_map['peso_compra']).value
                        if peso_val:
                            try:
                                peso_compra = float(peso_val)
                            except:
                                pass
                    
                    # Precio
                    precio_compra = None
                    if 'precio_compra' in col_map:
                        precio_val = ws.cell(row=row_num, column=col_map['precio_compra']).value
                        if precio_val:
                            try:
                                precio_compra = float(precio_val)
                            except:
                                pass
                    
                    # Salud
                    salud = "Sano"
                    if 'salud' in col_map:
                        salud_val = str(ws.cell(row=row_num, column=col_map['salud']).value or "").strip()
                        if salud_val:
                            salud = salud_val
                    
                    # Otros campos
                    color = None
                    if 'color' in col_map:
                        color_val = str(ws.cell(row=row_num, column=col_map['color']).value or "").strip()
                        if color_val:
                            color = color_val
                    
                    hierro = None
                    if 'hierro' in col_map:
                        hierro_val = str(ws.cell(row=row_num, column=col_map['hierro']).value or "").strip()
                        if hierro_val:
                            hierro = hierro_val
                    
                    comentarios = None
                    if 'comentarios' in col_map:
                        comentarios_val = str(ws.cell(row=row_num, column=col_map['comentarios']).value or "").strip()
                        if comentarios_val:
                            comentarios = comentarios_val
                    
                    # Insertar en BD
                    # Validar unicidad de código rápida
                    cursor.execute("SELECT 1 FROM animal WHERE codigo = ?", (codigo,))
                    if cursor.fetchone():
                        animales_errores.append(f"{codigo}: Código duplicado (omitido)")
                        continue

                    cursor.execute("""
                        INSERT INTO animal (
                            id_finca, codigo, nombre, tipo_ingreso, sexo, raza_id, id_potrero,
                            fecha_nacimiento, fecha_compra, peso_nacimiento, peso_compra,
                            precio_compra, salud, estado, inventariado, color, hierro, comentarios,
                            lote_id, id_sector, grupo_compra, condicion_corporal,
                            fecha_registro
                        )
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'Activo', 0, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (
                        id_finca, codigo, nombre, tipo_ingreso, sexo, raza_id, id_potrero,
                        fecha_nacimiento, fecha_compra, peso_nacimiento, peso_compra,
                        precio_compra, salud, color, hierro, comentarios,
                        lote_id, id_sector, grupo_compra, condicion_corporal,
                        datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    ))
                    
                    animales_importados += 1
                    
                except sqlite3.IntegrityError:
                    animales_errores.append(f"{codigo}: Código duplicado")
                except Exception as e:
                    animales_errores.append(f"{codigo}: {str(e)}")
            
            conn.commit()
        
        # Mostrar resultado
        # Escribir log de errores detallado
        try:
            if animales_errores:
                with open("import_animales_errores.txt", "w", encoding="utf-8") as f:
                    f.write("Errores de importación de animales\n")
                    f.write(f"Total errores: {len(animales_errores)}\n\n")
                    for e in animales_errores:
                        f.write(e + "\n")
            with open("import_animales_resumen.txt", "w", encoding="utf-8") as f:
                f.write(f"Importados correctamente: {animales_importados}\n")
                f.write(f"Errores: {len(animales_errores)}\n")
        except Exception:
            pass

        mensaje = f"✅ Importación completada\n\n"
        mensaje += f"Animales importados: {animales_importados}\n"
        if animales_errores:
            mensaje += f"Errores: {len(animales_errores)}\n\n"
            mensaje += "Primeros errores:\n"
            for error in animales_errores[:10]:
                mensaje += f"  • {error}\n"
            if len(animales_errores) > 10:
                mensaje += f"  ... y {len(animales_errores) - 10} más"
            mensaje += "\nSe generó archivo 'import_animales_errores.txt' con el detalle."
        else:
            mensaje += "✅ Todos los animales se importaron correctamente"

        messagebox.showinfo("Importación", mensaje)
        
    except Exception as e:
        messagebox.showerror("Error", f"Error al importar:\n{e}")

